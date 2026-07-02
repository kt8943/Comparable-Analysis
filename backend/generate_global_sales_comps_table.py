#!/usr/bin/env python3
"""
generate_global_sales_comps_table.py
=====================================
Global (non-Singapore) Land & Investment Sales Comparables table generator.
Same visual style as generate_sales_comps_table.py — navy headers, alternating
fills, subject/comp separation — but without SG-specific columns (leasehold,
Bala cap rate) and with dynamic currency/unit column headers.

Output Schema
-------------
  Type | Source | Property | Map Marker | Date
  Sale Price ({cur} {unit}) | Site Area ({area_unit}) | Unit Price ({cur}/{area_unit})
  Buyer | Zoning Type | Location Type | Quality Type | Note
"""

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Country → ISO currency code ───────────────────────────────────────────────

_COUNTRY_CURRENCY: dict[str, str] = {
    "singapore":          "SGD",
    "korea":              "KRW",
    "south korea":        "KRW",
    "republic of korea":  "KRW",
    "japan":              "JPY",
    "china":              "CNY",
    "prc":                "CNY",
    "hong kong":          "HKD",
    "taiwan":             "TWD",
    "australia":          "AUD",
    "new zealand":        "NZD",
    "united states":      "USD",
    "usa":                "USD",
    "us":                 "USD",
    "united kingdom":     "GBP",
    "uk":                 "GBP",
    "germany":            "EUR",
    "france":             "EUR",
    "india":              "INR",
    "indonesia":          "IDR",
    "malaysia":           "MYR",
    "thailand":           "THB",
    "vietnam":            "VND",
    "philippines":        "PHP",
}


def get_deal_currency(subject_cfg: dict) -> str:
    explicit = str(subject_cfg.get("currency", "")).strip()
    if explicit:
        return explicit
    country = subject_cfg.get("country_name", "").lower().strip()
    return _COUNTRY_CURRENCY.get(country, "USD")


def get_deal_price_unit(subject_cfg: dict) -> str:
    """'B' (billions) or 'M' (millions) — from deal config."""
    return str(subject_cfg.get("price_unit", "M")).upper()


def get_deal_area_unit(subject_cfg: dict) -> str:
    """'sqm' or 'SF' — from deal config."""
    return str(subject_cfg.get("area_unit", "sqm")).lower()


# ═══════════════════════════════════════════════════════════════════════════════
# OUTPUT SCHEMA
# ═══════════════════════════════════════════════════════════════════════════════

_SCHEMA_BASE = [
    ("Type",                "type",        "str",   "@",        10),  # hidden
    ("Source",              "source",      "str",   "@",        10),
    ("Property",            "property",    "str",   "@",        36),
    ("Map\nMarker",         "map_marker",  "str",   "@",         8),
    ("Date",                "sale_date",   "str",   "@",        13),
    ("Sale Price\n(-- --)", "price_raw",   "float", "#,##0.0",  18),
    ("Site Area\n(--)",     "site_area",   "int",   "#,##0",    14),
    ("Unit Price\n(--/--)", "unit_price",  "int",   "#,##0",    18),
    ("Buyer",               "buyer",       "str",   "@",        24),
    ("Zoning\nType",        "land_zoning", "str",   "@",        16),
    ("Location\nType",      "location",    "str",   "@",        20),
    ("Quality\nType",       "quality",     "str",   "@",        20),
    ("Note",                "remarks",     "str",   "@",        30),
]


def get_global_sales_schema(subject_cfg: dict) -> list:
    """Return schema with dynamic currency/unit column headers."""
    cur       = get_deal_currency(subject_cfg)
    unit      = get_deal_price_unit(subject_cfg)
    area_unit = get_deal_area_unit(subject_cfg)
    schema    = list(_SCHEMA_BASE)
    for i, entry in enumerate(schema):
        if entry[1] == "price_raw":
            schema[i] = (f"Sale Price\n({cur} {unit})", "price_raw",  "float", "#,##0.0", entry[4])
        elif entry[1] == "site_area":
            schema[i] = (f"Site Area\n({area_unit})",   "site_area",  "int",   "#,##0",   entry[4])
        elif entry[1] == "unit_price":
            schema[i] = (f"Unit Price\n({cur}/{area_unit})", "unit_price", "int", "#,##0", entry[4])
    return schema


# ═══════════════════════════════════════════════════════════════════════════════
# ROW CONVERTERS
# ═══════════════════════════════════════════════════════════════════════════════

_PRICE_SCALE = {"T": 1_000, "M": 1_000_000, "B": 1_000_000_000}


def _calc_unit_price(price_raw, price_unit: str, site_area) -> int | None:
    """Unit Price = Sale Price × scale / Site Area."""
    if not price_raw or not site_area:
        return None
    try:
        scale = _PRICE_SCALE.get(price_unit.upper(), 1_000_000)
        return round(float(price_raw) * scale / float(site_area))
    except (TypeError, ValueError, ZeroDivisionError):
        return None


def _source_label(src: str) -> str:
    mapping = {"excel": "Excel", "pdf": "PDF", "image": "Image", "manual": "Manual"}
    for prefix in ("pdf_", "excel_", "image_"):
        if src.startswith(prefix):
            return src.replace(prefix, prefix[:-1].title() + " ").replace("_", " ").title()
    return mapping.get(src, "")


def _m_to_display(price_m, price_unit: str):
    """Price is already stored in the display unit (no M-normalization for global comps)."""
    if price_m is None:
        return None
    try:
        return float(price_m)
    except (TypeError, ValueError):
        return price_m


def subject_to_row(cfg: dict, subject_cfg: dict) -> dict:
    price_unit = get_deal_price_unit(subject_cfg)
    price_raw  = _m_to_display(cfg.get("price_raw") or cfg.get("price_sgd_m"), price_unit)
    site_area  = cfg.get("site_area") or cfg.get("site_area_raw") or cfg.get("gfa_sf")
    return {
        "type":        "Subject",
        "source":      "",
        "property":    cfg.get("property_name", ""),
        "map_marker":  "★",
        "sale_date":   cfg.get("sale_date", ""),
        "price_raw":   price_raw,
        "site_area":   int(site_area) if site_area else None,
        "unit_price":  None,   # written as Excel formula by _data_row
        "buyer":       cfg.get("buyer", ""),
        "land_zoning": cfg.get("land_zoning", ""),
        "location":    cfg.get("location", ""),
        "quality":     cfg.get("quality", ""),
        "remarks":     cfg.get("remarks", ""),
    }


def comp_to_row(c: dict, subject_cfg: dict) -> dict:
    price_unit = get_deal_price_unit(subject_cfg)
    name       = str(c.get("property_name") or c.get("raw_description", "").split("\n")[0]).strip()
    price_raw  = _m_to_display(c.get("price_raw") or c.get("price_sgd_m"), price_unit)
    site_area  = c.get("site_area") or c.get("site_area_raw") or c.get("gfa_sf")
    return {
        "type":        "Comparable",
        "source":      _source_label(c.get("_source", "")),
        "property":    name,
        "map_marker":  str(c.get("map_marker", "")),
        "sale_date":   str(c.get("sale_date", "")),
        "price_raw":   float(price_raw) if price_raw else None,
        "site_area":   int(site_area) if site_area else None,
        "unit_price":  None,   # written as Excel formula by _data_row
        "buyer":       str(c.get("buyer", "")),
        "land_zoning": str(c.get("land_zoning", "")),
        "location":    str(c.get("location", "")),
        "quality":     str(c.get("quality", "")),
        "remarks":     str(c.get("remarks", "")),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# RENDERING HELPERS  (same palette as SG template)
# ═══════════════════════════════════════════════════════════════════════════════

_NAVY  = "FF1F3864"
_NAVYL = "FF2E4C7E"
_WHITE = "FFFFFFFF"
_LGRAY = "FFF2F2F2"
_NOTE  = "FFEBF3FB"
_DARK  = "FF1A1A1A"
_RED   = "FFC00000"
_AVGBG = "FFD6DCE4"

def _fill(h):  return PatternFill(patternType="solid", fgColor=h)
def _font(color=_WHITE, bold=True, sz=9, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=sz, italic=italic)
_T = Side(style="thin", color="FFBFBFBF")
def _border(): return Border(left=_T, right=_T, top=_T, bottom=_T)
def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _section_header(ws, row: int, label: str, nc: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row=row, column=1, value=label)
    c.fill = _fill(_NAVY); c.font = _font(_WHITE, bold=True, sz=10)
    c.alignment = _align("left", "center", wrap=False); c.border = _border()
    for col in range(2, nc + 1):
        x = ws.cell(row=row, column=col)
        x.fill = _fill(_NAVY); x.border = _border()
    ws.row_dimensions[row].height = 19


def _col_headers(ws, row: int, schema: list):
    for col, (hdr, *_) in enumerate(schema, 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.fill = _fill(_NAVY); c.font = _font(_WHITE, bold=True, sz=8.5)
        c.alignment = _align("center", "center", wrap=True); c.border = _border()
    ws.row_dimensions[row].height = 34


def _data_row(ws, row: int, row_dict: dict, schema: list,
              alt: bool = False, bold: bool = False, is_subject: bool = False,
              price_unit: str = "M"):
    bg = _fill(_LGRAY) if alt else _fill(_WHITE)
    # Pre-compute price and area column letters for the unit-price formula
    _price_col = get_column_letter(next(i + 1 for i, e in enumerate(schema) if e[1] == "price_raw"))
    _area_col  = get_column_letter(next(i + 1 for i, e in enumerate(schema) if e[1] == "site_area"))
    _scale     = _PRICE_SCALE.get(price_unit.upper(), 1_000_000)

    for col, entry in enumerate(schema, 1):
        key, dtype, fmt = entry[1], entry[2], entry[3]
        val     = row_dict.get(key)

        if key == "unit_price":
            display = f"=IFERROR({_price_col}{row}*{_scale}/{_area_col}{row},\"—\")"
            is_numeric = True
        else:
            display  = "—" if val is None or val == "" else val
            is_numeric = dtype in ("int", "float") and isinstance(val, (int, float))

        c = ws.cell(row=row, column=col, value=display)
        c.fill = bg; c.border = _border()
        c.font = _font(_DARK, bold=bold, sz=9)

        if col == 1:
            c.alignment = _align("left", "center", wrap=True)
        elif col == 2:
            c.alignment = _align("center", "center", wrap=False)
        elif col == 3:   # Property
            c.alignment = _align("center", "center", wrap=False)
            c.font = _font(_RED if is_subject else _DARK, bold=True, sz=12)
        elif is_numeric:
            c.alignment = _align("right", "center", wrap=False)
            c.number_format = fmt
        else:
            c.alignment = _align("center", "center", wrap=True)

    ws.row_dimensions[row].height = 48


def _avg_row(ws, row: int, first_r: int, last_r: int, schema: list):
    avg_keys = {"price_raw", "site_area", "unit_price"}
    for col, entry in enumerate(schema, 1):
        key, dtype, fmt = entry[1], entry[2], entry[3]
        c = ws.cell(row=row, column=col)
        c.fill = _fill(_AVGBG); c.border = _border()
        if key == "property":
            c.value = "Average"
            c.font = _font(_DARK, bold=True, sz=9)
            c.alignment = _align("left", "center", wrap=False)
        elif key in avg_keys:
            col_ltr = get_column_letter(col)
            c.value = f'=IFERROR(AVERAGEIF({col_ltr}{first_r}:{col_ltr}{last_r},">0"),"—")'
            c.number_format = fmt
            c.font = _font(_DARK, bold=True, sz=9)
            c.alignment = _align("right", "center", wrap=False)
        else:
            c.value = ""; c.font = _font(_DARK, bold=False, sz=9)
            c.alignment = _align("center", "center", wrap=False)
    ws.row_dimensions[row].height = 20


# ═══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════

def build_workbook(subject_row: dict, comp_rows: list,
                   subject_cfg: dict, output_path: str):
    schema    = get_global_sales_schema(subject_cfg)
    cur       = get_deal_currency(subject_cfg)
    unit      = get_deal_price_unit(subject_cfg)
    area_unit = get_deal_area_unit(subject_cfg)
    nc        = len(schema)
    deal_name = subject_cfg.get("deal_name", subject_cfg.get("property_name", "Deal"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transaction Comparables"
    ws.sheet_view.showGridLines = False

    for i, (*_, w) in enumerate(schema, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions["A"].hidden = True   # Type column hidden

    # ── Title block ──────────────────────────────────────────────────────────
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    t = ws.cell(row=r, column=1,
                value=f"{deal_name.upper()}  —  TRANSACTION COMPARABLES (ASSET SALES)")
    t.fill = _fill(_NAVY); t.font = _font(_WHITE, bold=True, sz=13)
    t.alignment = _align("center", "center", wrap=False)
    ws.row_dimensions[r].height = 24

    r = 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    s = ws.cell(row=r, column=1, value="Confidential — For Discussion Purposes Only")
    s.fill = _fill(_NAVYL); s.font = _font(_WHITE, bold=False, sz=9, italic=True)
    s.alignment = _align("center", "center", wrap=False)
    ws.row_dimensions[r].height = 14

    # ── Table 1: Subject Sale ─────────────────────────────────────────────────
    r = 4
    _section_header(ws, r, "  Subject Sale — Based on Underwriting", nc)
    r = 5; _col_headers(ws, r, schema)
    r = 6; _data_row(ws, r, subject_row, schema, bold=True, is_subject=True, price_unit=unit)

    ws.row_dimensions[7].height = 6

    # ── Table 2: Comparable Sales ─────────────────────────────────────────────
    r = 8
    _section_header(ws, r, "  Comparable Asset Sales / Valuations", nc)
    r = 9; _col_headers(ws, r, schema)
    for i, crow in enumerate(comp_rows):
        r = 10 + i
        _data_row(ws, r, crow, schema, alt=(i % 2 == 1), price_unit=unit)

    if comp_rows:
        _avg_row(ws, 10 + len(comp_rows), 10, 9 + len(comp_rows), schema)

    # ── Notes footer ──────────────────────────────────────────────────────────
    r = 10 + len(comp_rows) + 2
    notes = (
        f"Notes:  (1) ★ = {deal_name}; markers 1–{len(comp_rows)} = comparable transactions.  "
        f"(2) Sale Price in {cur} {unit} as reported in source.  "
        f"(3) Unit Price ({cur}/{area_unit}) = Sale Price × scale ÷ Site Area "
        f"(calculated if not provided in source).  "
        "Source: JLL, Savills, Colliers, CBRE, local market announcements."
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    n = ws.cell(row=r, column=1, value=notes)
    n.fill = _fill(_NOTE); n.font = _font("FF404040", bold=False, sz=8, italic=True)
    n.alignment = _align("left", "top", wrap=True)
    ws.row_dimensions[r].height = 60

    wb.save(output_path)
    print(f"  Saved → {output_path}")
