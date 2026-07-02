#!/usr/bin/env python3
"""
generate_bala_table_excel.py
============================
Reads the official Singapore Bala Table from Input_files/bala table.pdf
(Appendix 2 — Table Showing Leasehold Values as Percentage of Freehold Value,
Singapore Land Authority / SISV) and writes a clean Excel file to
Input_files/bala_table.xlsx.

Run this once on any machine to (re)generate the Excel from the source PDF:

    # Mac / Linux (from project root):
    python3 backend/generate_bala_table_excel.py

    # Windows (from project root):
    python backend/generate_bala_table_excel.py

Requirements:  pypdf  openpyxl
    pip install pypdf openpyxl

Output columns in 'Bala Table' sheet:
    A — Remaining Leasehold (Years)    1 … 99
    B — Leasehold Value (% of Freehold)   3.8 … 96.0
    C — Factor (Decimal)               0.038 … 0.960

The generate_sales_comps_table.py and generate_land_comps_table.py scripts
read column A + B from this file at import time for all Bala calculations.
"""

import re
import sys
from pathlib import Path

# ── Windows UTF-8 fix ─────────────────────────────────────────────────────────
for _stream in (sys.stdout, sys.stderr):
    try:
        if getattr(_stream, "encoding", "utf-8").lower().replace("-", "") != "utf8":
            _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, OSError):
        pass

# ── dependency check ─────────────────────────────────────────────────────────
try:
    import pypdf
except ImportError:
    sys.exit("ERROR: pypdf not installed.  Run:  pip install pypdf")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    sys.exit("ERROR: openpyxl not installed.  Run:  pip install openpyxl")


# ── paths ────────────────────────────────────────────────────────────────────
BASE_DIR  = Path(__file__).parent.parent   # backend/ → project root
PDF_PATH  = BASE_DIR / "Input_files" / "bala table.pdf"
XLSX_PATH = BASE_DIR / "Input_files" / "bala_table.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — Parse PDF
# ─────────────────────────────────────────────────────────────────────────────

def parse_bala_pdf(pdf_path: Path) -> list[tuple[int, float]]:
    """
    Extract (years, pct) pairs from the Bala Table PDF.
    Handles the 3-column layout: years col1 | pct col1 | years col2 | pct col2 ...
    Returns a list sorted by years, e.g. [(1, 3.8), (2, 7.5), ..., (99, 96.0)].
    """
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    reader = pypdf.PdfReader(str(pdf_path))
    full_text = ""
    for page in reader.pages:
        full_text += page.extract_text() + "\n"

    # Find all  <integer>  <decimal>  pairs in the text
    # The PDF lays out 3 column-pairs on a single page; pypdf extracts them
    # row-by-row left-to-right, so the pattern "N  X.X" repeats throughout.
    pairs = re.findall(r'\b(\d{1,2})\s+([\d]+\.[\d]+)\b', full_text)

    result = {}
    for yrs_str, pct_str in pairs:
        yrs = int(yrs_str)
        pct = float(pct_str)
        if 1 <= yrs <= 99 and 0 < pct < 100:
            result[yrs] = pct

    if len(result) < 90:
        raise ValueError(
            f"Only {len(result)} rows extracted from PDF — expected 99.  "
            "Check that the correct PDF is in Input_files/bala table.pdf"
        )

    return sorted(result.items())   # [(1, 3.8), (2, 7.5), ...]


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — Validate key checkpoints
# ─────────────────────────────────────────────────────────────────────────────

CHECKPOINTS = {30: 60.0, 60: 80.0, 99: 96.0}

def validate(data: list[tuple[int, float]]) -> None:
    d = dict(data)
    ok = True
    for yrs, expected in CHECKPOINTS.items():
        actual = d.get(yrs)
        if actual != expected:
            print(f"  WARN  n={yrs}: expected {expected}%, got {actual}%")
            ok = False
    if ok:
        print(f"  Checkpoints OK  (n=30→{d[30]}%  n=60→{d[60]}%  n=99→{d[99]}%)")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Write Excel
# ─────────────────────────────────────────────────────────────────────────────

# ── palette ──────────────────────────────────────────────────────────────────
_NAVY  = "FF1F3864"
_WHITE = "FFFFFFFF"
_DARK  = "FF1A1A1A"
_LGRAY = "FFF2F2F2"
_NOTE  = "FFEBF3FB"

def _fill(h):
    return PatternFill(patternType="solid", fgColor=h)

def _font(color=_WHITE, bold=True, sz=9, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=sz, italic=italic)

_T = Side(style="thin", color="FFBFBFBF")
def _border():
    return Border(left=_T, right=_T, top=_T, bottom=_T)

def _align(h="center"):
    return Alignment(horizontal=h, vertical="center", wrap_text=False)


def write_excel(data: list[tuple[int, float]], out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bala Table"
    ws.sheet_view.showGridLines = False

    # ── Header row ────────────────────────────────────────────────────────────
    headers = [
        "Remaining Leasehold (Years)",
        "Leasehold Value (% of Freehold)",
        "Factor (Decimal)",
    ]
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=hdr)
        c.fill      = _fill(_NAVY)
        c.font      = _font(_WHITE, bold=True, sz=9)
        c.alignment = _align("center")
        c.border    = _border()
    ws.row_dimensions[1].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, (yrs, pct) in enumerate(data, 2):
        bg = _fill(_LGRAY) if i % 2 == 0 else _fill(_WHITE)

        ca = ws.cell(row=i, column=1, value=yrs)
        cb = ws.cell(row=i, column=2, value=pct)
        cc = ws.cell(row=i, column=3, value=round(pct / 100.0, 5))

        for c in (ca, cb, cc):
            c.fill      = bg
            c.font      = _font(_DARK, bold=False, sz=9)
            c.alignment = _align("center")
            c.border    = _border()

        ca.number_format = "0"
        cb.number_format = '0.0"%"'
        cc.number_format = "0.00000"
        ws.row_dimensions[i].height = 15

    # ── Source note ───────────────────────────────────────────────────────────
    note_row = len(data) + 3
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=3)
    n = ws.cell(
        row=note_row, column=1,
        value=(
            "Source: Appendix 2 — Table Showing Leasehold Values as Percentage of Freehold Value "
            "(Singapore Land Authority / SISV).  "
            "Values for n > 99 yrs: linear interpolation between 99 yrs (96%) and freehold (100%).  "
            "Freehold / 999-yr leases: Bala factor = 1.0 (no adjustment)."
        )
    )
    n.fill      = _fill(_NOTE)
    n.font      = _font("FF404040", bold=False, sz=8, italic=True)
    n.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[note_row].height = 48

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print(f"Reading PDF  : {PDF_PATH}")
    data = parse_bala_pdf(PDF_PATH)
    print(f"  Extracted {len(data)} rows")

    print("Validating   :")
    validate(data)

    print(f"Writing Excel: {XLSX_PATH}")
    write_excel(data, XLSX_PATH)

    print(f"\nDone — {XLSX_PATH.name} saved with {len(data)} rows.")
    print("Run this script again any time to regenerate from the PDF.")


if __name__ == "__main__":
    main()
