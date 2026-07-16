#!/usr/bin/env python3
"""
search_online_rent_comps.py
===========================
Searches online for comparable rental transactions near the subject property
using OpenAI's web search capability.

Search strategy (same 3-tier proximity-first ladder as sales comps):
  Tier 1 — Proximity   : within proximity_km of subject
  Tier 2 — City        : within city_km of subject
  Tier 3 — Country     : country-wide, no distance cap

Usage
-----
    python3 search_online_rent_comps.py --config configs/my_deal.json
    python3 search_online_rent_comps.py --config configs/my_deal.json --map
    python3 search_online_rent_comps.py --config configs/my_deal.json --map --refresh

Output
------
    output/<DealName>/Online_Rent_Comps_<DealName>.xlsx
    output/<DealName>/Online_Rent_Comps_<DealName>_map.png   (if --map)
    output/<DealName>/Online_Rent_Comps_<DealName>_search_cache.json
"""

import sys as _sys
from pathlib import Path as _Path
_sys.path.insert(0, str(_Path(__file__).parent))
# ── Windows UTF-8 fix ─────────────────────────────────────────────────────────
for _stream in (_sys.stdout, _sys.stderr):
    try:
        if getattr(_stream, "encoding", "utf-8").lower().replace("-", "") != "utf8":
            _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, OSError):
        pass
import argparse
import hashlib
import json
import math
import os
import re
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from generate_rent_comps_table import (
    get_rent_schema, build_workbook, compute_eff_rent, RENT_SCHEMA_BASE,
    _fill, _font, _border, _align,
)
from generate_rent_comps_map import geocode_with_fallbacks, _parse_property_text
from tools.house_rules import search_rules as _house_rules, warn_window_vs_recency
from tools.calculations import find_same_building as _find_same_building


def _shared_mapbox_token() -> str:
    """Mapbox token fallback: shared_settings.json (single source of truth) → env."""
    try:
        p = Path(__file__).parent.parent / "configs" / "shared_settings.json"
        if p.exists():
            tok = (json.loads(p.read_text(encoding="utf-8")) or {}).get("mapbox_token", "")
            if tok:
                return tok
    except Exception:
        pass
    return os.environ.get("MAPBOX_TOKEN", "") or os.environ.get("MAPBOX_API_KEY", "")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — QUERY BUILDING
# ═══════════════════════════════════════════════════════════════════════════════

def _year_window(years_back: int) -> str:
    now = datetime.now().year
    return " OR ".join(str(now - i) for i in range(years_back + 1))


def build_rent_queries(subject_cfg: dict, level: str, years_back: int = 2) -> list:
    """
    Build web search queries for rental comps.
    level: "proximity" | "city" | "country"
    """
    prop_name    = subject_cfg["property_name"]
    address      = subject_cfg.get("address", "")
    asset_class  = subject_cfg.get("asset_class", "office")
    currency     = subject_cfg.get("currency", "SGD")
    country_name = subject_cfg.get("country_name", "Singapore")
    kws          = subject_cfg.get("submarket_keywords", [country_name])
    primary      = kws[0] if kws else country_name
    broader      = subject_cfg.get("rent_broader_query") or \
                   subject_cfg.get("broader_market_query", f"{country_name} {asset_class} rent")
    asset_kw     = subject_cfg.get("rent_search_keyword") or \
                   subject_cfg.get("asset_search_keyword", asset_class)
    yrs          = _year_window(years_back)

    if level == "proximity":
        parts    = [p.strip() for p in address.split(",") if p.strip()]
        precinct = parts[-3] if len(parts) >= 3 else (parts[-2] if len(parts) >= 2 else primary)
        return [
            f'{country_name} "{precinct}" {asset_kw} lease rent asking ({yrs}) {currency}',
            f'{country_name} "{primary}" {asset_kw} rental transaction ({yrs})',
        ]

    elif level == "city":
        kw_expr = " OR ".join(f'"{k}"' for k in kws[:3])
        return [
            f'{country_name} ({kw_expr}) {asset_kw} lease rent ({yrs}) {currency}',
            f'{broader} lease rental ({yrs})',
        ]

    else:  # country
        return [
            f'{broader} lease rental ({yrs})',
            f'{country_name} {asset_kw} asking rent market ({yrs}) JLL Savills CBRE Colliers',
        ]


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — OPENAI SEARCH + EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════════

def search_and_extract_rent(query: str, client, search_model: str,
                            extract_model: str, subject_cfg: dict = None) -> tuple:
    """
    Run one web search query and extract rental comp records.
    Returns (records: list[dict], sources: list[dict]).
    """
    country_name = (subject_cfg or {}).get("country_name", "Singapore")
    asset_class  = (subject_cfg or {}).get("asset_class",  "office")
    currency     = (subject_cfg or {}).get("currency",     "SGD")
    gfa_unit     = (subject_cfg or {}).get("gfa_unit",     "sf")
    period       = (subject_cfg or {}).get("rent_period",  "mth")

    system_prompt = (
        f"You search for {country_name} {asset_class} rental transaction data. "
        "Only include CONFIRMED lease transactions with actual rent figures. "
        "Exclude: general market reports without specific deals, "
        "property listings (not yet leased), rent guides without deal details."
    )

    extract_prompt = f"""From the search results, extract {asset_class} rental comparable transactions.

IMPORTANT: Only include {asset_class} leases. Exclude residential apartments, hotels,
serviced apartments, and other non-{asset_class} property types.

Return a JSON array. Each element must have these exact keys:
  property_name   : full building name
  address         : street address
  lease_date      : date or quarter/year the lease commenced (e.g. "Q1 2024", "Mar 2024")
  nla_sf          : leased GLA / NLA as a number in {gfa_unit}; null if unknown
  asking_rent     : gross face rent per {gfa_unit} per {period} as a number; null if unknown
  eff_rent        : effective rent per {gfa_unit} per {period} as a number; null if unknown
  lease_term_yrs  : lease term in years as a number; null if unknown (used to compute eff rent)
  rent_free_mths  : rent-free period in months as a number; null if unknown
  lease_type      : type of leased area and any comments, e.g. "Whole Floor ({asset_class.title()})", "Partial Floor"; "" if unknown
  currency        : "{currency}"
  source_url      : the URL where this data was found

Only include records that have ALL THREE of: property_name, nla_sf, and asking_rent (or eff_rent).
Records missing leased GLA or any rent figure must be excluded entirely.
Return [] if no qualifying records found."""

    # Search
    search_resp = client.chat.completions.create(
        model=search_model,
        messages=[
            {"role": "system",  "content": system_prompt},
            {"role": "user",    "content": f"Search for: {query}"},
        ],
    )
    search_text = search_resp.choices[0].message.content or ""

    # Collect source URLs from annotations
    sources = []
    annotations = getattr(search_resp.choices[0].message, "annotations", None) or []
    for ann in annotations:
        url   = getattr(ann, "url", None) or getattr(getattr(ann, "url_citation", None), "url", None)
        title = getattr(ann, "title", None) or getattr(getattr(ann, "url_citation", None), "title", None)
        if url:
            sources.append({"url": url, "title": title or url[:60]})

    # Extract
    extract_resp = client.chat.completions.create(
        model=extract_model,
        messages=[
            {"role": "system",  "content": "Extract structured JSON from the text. Return ONLY valid JSON."},
            {"role": "user",    "content": f"{extract_prompt}\n\nSearch results:\n{search_text}"},
        ],
        response_format={"type": "json_object"},
    )
    raw = extract_resp.choices[0].message.content or "[]"
    raw = re.sub(r"^```[a-z]*\n?", "", raw.strip())
    raw = re.sub(r"\n?```$", "", raw)

    try:
        parsed = json.loads(raw)
        if isinstance(parsed, dict):
            parsed = parsed.get("records") or parsed.get("comparables") or \
                     next((v for v in parsed.values() if isinstance(v, list)), [])
        records = parsed if isinstance(parsed, list) else []
    except Exception:
        records = []

    return records, sources


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — DEDUP + VALIDATION
# ═══════════════════════════════════════════════════════════════════════════════

# Shared asset-type guard (same logic as search_online_sales_comps.py)
_ASSET_ACCEPT_KWS = {
    "office":      ["office"],
    "logistics":   ["logistic", "warehouse", "cold storage", "distribution centre",
                    "distribution center", "fulfillment", "fulfilment"],
    "industrial":  ["industrial", "business park", "factory"],
    "retail":      ["retail", "shophouse", "mall"],
    "residential": ["residential", "apartment", "condominium", "condo", "landed"],
}
_ASSET_REJECT_KWS = {
    "office":      ["residential", "apartment", "condominium", "condo",
                    "hotel", "serviced apartment", "hospitality", "columbarium"],
    "logistics":   ["residential", "apartment", "condominium", "condo",
                    "hotel", "serviced apartment", "office"],
    "industrial":  ["residential", "apartment", "condominium", "condo",
                    "hotel", "serviced apartment"],
    "retail":      ["residential", "apartment", "condominium", "condo",
                    "hotel", "serviced apartment",
                    "logistics park", "warehouse", "industrial estate"],
    "residential": ["office", "logistics park", "logistics centre",
                    "warehouse", "industrial estate"],
}


def _asset_type_matches(record: dict, asset_class: str) -> bool:
    """Return True if the record's lease type is compatible with asset_class."""
    if not asset_class:
        return True
    combined = " ".join(str(record.get(f) or "") for f in (
        "lease_type", "asset_type", "property_type", "land_zoning", "comment",
    )).lower()
    if not combined.strip():
        return True
    ac = asset_class.lower()
    if any(kw in combined for kw in _ASSET_ACCEPT_KWS.get(ac, [])):
        return True
    if any(kw in combined for kw in _ASSET_REJECT_KWS.get(ac, [])):
        return False
    return True


def validate_dedup_rent(records: list, subject_name: str = "",
                        subject_country: str = "",
                        subject_asset_class: str = "") -> list:
    """Filter out subject property, wrong asset type, invalid records, and duplicates."""
    seen    = {}
    cleaned = []
    subj_tokens = set(re.sub(r"\W+", " ", subject_name.lower()).split())

    for r in records:
        name = str(r.get("property_name") or "").strip()
        if not name:
            continue
        # Skip if same as subject
        name_tokens = set(re.sub(r"\W+", " ", name.lower()).split())
        if len(name_tokens & subj_tokens) >= max(2, len(subj_tokens) * 0.6):
            continue
        # Asset-type filter: skip if clearly a different property type
        if subject_asset_class and not _asset_type_matches(r, subject_asset_class):
            print(f"    [type-filter] skipped '{name}' "
                  f"(not {subject_asset_class})")
            continue
        # Require both leased GLA and at least one rent figure
        if not r.get("nla_sf"):
            continue
        if not r.get("asking_rent") and not r.get("eff_rent"):
            continue

        rent = float(r.get("asking_rent") or r.get("eff_rent") or 0)
        key  = re.sub(r"\W+", "", name.lower())[:24] + f"_{rent:.0f}"
        if key in seen:
            continue
        seen[key] = True
        cleaned.append(r)

    return cleaned


def geocode_records(records: list, mapbox_token: str,
                    country_code: str = "", country_name: str = "") -> list:
    """Add lon/lat to each record. Only geocodes records that have a real address.
    Property names and descriptions are NOT used as geocoding queries — they
    produce unreliable or wrong coordinates."""
    suffix = f", {country_name}" if country_name else ""
    for r in records:
        addr = str(r.get("address") or "").strip()
        if not addr:
            r["lon"], r["lat"] = None, None
            continue
        query = f"{addr}{suffix}" if suffix not in addr else addr
        try:
            lon, lat, _ = geocode_with_fallbacks([query], mapbox_token, country_code)
            r["lon"], r["lat"] = lon, lat
        except Exception:
            r["lon"], r["lat"] = None, None
    return records


def _haversine_km(lon1, lat1, lon2, lat2) -> float:
    R = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — LLM CLASSIFICATION
# ═══════════════════════════════════════════════════════════════════════════════

def classify_rent_records(records: list, subject_cfg: dict,
                          client, extract_model: str) -> list:
    if not records:
        return []
    asset_class  = subject_cfg.get("asset_class",  "office")
    country_name = subject_cfg.get("country_name", "Singapore")
    entries = "\n".join(
        f'{i+1}. {r.get("property_name","")} — {r.get("address","")}'
        for i, r in enumerate(records)
    )
    prompt = (
        f"Classify these {country_name} {asset_class} rental comparables. "
        "Return a JSON array, one object per item, with keys:\n"
        "  index, location, quality, lease_type, relevance (0-10)\n\n"
        f"Items:\n{entries}"
    )
    try:
        resp = client.chat.completions.create(
            model=extract_model,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
        )
        raw = resp.choices[0].message.content or "[]"
        raw = re.sub(r"^```[a-z]*\n?", "", raw.strip())
        raw = re.sub(r"\n?```$", "", raw)
        parsed = json.loads(raw)
        if isinstance(parsed, dict):
            parsed = parsed.get("comparables") or next(
                (v for v in parsed.values() if isinstance(v, list)), []
            )
        for item in parsed:
            idx = int(item.get("index", 0)) - 1
            if 0 <= idx < len(records):
                records[idx].update({
                    "location":   item.get("location",   ""),
                    "quality":    item.get("quality",    ""),
                    "lease_type": item.get("lease_type", ""),
                    "relevance":  int(item.get("relevance", 5)),
                })
    except Exception as e:
        print(f"  [Classify] {e}")

    records.sort(key=lambda x: -int(x.get("relevance", 5)))
    for i, r in enumerate(records):
        r["map_marker"] = str(i + 1)
    return records


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — CACHE
# ═══════════════════════════════════════════════════════════════════════════════

def _cache_key(subject_cfg: dict, sc_cfg: dict) -> str:
    s = json.dumps({"name": subject_cfg.get("deal_name"),
                    "sc":   sc_cfg}, sort_keys=True)
    return hashlib.md5(s.encode()).hexdigest()[:12]

def load_cache(path: str, expected_key: str):
    try:
        with open(path, encoding="utf-8") as f:
            d = json.load(f)
        if d.get("key") == expected_key:
            return d.get("records", []), d.get("levels", [])
    except Exception:
        pass
    return None, None

def save_cache(path: str, records: list, levels: list, key: str):
    with open(path, "w", encoding="utf-8") as f:
        json.dump({"key": key, "records": records, "levels": levels}, f,
                  indent=2, ensure_ascii=False)
    print(f"  Cache saved → {path}")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — EXCEL WORKBOOK (ONLINE VERSION WITH SOURCES TAB)
# ═══════════════════════════════════════════════════════════════════════════════

def record_to_rent_row(r: dict, subject_cfg: dict) -> dict:
    name = str(r.get("property_name") or "").strip()
    addr = str(r.get("address") or "").strip()
    prop = f"{name}\n{addr}" if addr else name

    return {
        # Search origin ("Web search", "URA PMI", …). OUTPUT_SCHEMA has always had a
        # Source column; without this key it rendered blank. URLs are on 'Sources'.
        "source":         " + ".join(_record_origins(r)),
        "property":       prop,
        "map_marker":     str(r.get("map_marker", "")),
        "lease_date":     str(r.get("lease_date") or ""),
        "nla_sf":         r.get("nla_sf"),
        "lease_term_yrs": r.get("lease_term_yrs"),
        "asking_rent":    r.get("asking_rent"),
        "eff_rent":       r.get("eff_rent"),
        "location":       str(r.get("location") or ""),
        "quality":        str(r.get("quality") or ""),
        "lease_type":     str(r.get("lease_type") or r.get("asset_type") or ""),
    }


# Human-readable origin labels for each connector (source_name → label).
_SOURCE_LABELS = {
    "web_search":     "Web search",
    "ura_pmi":        "URA PMI",
    "broker_reports": "Broker report",
    "ura_gls":        "URA GLS",
}


def _source_label(name: str) -> str:
    if not name:
        return "Web search"
    return _SOURCE_LABELS.get(name, name.replace("_", " ").title())


def _record_origins(rec: dict) -> list:
    """Distinct origin labels for a record (e.g. ['Web search', 'Broker report'])."""
    out = []
    for s in rec.get("sources") or []:
        lbl = _source_label(s.get("source_name") or "")
        if lbl not in out:
            out.append(lbl)
    return out or ["Web search"]


def build_workbook_online_rent(subject_row: dict, comp_rows: list,
                                subject_cfg: dict, output_path: str,
                                records_with_sources: list = None):
    schema = get_rent_schema(subject_cfg)

    from generate_rent_comps_table import build_workbook as _bw
    _bw(subject_cfg, comp_rows, output_path, schema)

    # Reload to add Sources sheet
    wb = openpyxl.load_workbook(output_path)
    ws_src = wb.create_sheet("Sources")
    ws_src.column_dimensions["A"].width = 6
    ws_src.column_dimensions["B"].width = 38
    ws_src.column_dimensions["C"].width = 16
    ws_src.column_dimensions["D"].width = 80

    # Header
    for col, hdr in enumerate(["#", "Property", "Source", "Source URL"], 1):
        c = ws_src.cell(row=1, column=col, value=hdr)
        c.fill   = _fill("1A3A5C")
        c.font   = _font(bold=True)
        c.border = _border()
        c.alignment = _align()

    row_idx = 2
    for r in (records_with_sources or []):
        for src in r.get("sources", []):
            ws_src.cell(row=row_idx, column=1, value=r.get("map_marker", ""))
            ws_src.cell(row=row_idx, column=2, value=r.get("property_name", ""))
            sc = ws_src.cell(row=row_idx, column=3, value=_source_label(src.get("source_name") or ""))
            sc.font = _font(bold=True)
            url_cell = ws_src.cell(row=row_idx, column=4, value=src.get("url", ""))
            url_cell.hyperlink = src.get("url", "")
            url_cell.style = "Hyperlink"
            row_idx += 1

    wb.save(output_path)
    print(f"  Saved → {output_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 7 — MAIN RUN
# ═══════════════════════════════════════════════════════════════════════════════

def run(config_path: str = "configs/deal_config.json",
        generate_map: bool = False, refresh: bool = False):
    with open(config_path, encoding="utf-8") as f:
        cfg = json.load(f)

    subject_cfg   = cfg["subject_property"]
    mb_cfg        = cfg.get("mapbox", {})
    mapbox_tok    = mb_cfg.get("token", "") or _shared_mapbox_token()
    oa_cfg        = cfg.get("openai", {})
    api_key       = oa_cfg.get("api_key") or os.environ.get("OPENAI_API_KEY", "")
    search_model  = oa_cfg.get("search_model",  "gpt-4o-mini-search-preview")
    extract_model = oa_cfg.get("extract_model", "gpt-4o-mini")
    country_code  = cfg.get("country_code", "")
    country_name  = subject_cfg.get("country_name", "")
    prop_name     = subject_cfg["property_name"]

    # rent_search overrides online_search for rent-specific settings
    sales_cfg       = cfg.get("online_search", {})
    # House rules (configs/shared_settings.json → search_rules) apply to every
    # deal; rent_search (falling back to online_search) overrides per deal.
    sc_cfg          = _house_rules("rent", {**sales_cfg, **cfg.get("rent_search", {})},
                                   subject_cfg.get("asset_class", ""))

    proximity_km    = sc_cfg.get("proximity_km",    3.0)
    city_km         = sc_cfg.get("city_km",        25.0)
    min_results     = sc_cfg.get("min_results",     5)     # rent default: 5 (vs 3 for sales)
    max_results     = sc_cfg.get("max_results",     15)
    years_back      = sc_cfg.get("years_back",      2)
    years_back_max  = sc_cfg.get("years_back_max",  3)
    years_back_step = sc_cfg.get("years_back_step", 1)
    max_level       = sc_cfg.get("max_level",       3)
    # Recency filter (web-search + connectors): keep only leases within recency_months.
    # 36mo, tighter than the 60mo used for sales and land: rental evidence dates
    # faster than capital evidence, so a lease from the last cycle says little about
    # today's achievable rent. Unparseable dates are KEPT.
    from sources.base import months_ago as _months_ago
    _rec_m = int(sc_cfg.get("recency_months", 36) or 36)
    _w = warn_window_vs_recency(years_back_max, _rec_m)
    if _w:
        print(_w)


    # Hard budget on web queries per run. One query = 1 web search + 1 extract call,
    # so max_queries=5 costs 5 searches + 5 extracts, plus 1 classification = 11
    # OpenAI calls. Config overrides via online_search.max_queries.
    max_queries = int(sc_cfg.get("max_queries", 5) or 5)
    _queries    = {"n": 0}

    if not api_key:
        raise ValueError("OpenAI API key not found. "
                         "Set OPENAI_API_KEY or openai.api_key in config.")
    if not mapbox_tok:
        raise ValueError("Mapbox token missing — add mapbox.token to config.")

    try:
        from openai import OpenAI
    except ImportError:
        raise ImportError("Run: pip3 install openai")

    client    = OpenAI(api_key=api_key)
    deal_name = subject_cfg.get("deal_name", prop_name)
    deal_slug = deal_name.replace(" ", "_")

    output_file = cfg.get("output_file",
                  f"output/{deal_slug}/Transaction_Comparables_{deal_slug}.xlsx")
    out_dir     = Path(output_file).parent
    out_dir.mkdir(parents=True, exist_ok=True)

    out_excel  = str(out_dir / f"Online_Rent_Comps_{deal_slug}.xlsx")
    out_map    = str(out_dir / f"Online_Rent_Comps_{deal_slug}_map.png")
    cache_path = str(out_dir / f"Online_Rent_Comps_{deal_slug}_search_cache.json")

    print(f"\n{'='*62}\n  Online Rent Comps Search : {deal_name}\n{'='*62}")

    # Geocode subject
    address = subject_cfg.get("address", "")
    print(f"\n[0/5] Geocoding subject property")
    s_lon, s_lat, _ = geocode_with_fallbacks(
        [f"{prop_name}, {address}", address, prop_name],
        mapbox_tok, country_code,
    )
    print(f"      {prop_name} → ({s_lon}, {s_lat})")

    # Cache check
    c_key = _cache_key(subject_cfg, sc_cfg)
    if not refresh:
        print(f"\n[Search] Checking cache…")
        cached_records, cached_levels = load_cache(cache_path, c_key)
    else:
        print(f"\n[Search] --refresh: discarding any existing cache.")
        cached_records, cached_levels = None, None

    if cached_records is not None:
        all_records = cached_records
        levels_used = cached_levels
    else:
        GEO_LEVELS  = [
            ("proximity", proximity_km, f"Proximity ≤{proximity_km}km"),
            ("city",      city_km,      f"City ≤{city_km}km"),
        ]
        # Tier 3: whole country — max_km None means no distance filter at all.
        BROAD_LEVEL = ("country", None, f"Country-wide ({country_name or 'all'})")

        all_records:     list = []
        levels_used:     list = []
        seen_keys:       dict = {}
        # cross-source dupes are matched by location+rent via _find_same_building.
        yrs = years_back

        def _run_level(level_id, max_km, level_label, yrs_used):
            queries   = build_rent_queries(subject_cfg, level_id, yrs_used)
            level_new = 0
            print(f"\n[Search] {level_label}  (years back: {yrs_used})")
            for q in queries:
                if _queries["n"] >= max_queries:
                    print(f"  · query budget reached ({max_queries}) — stopping search")
                    break
                print(f"  Query: {q[:80]}…" if len(q) > 80 else f"  Query: {q}")
                _queries["n"] += 1
                try:
                    raw, q_sources = search_and_extract_rent(
                        q, client, search_model, extract_model, subject_cfg
                    )
                except Exception as e:
                    print(f"  ✗ {e.__class__.__name__}: {e}")
                    continue
                if q_sources:
                    print(f"  ✦ {len(q_sources)} source(s) found")
                cleaned  = validate_dedup_rent(raw, subject_name=prop_name,
                                               subject_country=country_name,
                                               subject_asset_class=subject_cfg.get("asset_class", ""))
                _b = len(cleaned)
                cleaned  = [r for r in cleaned
                            if (_months_ago(str(r.get("lease_date") or "")) or 0) <= _rec_m]
                if len(cleaned) != _b:
                    print(f"    [recency] dropped {_b - len(cleaned)} lease(s) older than "
                          f"{_rec_m}mo; kept {len(cleaned)}")
                geocoded = geocode_records(cleaned, mapbox_tok, country_code, country_name)
                for r in geocoded:
                    rent  = float(r.get("asking_rent") or r.get("eff_rent") or 0)
                    key   = re.sub(r"\W+", "", str(r.get("property_name","")).lower())[:24] \
                            + f"_{rent:.1f}"
                    if max_km is not None and r.get("lon") is not None:
                        if _haversine_km(r["lon"], r["lat"], s_lon, s_lat) > max_km:
                            continue
                    if _find_same_building(all_records, r.get("lon"), r.get("lat"), rent,
                                           lambda x: x.get("asking_rent") or x.get("eff_rent")) is not None:
                        continue
                    if key in seen_keys:
                        idx = seen_keys[key]
                        existing_urls = {s["url"] for s in all_records[idx].get("sources",[])}
                        for s in q_sources:
                            if s["url"] and s["url"] not in existing_urls:
                                all_records[idx].setdefault("sources",[]).append(
                                    {**s, "source_name": "web_search"})
                        continue
                    r["sources"] = [{**s, "source_name": "web_search"} for s in q_sources]
                    seen_keys[key] = len(all_records)
                    all_records.append(r)
                    level_new += 1
            label_str = f"{level_label} (yrs:{yrs_used})"
            if label_str not in levels_used:
                levels_used.append(label_str)
            print(f"  → {level_new} new  |  total: {len(all_records)}")
            return level_new

        # Grounded data sources to combine with (or instead of) OpenAI web search.
        # Defaults to web-search only → identical behaviour to before.
        sources_cfg   = sc_cfg.get("sources") or ["web_search"]
        _web          = "web_search" in sources_cfg

        active_levels = GEO_LEVELS[:max_level]
        while _web and len(all_records) < min_results:
            for level_id, max_km, level_label in active_levels:
                _run_level(level_id, max_km, level_label, yrs)
                if len(all_records) >= min_results:
                    break
            if len(all_records) >= min_results:
                print(f"  ✓ min_results ({min_results}) reached.")
                break
            if yrs < years_back_max:
                yrs += years_back_step
                print(f"\n  ↩  Sparse results ({len(all_records)}). "
                      f"Extending to {yrs} years back…")
            else:
                print(f"\n  ↩  Temporal limit ({years_back_max} yrs) reached. "
                      f"{len(all_records)} record(s) found in geo levels.")
                break

        if _web and len(all_records) < min_results and max_level >= 3:
            print(f"\n  Falling back to broader market search.")
            _run_level(*BROAD_LEVEL, yrs)
            if len(all_records) < min_results:
                print(f"  ⚠  All levels exhausted — {len(all_records)} record(s).")

        # ── Grounded connectors (URA PMI rents, …): fetch once, ingest ──────────
        from sources.registry import get_grounded
        from sources.base import months_ago as _months_ago
        _conn_params = {"country_code": country_code, "country_name": country_name,
                        "years_back": yrs, "s_lon": s_lon, "s_lat": s_lat,
                        "proximity_km": proximity_km, "city_km": city_km,
                        "comp_type": "rent",
                        "client": client, "extract_model": extract_model,
                        "ura_max_rows": sc_cfg.get("ura_max_rows", 300),
                        "broker_pages": sc_cfg.get("broker_pages"),
                        "broker_max_pdfs": sc_cfg.get("broker_max_pdfs", 4)}
        # _rec_m comes from run scope: one recency cap for every source.
        for _conn in get_grounded((country_code or "sg").lower(), "rent",
                                  sources_cfg, _conn_params):
            print(f"\n[Source] {_conn.label or _conn.name}")
            try:
                _raw, _srcs = _conn.fetch(subject_cfg, _conn_params)
            except Exception as e:
                print(f"  ✗ {e.__class__.__name__}: {e}")
                continue
            if not _raw:
                print("  → 0 records")
                continue
            _cleaned = validate_dedup_rent(
                _raw, subject_name=prop_name, subject_country=country_name,
                subject_asset_class=subject_cfg.get("asset_class", ""))
            _geo = geocode_records(_cleaned, mapbox_tok, country_code, country_name)
            _base_srcs = _srcs or [{"title": _conn.label or _conn.name, "url": ""}]
            _added = 0
            for r in _geo:
                if r.get("lon") is None:
                    continue
                if _haversine_km(r["lon"], r["lat"], s_lon, s_lat) > city_km:
                    continue
                rent = float(r.get("asking_rent") or r.get("eff_rent") or 0)
                key  = re.sub(r"\W+", "", str(r.get("property_name", "")).lower())[:24] \
                       + f"_{rent:.1f}"
                if key in seen_keys:
                    continue
                r["sources"] = [{**s, "source_name": _conn.name} for s in _base_srcs]
                seen_keys[key] = len(all_records)
                all_records.append(r)
                _added += 1
            _lbl = _conn.label or _conn.name
            if _lbl not in levels_used:
                levels_used.append(_lbl)
            print(f"  → {_added} new  |  total: {len(all_records)}")

        save_cache(cache_path, all_records, levels_used, c_key)

    records = all_records
    print(f"\n[1/5] SEARCH complete — {len(records)} record(s) for processing")

    if not records:
        print("\n  No rental comps found. Consider widening search parameters.\n")
        return

    # Classify
    print(f"\n[2/5] CLASSIFY  ({extract_model})")
    classified = classify_rent_records(records, subject_cfg, client, extract_model)
    print(f"      → {len(classified)} records classified")

    # Trim AFTER classify, not before: classify_rent_records is what scores relevance
    # and sorts by it. Trimming first kept whichever comps the earliest query happened
    # to find and threw away better ones unseen.
    if len(classified) > max_results:
        print(f"      → keeping the {max_results} most relevant of {len(classified)}")
        classified = classified[:max_results]
        for i, r in enumerate(classified, 1):
            r["map_marker"] = str(i)

    # ── Location competitiveness (SG, URA proximity vs subject) ───────────────
    # Same Superior/Comparable/Inferior logic as the internal pipeline, reusing the
    # map-resolved lon/lat on each comp + the subject. SG-only; others left blank.
    try:
        from tools.location_score import apply_location as _apply_loc
        classified = _apply_loc(classified,
                                subject_cfg.get("property_name", ""),
                                subject_cfg.get("address", ""),
                                subject_cfg.get("asset_class", ""),
                                subj_lonlat=(s_lon, s_lat))
    except Exception as _le:
        print(f"  [location] skipped: {_le}")

    # Compute effective rent
    print(f"\n[3/5] CALCULATE  (effective rent from rent-free)")
    classified = compute_eff_rent(classified)
    params      = cfg.get("parameters", {})
    bala_yield  = params.get("bala_yield", 0.06)

    print(f"\n  {'#':<3} {'Property':<42} {'km':>5} {'Asking':>8} {'Eff.Rent':>8}")
    print("  " + "─" * 70)
    for r in classified:
        km      = _haversine_km(r["lon"], r["lat"], s_lon, s_lat) if r.get("lon") else 0
        asking  = float(r.get("asking_rent") or 0)
        eff     = float(r.get("eff_rent")    or 0)
        name    = str(r.get("property_name", ""))[:42]
        print(f"  {r.get('map_marker',''):<3} {name:<42} {km:>5.1f} {asking:>8.2f} {eff:>8.2f}")

    # Sources summary
    total_sources = sum(len(r.get("sources", [])) for r in classified)
    print(f"\n  Sources ({total_sources} URL(s) across {len(classified)} comp(s)):")
    for r in classified:
        srcs = r.get("sources", [])
        if srcs:
            print(f"  {r.get('map_marker','?')}. {r.get('property_name','')}")
            for s in srcs:
                title = (s.get("title") or s.get("url", ""))[:60]
                print(f"       ✦ {title}")
                print(f"         {s.get('url','')}")

    # Render Excel
    print(f"\n[4/5] RENDER   {out_excel}")
    subj_row  = {
        "property":       f"{subject_cfg['property_name']}\n{subject_cfg.get('address','')}",
        "map_marker":     "★",
        "lease_date":     subject_cfg.get("sale_date", ""),
        "nla_sf":         subject_cfg.get("gfa_sf"),
        "lease_term_yrs": None,
        "asking_rent":    subject_cfg.get("asking_rent"),
        "eff_rent":       subject_cfg.get("eff_rent"),
        "location":       subject_cfg.get("location", ""),
        "quality":        subject_cfg.get("quality", ""),
        "lease_type":     subject_cfg.get("asset_type", ""),
    }
    comp_rows = [record_to_rent_row(r, subject_cfg) for r in classified]
    build_workbook_online_rent(subj_row, comp_rows, subject_cfg, out_excel,
                               records_with_sources=classified)

    # Render map
    if generate_map:
        print(f"\n[5/5] MAP   {out_map}")
        from generate_rent_comps_map import render_map
        comps_geo = [
            (r["map_marker"], r["lon"], r["lat"])
            for r in classified
            if r.get("lon") is not None
        ]
        render_map(
            subject_lonlat=(s_lon, s_lat),
            comps=comps_geo,
            token=mapbox_tok,
            output_path=out_map,
            style=mb_cfg.get("style", "streets-v12"),
            width=mb_cfg.get("width", 1200),
            height=mb_cfg.get("height", 900),
            padding=mb_cfg.get("padding", 100),
            pin_size=mb_cfg.get("pin_size", "l"),
        )
    else:
        print(f"\n[5/5] MAP   skipped (pass --map to generate)")

    print(f"\nDone.\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Search online for rental comparables")
    parser.add_argument("--config",  default="configs/deal_config.json")
    parser.add_argument("--map",     action="store_true", help="Generate map PNG")
    parser.add_argument("--refresh", action="store_true", help="Ignore cache")
    args = parser.parse_args()
    run(args.config, generate_map=args.map, refresh=args.refresh)
