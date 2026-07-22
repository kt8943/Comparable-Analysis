#!/usr/bin/env python3
"""
generate_sales_comps_table.py
=============================
Schema definition and Excel workbook builder for Asset Sales Comparables.

Mirrors the pattern of generate_rent_comps_table.py and
generate_land_comps_table.py — all rendering logic lives here so both
scan_input_sales_comps.py and search_online_sales_comps.py share
identical output formatting.

Fixed Output Schema  (14 visible columns + 1 hidden stake % column)
--------------------------------------------------------------------
  Source | Property | Map Marker | Sale Date | Land Zoning | Remaining Leasehold (Y)
  GFA (SF) | Price (SGD M) | Price (SGD psf GFA) | FTM NOI Cap Rate
  Adj. Cap Rate | Location | Quality | Asset Type
  + hidden col O = Stake % (used by psf formula; not printed)

Calculation transparency
------------------------
  A 'Params' worksheet is added to every output workbook:
    B2 = Subject Remaining Leasehold (yrs)  ← referenced by Adj CR formula
    B3 = Bala Table Yield (y)               ← change to recalculate all rows
  Live Excel formulas (auditable, not hard-coded values):
    Price psf  (col H)  =  G / N × 1,000,000 / F
    Adj CR     (col J)  =  I × Bala(E) / Bala(Params!B2)
                           where Bala(n) = 1 − (1/(1+Params!$B$3))^n
"""

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from tools.calculations import (bala_factor, bala_expr as _bala_expr, _BALA_TABLE,
                                parse_cap_rate as _cap_rate,
                                parse_remaining_yrs as _parse_remaining_yrs)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — FIXED OUTPUT SCHEMA
#
# WHAT  : The 13 visible output columns. Never changes between deals.
# HOW   : (display_header, internal_key, dtype, excel_format, col_width)
#         dtype drives alignment and number_format:
#           "str"   → centre-align, text format
#           "int"   → right-align, integer format
#           "float" → right-align, decimal format
#           "pct"   → centre-align, percentage format
# ═══════════════════════════════════════════════════════════════════════════════

OUTPUT_SCHEMA = [
    ("Type",                                              "type",          "str",   "@",               10),
    ("Source",                                            "source",        "str",   "@",               10),
    ("Property",                                          "property",      "str",   "@",               36),
    ("Map\nMarker",                                       "map_marker",    "str",   "@",                8),
    ("Sale Date",                                         "sale_date",     "str",   "@",               13),
    ("Land Zoning",                                       "land_zoning",   "str",   "@",               20),
    ("Remaining\nLeasehold (Y)",                          "remaining_yrs", "float", '#,##0.# "yrs"',   14),
    ("GFA (SF)",                                          "gfa_sf",        "int",   "#,##0",           13),
    ("Price\n(SGD M)",                                    "price_sgd_m",   "float", '"S$"#,##0.0',  13),
    ("Price\n(SGD psf GFA)",                             "price_psf_gfa", "int",   '"S$"#,##0',       16),
    ("FTM NOI\nCapitalisation Rate",                     "ftm_cap_rate",  "pct",   "0.00%",           18),
    ("Adj. Capitalisation Rate\n(Subject {subj_yrs} Yrs Remaining)",
                                                          "adj_cap_rate",  "pct",   "0.00%",           28),
    ("Location",                                          "location",      "str",   "@",               24),
    ("Quality",                                           "quality",       "str",   "@",               28),
    ("Asset Type",                                        "asset_type",    "str",   "@",               24),
]

# Hidden stake % column sits one past the last visible column (auto-adjusts)
_STAKE_COL = get_column_letter(len(OUTPUT_SCHEMA) + 1)   # col P

def _schema_idx(key: str, schema: list) -> int:
    """Return 1-based column index for the given internal key."""
    for i, entry in enumerate(schema, 1):
        if entry[1] == key:
            return i
    raise KeyError(f"Schema key {key!r} not found")


def get_output_schema(subject_cfg: dict = None) -> list:
    """
    Return OUTPUT_SCHEMA adapted for this deal's currency and area unit.
    Uses key-based lookup so column order changes never break indices.
    """
    if subject_cfg is None:
        return list(OUTPUT_SCHEMA)
    currency   = subject_cfg.get("currency", "SGD")
    price_unit = subject_cfg.get("price_unit", "M").upper()
    gfa_unit   = subject_cfg.get("gfa_unit", "sf").lower()
    sym        = subject_cfg.get("currency_symbol", currency)
    gfa_lbl    = "sqm" if gfa_unit == "sqm" else "SF"
    area_lbl   = "psm" if gfa_unit == "sqm" else "psf"

    schema = list(OUTPUT_SCHEMA)
    for i, entry in enumerate(schema):
        if entry[1] == "gfa_sf":
            schema[i] = (f"GFA ({gfa_lbl.upper()})", "gfa_sf", "int",
                         "#,##0", entry[4])
        elif entry[1] == "price_sgd_m":
            schema[i] = (f"Price\n({currency} {price_unit})", "price_sgd_m", "float",
                         f'"{sym}"#,##0.0', entry[4])
        elif entry[1] == "price_psf_gfa":
            schema[i] = (f"Price\n({currency} {area_lbl} GFA)", "price_psf_gfa", "int",
                         f'"{sym}"#,##0', entry[4])
    return schema


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — ROW CONVERTERS
# ═══════════════════════════════════════════════════════════════════════════════

def _blank_if_absent(v):
    """Numeric value, or None when the source reported none for this field.

    A missing GFA / remaining leasehold / cap rate is UNKNOWN, not zero. Coercing
    it to 0 prints a confident '0.0' (or '0.00%') that reads as a measured value
    the report never had — the analyst cannot tell it apart from a real figure.
    _data_row renders None as '—', which says 'not reported' honestly. 0 is never
    a real GFA/leasehold/cap rate, so treat it as absent too.
    """
    try:
        f = float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return None
    return f or None


def subject_to_row(cfg: dict) -> dict:
    price_m    = cfg.get("price_sgd_m")
    price_unit = cfg.get("price_unit", "M")
    gfa        = _blank_if_absent(cfg.get("gfa_sf"))
    prop_name  = cfg["property_name"]
    address    = cfg.get("address", "")
    prop_cell  = (f"{prop_name}\n{address}"
                  if address and address.strip() != prop_name.strip()
                  else prop_name)
    return {
        "type":          "Subject",
        "source":        "",
        "property":      prop_cell,
        "map_marker":    "★",
        "sale_date":     cfg.get("sale_date", ""),
        "land_zoning":   cfg.get("land_zoning", ""),
        "remaining_yrs": _parse_remaining_yrs(cfg.get("remaining_leasehold_yrs")),
        "gfa_sf":        int(gfa) if gfa else None,
        "price_sgd_m":   _m_to_display(price_m, price_unit),
        # gfa used to default to 1 so this division could not raise — which
        # printed the subject's GFA as a literal "1" and a psf of price×1e6.
        # With no GFA there is no psf to report.
        "price_psf_gfa": round(float(price_m) * 1e6 / gfa) if (price_m and gfa) else None,
        "ftm_cap_rate":  _cap_rate(cfg.get("ftm_noi_cap_rate")),
        "adj_cap_rate":  None,    # subject IS the reference
        "location":      cfg.get("location", ""),
        "quality":       cfg.get("quality", ""),
        "asset_type":    cfg.get("asset_type", ""),
        "stake_pct":     1.0,
    }


def _calc_price_psf(c: dict):
    """Price psf GFA. price_sgd_m is always M-normalized, so always × 1e6.

    RULE (precedence): use the source's directly-reported unit price when the
    input provides one (e.g. Colliers' 'Unit Price (SGD/psf)' = 3,757 for
    Northpoint City South Wing). ONLY when the input has no reported unit price
    do we derive it by calculation, price ÷ GFA (stake-adjusted to a 100% basis)."""
    # 1) Directly reported by the source input — take it as-is.
    try:
        reported = float(str(c.get("price_psf_gfa")).replace(",", "").strip())
    except (TypeError, ValueError):
        reported = 0.0
    if reported > 0:
        return round(reported)
    # 2) No reported unit price → derive from price ÷ GFA.
    price_m = float(c.get("price_sgd_m") or 0) or None
    gfa_sf  = int(c.get("gfa_sf") or 0) or None
    stake   = float(c.get("stake_pct") or 1.0) or 1.0
    if price_m and gfa_sf:
        return round(price_m * 1_000_000 / stake / gfa_sf)
    return None


def _m_to_display(price_m, price_unit: str):
    """Convert M-normalized stored price to display unit (B → ÷1000, else no-op)."""
    if price_m is None:
        return None
    try:
        v = float(price_m)
        return round(v / 1000, 4) if price_unit.upper() == "B" else v
    except (TypeError, ValueError):
        return price_m


def comp_to_row(c: dict, price_unit: str = "M") -> dict:
    name = str(c.get("property_name") or
               c.get("raw_description", "").split("\n")[0]).strip()
    prop = name
    _src = c.get("_source", "")
    _src_map = {"excel": "Excel", "pdf": "PDF", "image": "Image", "manual": "Manual"}
    if _src.startswith("pdf_"):
        _src_map[_src] = "PDF " + _src[4:]
    if _src.startswith("excel_"):
        _src_map[_src] = "Excel " + _src[6:]
    if _src.startswith("image_"):
        _src_map[_src] = "Image " + _src[6:]
    # parse_remaining_yrs, not a bare float(): the scan modules normally hand us
    # a number, but a value typed into the preview ("99 years from 2004",
    # "Freehold") arrives raw, and a lease we can DERIVE must be derived rather
    # than blanked. Returns None only when there is genuinely nothing to compute.
    _yrs = _parse_remaining_yrs(c.get("remaining_yrs"))
    _gfa = _blank_if_absent(c.get("gfa_sf"))
    return {
        "type":          "Comparable",
        "source":        _src_map.get(c.get("_source", ""), ""),
        "property":      prop,
        "map_marker":    str(c.get("map_marker", "")),
        "sale_date":     str(c.get("sale_date", "")),
        "land_zoning":   str(c.get("land_zoning", "")),
        "remaining_yrs": (None if _yrs is None else
                          "FH" if _yrs >= 999 else round(_yrs, 1)),
        "gfa_sf":        int(_gfa) if _gfa else None,
        # If the source reported a price range (e.g. "600-630"), show the original
        # string as the display value; use the numeric midpoint for psf computation.
        "price_sgd_m":   (c.get("price_sgd_m_display")
                          or _m_to_display(c.get("price_sgd_m"), price_unit)),
        # If the source reported a scaled/qualified value we can't compute a psf
        # from (e.g. a hotel "1.59 million/per key"), show that original figure
        # rather than a blank — same "preserve what was reported" precedent as
        # price_sgd_m_display above, just for this column.
        "price_psf_gfa": (c.get("price_psf_gfa_display") or _calc_price_psf(c)),
        "ftm_cap_rate":  _blank_if_absent(c.get("ftm_cap_rate")),
        # Written statically here; _write_formulas replaces it with the live Bala
        # formula UNLESS the source reported an Adj. Cap Rate of its own, which
        # this flag marks — a direct input is never overwritten by a calculation.
        "adj_cap_rate":  _blank_if_absent(c.get("adj_cap_rate")),
        "_adj_reported": _cap_rate(c.get("adj_npi_yield")) is not None,
        "location":      str(c.get("location", "")),
        "quality":       str(c.get("quality", "")),
        "asset_type":    str(c.get("asset_type", "")),
        "stake_pct":     float(c.get("stake_pct") or 1.0),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — STAGE 4: RENDER
#
# WHAT  : Write the formatted Excel workbook.
# HOW   : Two-table layout on "Transaction Comparables" sheet.
#         "Params" sheet holds Bala parameters referenced by formulas.
#         Live Excel formulas for Price psf and Adj Cap Rate (auditable).
#         8-char ARGB colours (FF prefix = fully opaque).
# ═══════════════════════════════════════════════════════════════════════════════

# ── Palette (8-char ARGB — FF prefix = fully opaque) ─────────────────────────
_NAVY  = "FF1F3864"
_NAVYL = "FF2E4C7E"
_WHITE = "FFFFFFFF"
_LGRAY = "FFF2F2F2"
_NOTE  = "FFEBF3FB"
_DARK  = "FF1A1A1A"
_RED   = "FFC00000"   # subject property star marker colour
_AVGBG = "FFD6DCE4"   # blue-grey for average row

def _fill(h):  return PatternFill(patternType="solid", fgColor=h)
def _font(color=_WHITE, bold=True, sz=9, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=sz, italic=italic)
_T = Side(style="thin", color="FFBFBFBF")
def _border(): return Border(left=_T, right=_T, top=_T, bottom=_T)
def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _section_header(ws, row: int, label: str, nc: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row=row, column=1, value=label)
    c.fill = _fill(_NAVY); c.font = _font(_WHITE, bold=True, sz=10)
    c.alignment = _align("left", "center", wrap=False); c.border = _border()
    for col in range(2, nc + 1):
        x = ws.cell(row=row, column=col)
        x.fill = _fill(_NAVY); x.border = _border()
    ws.row_dimensions[row].height = 19


def _col_headers(ws, row: int, subj_yrs: int, schema=None):
    schema = schema or OUTPUT_SCHEMA
    for col, (hdr, *_) in enumerate(schema, 1):
        c = ws.cell(row=row, column=col, value=hdr.replace("{subj_yrs}", str(subj_yrs)))
        c.fill = _fill(_NAVY); c.font = _font(_WHITE, bold=True, sz=8.5)
        c.alignment = _align("center", "center", wrap=True); c.border = _border()
    ws.row_dimensions[row].height = 34


def _data_row(ws, row: int, row_dict: dict, alt: bool = False, bold: bool = False,
              is_subject: bool = False, schema=None):
    """Write static values for all 13 visible columns + hidden stake % (col N)."""
    schema = schema or OUTPUT_SCHEMA
    bg = _fill(_LGRAY) if alt else _fill(_WHITE)
    for col, entry in enumerate(schema, 1):
        key, dtype, fmt = entry[1], entry[2], entry[3]
        val     = row_dict.get(key)
        display = "—" if val is None else val

        c = ws.cell(row=row, column=col, value=display)
        c.fill = bg; c.border = _border()
        c.font = _font(_DARK, bold=bold, sz=9)

        if col == 1:
            c.alignment = _align("left", "center", wrap=True)
        elif col == 2:
            # Source column — centred, normal weight
            c.alignment = _align("center", "center", wrap=False)
        elif col == 3:
            c.alignment = _align("center", "center", wrap=False)
            # Subject property name is red; comp names stay dark
            marker_color = _RED if is_subject else _DARK
            c.font = _font(marker_color, bold=True, sz=12)
        elif dtype in ("int", "float") and isinstance(val, (int, float)):
            c.alignment = _align("right", "center", wrap=False)
            c.number_format = fmt
        elif dtype == "pct" and isinstance(val, float):
            c.alignment = _align("center", "center", wrap=False)
            c.number_format = fmt
        else:
            c.alignment = _align("center", "center", wrap=True)

    # Hidden stake % in col N (used by psf formula)
    sc = ws.cell(row=row, column=len(schema) + 1,
                 value=row_dict.get("stake_pct", 1.0))
    sc.number_format = "0.0%"

    ws.row_dimensions[row].height = 48


def _write_formulas(ws, row: int, is_subject: bool = False, schema=None,
                    row_dict: dict = None):
    """
    Write live Excel formula for Adj CR (col J).

    Adj CR = FTM(I) × Bala(comp_yrs=E) / Bala(subj_yrs=Params!B2)
             Bala(n) = 1 − (1 / (1 + Params!$B$3))^n

    Price psf/psm is now written as a Python-computed static value in _data_row
    so it appears correctly in the Streamlit preview (openpyxl cannot evaluate
    Excel formulas when reading with data_only=True).
    """
    schema  = schema or OUTPUT_SCHEMA
    I       = get_column_letter(_schema_idx("ftm_cap_rate",  schema))
    E       = get_column_letter(_schema_idx("remaining_yrs", schema))
    adj_col = _schema_idx("adj_cap_rate",  schema)

    if is_subject:
        return   # subject Adj CR stays as "—"

    # The source printed its own Adj. Cap Rate — _data_row already wrote that
    # direct input, so leave the cell alone instead of overwriting it with the
    # derived Bala figure.
    if (row_dict or {}).get("_adj_reported"):
        return

    # Adj Cap Rate via official Singapore Bala Table (VLOOKUP in 'Bala Tbl' sheet)
    bala_comp = _bala_expr(f"{E}{row}")
    bala_subj = _bala_expr("Params!$B$2")
    adj       = ws.cell(row=row, column=adj_col)
    # bala_expr already treats a "—" leasehold as freehold, but an unreported FTM
    # cap rate (I) is now "—" rather than 0, and "—"×factor is #VALUE!. No cap
    # rate means no adjusted cap rate — show the same "—" the other cells use.
    adj.value        = f'=IFERROR({I}{row}*{bala_comp}/{bala_subj},"—")'
    adj.number_format = "0.00%"
    adj.alignment    = _align("center", "center", wrap=False)
    adj.font         = _font(_DARK, bold=False, sz=9)


def _avg_row(ws, row: int, first_r: int, last_r: int, schema=None):
    """Write Average row — Price psf GFA, FTM Cap Rate, Adj. Cap Rate."""
    schema   = schema or OUTPUT_SCHEMA
    avg_keys = {"price_psf_gfa", "ftm_cap_rate", "adj_cap_rate"}
    for col, entry in enumerate(schema, 1):
        key, dtype, fmt = entry[1], entry[2], entry[3]
        c = ws.cell(row=row, column=col)
        c.fill   = _fill(_AVGBG)
        c.border = _border()
        if key == "property":
            c.value     = "Average"
            c.font      = _font(_DARK, bold=True, sz=9)
            c.alignment = _align("left", "center", wrap=False)
        elif key in avg_keys:
            col_ltr         = get_column_letter(col)
            c.value         = (f'=IFERROR(AVERAGEIF({col_ltr}{first_r}:'
                               f'{col_ltr}{last_r},">0"),"—")')
            c.number_format = fmt
            c.font          = _font(_DARK, bold=True, sz=9)
            c.alignment     = _align("center" if dtype == "pct" else "right",
                                     "center", wrap=False)
        else:
            c.value     = ""
            c.font      = _font(_DARK, bold=False, sz=9)
            c.alignment = _align("center", "center", wrap=False)
    ws.row_dimensions[row].height = 20


def _build_params_sheet(wb, subject_cfg: dict, bala_yield: float = 0.06):
    """
    'Params' worksheet — holds parameters referenced by the Adj CR formula.
    B2 = Subject Remaining Leasehold (yrs) — referenced by every comp row formula.
    Bala factors are now looked up from the official Singapore Bala Table
    in the 'Bala Tbl' sheet (not calculated from a parametric yield).
    """
    ws = wb.create_sheet("Params")
    ws.sheet_view.showGridLines = False

    rows = [
        ("Parameter",                                    "Value",   None),
        ("Subject Remaining Leasehold (yrs)",
         subject_cfg["remaining_leasehold_yrs"],         '#,##0 "yrs"'),
        ("", "", None),
        ("Bala Table",
         "See 'Bala Tbl' sheet — official SG table",    None),
        ("Formula: Adj CR = FTM NOI Cap Rate × Bala(comp yrs) / Bala(subject yrs)", "", None),
        ("Source: Appendix 2, SLA/SISV Singapore Bala Table", "", None),
    ]
    for r, row_data in enumerate(rows, 1):
        a, b = row_data[0], row_data[1]
        fmt  = row_data[2] if len(row_data) > 2 else None
        ca = ws.cell(row=r, column=1, value=a)
        cb = ws.cell(row=r, column=2, value=b)
        ca.font = _font(_DARK, bold=(r == 1), sz=9)
        cb.font = _font(_DARK, bold=(r == 1), sz=9)
        ca.alignment = _align("left",  "center", wrap=False)
        cb.alignment = _align("right", "center", wrap=False)
        ca.border = _border(); cb.border = _border()
        if fmt and isinstance(b, (int, float)):
            cb.number_format = fmt

    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 36
    for r in range(1, len(rows) + 1):
        ws.row_dimensions[r].height = 18


def _build_bala_sheet(wb):
    """
    Add 'Bala Tbl' lookup sheet containing the official Singapore Bala Table.
    Column A = remaining years (1–99), Column B = factor as decimal (0.038–0.960).
    Referenced by the Adj CR / Adj Price VLOOKUP formulas in every comp row.
    Source: Appendix 2, Table Showing Leasehold Values as Percentage of Freehold Value.
    """
    if "Bala Tbl" in wb.sheetnames:
        del wb["Bala Tbl"]
    ws = wb.create_sheet("Bala Tbl")
    ws.sheet_view.showGridLines = False

    # Header
    for col, hdr in enumerate(["Years", "Factor (decimal)"], 1):
        c = ws.cell(row=1, column=col, value=hdr)
        c.font      = _font(_WHITE, bold=True, sz=9)
        c.fill      = _fill(_NAVY)
        c.alignment = _align("center", "center", wrap=False)
        c.border    = _border()

    # Data rows: years 1–99
    for row_i, (yrs, pct) in enumerate(_BALA_TABLE.items(), 2):
        ca = ws.cell(row=row_i, column=1, value=yrs)
        cb = ws.cell(row=row_i, column=2, value=round(pct / 100.0, 5))
        for c in (ca, cb):
            c.font      = _font(_DARK, bold=False, sz=9)
            c.alignment = _align("center", "center", wrap=False)
            c.border    = _border()
        ca.number_format = "0"
        cb.number_format = "0.000"

    # Source note
    note_row = len(_BALA_TABLE) + 3
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=2)
    n = ws.cell(row=note_row, column=1,
                value=("Source: Appendix 2 — Table Showing Leasehold Values as Percentage "
                       "of Freehold Value (Singapore Land Authority / SISV)"))
    n.font      = _font("FF555555", bold=False, sz=8, italic=True)
    n.alignment = _align("left", "center", wrap=True)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    for r in range(1, note_row + 1):
        ws.row_dimensions[r].height = 15


def build_workbook(subject_row: dict, comp_rows: list,
                   subject_cfg: dict, output_path: str, bala_yield: float = 0.06):
    schema = get_output_schema(subject_cfg)
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Transaction Comparables"
    ws.sheet_view.showGridLines = False

    nc        = len(schema)
    subj_yrs  = subject_cfg["remaining_leasehold_yrs"]
    deal_name = subject_cfg.get("deal_name", subject_cfg["property_name"])

    for i, (*_, w) in enumerate(schema, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Type (col A) — hidden by default (redundant with row position)
    ws.column_dimensions["A"].hidden = True
    # Source (col B) — visible; manually hide in Excel if not needed for the report
    ws.column_dimensions["B"].hidden = False
    # Hide the stake % helper column (auto-positioned after last schema col)
    ws.column_dimensions[_STAKE_COL].width  = 0
    ws.column_dimensions[_STAKE_COL].hidden = True

    # ── Title block ──────────────────────────────────────────────────────────
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    t = ws.cell(row=r, column=1,
                value=f"{deal_name.upper()}  —  TRANSACTION COMPARABLES (ASSET SALES)")
    t.fill = _fill(_NAVY); t.font = _font(_WHITE, bold=True, sz=13)
    t.alignment = _align("center", "center", wrap=False)
    ws.row_dimensions[r].height = 24

    r = 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    s = ws.cell(row=r, column=1, value="Confidential — For Discussion Purposes Only")
    s.fill = _fill(_NAVYL); s.font = _font(_WHITE, bold=False, sz=9, italic=True)
    s.alignment = _align("center", "center", wrap=False)
    ws.row_dimensions[r].height = 14

    # ── Table 1: Subject Sale ────────────────────────────────────────────────
    r = 4
    _section_header(ws, r, "  Subject Sale — Based on Underwriting", nc)
    r = 5;  _col_headers(ws, r, subj_yrs, schema=schema)
    r = 6;  _data_row(ws, r, subject_row, bold=True, is_subject=True, schema=schema)
    _write_formulas(ws, r, is_subject=True, schema=schema)

    ws.row_dimensions[7].height = 6

    # ── Table 2: Comparable Sales ────────────────────────────────────────────
    r = 8
    _section_header(ws, r, "  Comparable Asset Sales / Valuations", nc)
    r = 9;  _col_headers(ws, r, subj_yrs, schema=schema)
    for i, crow in enumerate(comp_rows):
        r = 10 + i
        _data_row(ws, r, crow, alt=(i % 2 == 1), schema=schema)
        _write_formulas(ws, r, is_subject=False, schema=schema, row_dict=crow)

    # ── Average row ───────────────────────────────────────────────────────────
    if comp_rows:
        _avg_row(ws, 10 + len(comp_rows), 10, 9 + len(comp_rows), schema=schema)

    # ── Notes footer ─────────────────────────────────────────────────────────
    r = 10 + len(comp_rows) + 2
    currency  = subject_cfg.get("currency", "SGD")
    gfa_unit  = subject_cfg.get("gfa_unit", "sf").lower()
    area_lbl  = "psm" if gfa_unit == "sqm" else "psf"
    notes = (
        f"Notes:  (1) ★ = {deal_name}; markers 1–{len(comp_rows)} = comparable transactions.  "
        f"(2) Price ({currency} {area_lbl} GFA) = implied 100%-ownership price ÷ Total GFA (live Excel formula).  "
        "(3) FTM NOI Cap Rate = Stabilised NPI Yield at time of transaction (from input file).  "
        "(4) Adj. Cap Rate uses the official Singapore Bala Table (Appendix 2, SLA/SISV).  "
        "Formula: Adj CR = FTM × Bala(comp yrs) / Bala(subject yrs).  "
        "Bala factors are looked up from the 'Bala Tbl' sheet — see Params for subject yrs.  "
        "(5) Remaining Leasehold = 0 denotes Freehold.  "
        "Source: JLL, Savills, Colliers, CBRE, local market announcements."
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    n = ws.cell(row=r, column=1, value=notes)
    n.fill = _fill(_NOTE); n.font = _font("FF404040", bold=False, sz=8, italic=True)
    n.alignment = _align("left", "top", wrap=True)
    ws.row_dimensions[r].height = 72

    # ── Params + Bala Tbl sheets ─────────────────────────────────────────────
    _build_params_sheet(wb, subject_cfg)
    _build_bala_sheet(wb)

    wb.save(output_path)
    print(f"  Saved → {output_path}")


