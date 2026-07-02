#!/usr/bin/env python3
"""
generate_global_rent_comps_table.py
=====================================
Global (non-Singapore) Rental Comparables table generator.
Same visual style as the SG rent template — navy headers, alternating fills,
subject/comp separation — but with dynamic currency/unit column headers and
without SG-specific columns.

Output Schema
-------------
  Type | Property | Map Marker | Date
  Net Rental Rate ({cur}/{area_unit}/mth) | Site Area ({area_unit})
  Location Type | Quality Type | Note
"""

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from generate_global_sales_comps_table import (
    get_deal_currency,
    get_deal_price_unit,
    get_deal_area_unit,
)


# ═══════════════════════════════════════════════════════════════════════════════
# OUTPUT SCHEMA
# ═══════════════════════════════════════════════════════════════════════════════

_SCHEMA_BASE = [
    ("Type",                         "type",         "str",   "@",       10),  # hidden
    ("Property",                     "property",     "str",   "@",       36),
    ("Map\nMarker",                  "map_marker",   "str",   "@",        8),
    ("Date",                         "lease_date",   "str",   "@",       13),
    ("Net Rental Rate\n(--/--/mth)", "rental_rate",  "float", "#,##0.00",18),
    ("Site Area\n(--)",              "site_area",    "int",   "#,##0",   14),
    ("Location\nType",               "location",     "str",   "@",       20),
    ("Quality\nType",                "quality",      "str",   "@",       20),
    ("Note",                         "remarks",      "str",   "@",       30),
]


def get_global_rent_schema(subject_cfg: dict) -> list:
    """Return schema with dynamic currency/unit column headers."""
    cur       = get_deal_currency(subject_cfg)
    area_unit = get_deal_area_unit(subject_cfg)
    schema    = list(_SCHEMA_BASE)
    for i, entry in enumerate(schema):
        if entry[1] == "rental_rate":
            schema[i] = (
                f"Net Rental Rate\n({cur}/{area_unit}/mth)",
                "rental_rate", "float", "#,##0.00", entry[4],
            )
        elif entry[1] == "site_area":
            schema[i] = (
                f"Site Area\n({area_unit})",
                "site_area", "int", "#,##0", entry[4],
            )
    return schema


# ═══════════════════════════════════════════════════════════════════════════════
# ROW CONVERTERS
# ═══════════════════════════════════════════════════════════════════════════════

def subject_to_row(cfg: dict, subject_cfg: dict) -> dict:
    rental_rate = (cfg.get("rental_rate") or cfg.get("net_rental_rate")
                   or cfg.get("rent_psf") or cfg.get("net_rent_psm"))
    site_area   = cfg.get("site_area") or cfg.get("site_area_raw") or cfg.get("gfa_sf")
    return {
        "type":        "Subject",
        "property":    cfg.get("property_name", ""),
        "map_marker":  "★",
        "lease_date":  cfg.get("lease_date", "") or cfg.get("sale_date", ""),
        "rental_rate": float(rental_rate) if rental_rate else None,
        "site_area":   int(site_area) if site_area else None,
        "location":    cfg.get("location", ""),
        "quality":     cfg.get("quality", ""),
        "remarks":     cfg.get("remarks", ""),
    }


def comp_to_row(c: dict, subject_cfg: dict) -> dict:
    name = str(c.get("property_name") or c.get("raw_description", "").split("\n")[0]).strip()
    rental_rate = (c.get("rental_rate") or c.get("net_rental_rate")
                   or c.get("rent_psf") or c.get("net_rent_psm"))
    site_area   = c.get("site_area") or c.get("site_area_raw") or c.get("gfa_sf")
    return {
        "type":        "Comparable",
        "property":    name,
        "map_marker":  str(c.get("map_marker", "")),
        "lease_date":  str(c.get("lease_date", "") or c.get("sale_date", "")),
        "rental_rate": float(rental_rate) if rental_rate else None,
        "site_area":   int(site_area) if site_area else None,
        "location":    str(c.get("location", "")),
        "quality":     str(c.get("quality", "")),
        "remarks":     str(c.get("remarks", "")),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# RENDERING HELPERS  (same palette as SG template)
# ═══════════════════════════════════════════════════════════════════════════════

_NAVY  = "FF1F3864"
_NAVYL = "FF2E4C7E"
_WHITE = "FFFFFFFF"
_LGRAY = "FFF2F2F2"
_NOTE  = "FFEBF3FB"
_DARK  = "FF1A1A1A"
_RED   = "FFC00000"
_AVGBG = "FFD6DCE4"

def _fill(h):  return PatternFill(patternType="solid", fgColor=h)
def _font(color=_WHITE, bold=True, sz=9, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=sz, italic=italic)
_T = Side(style="thin", color="FFBFBFBF")
def _border(): return Border(left=_T, right=_T, top=_T, bottom=_T)
def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _section_header(ws, row: int, label: str, nc: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row=row, column=1, value=label)
    c.fill = _fill(_NAVY); c.font = _font(_WHITE, bold=True, sz=10)
    c.alignment = _align("left", "center", wrap=False); c.border = _border()
    for col in range(2, nc + 1):
        x = ws.cell(row=row, column=col)
        x.fill = _fill(_NAVY); x.border = _border()
    ws.row_dimensions[row].height = 19


def _col_headers(ws, row: int, schema: list):
    for col, (hdr, *_) in enumerate(schema, 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.fill = _fill(_NAVY); c.font = _font(_WHITE, bold=True, sz=8.5)
        c.alignment = _align("center", "center", wrap=True); c.border = _border()
    ws.row_dimensions[row].height = 34


def _data_row(ws, row: int, row_dict: dict, schema: list,
              alt: bool = False, bold: bool = False, is_subject: bool = False):
    bg = _fill(_LGRAY) if alt else _fill(_WHITE)
    for col, entry in enumerate(schema, 1):
        key, dtype, fmt = entry[1], entry[2], entry[3]
        val     = row_dict.get(key)
        display = "—" if val is None or val == "" else val

        c = ws.cell(row=row, column=col, value=display)
        c.fill = bg; c.border = _border()
        c.font = _font(_DARK, bold=bold, sz=9)

        if col == 1:   # Type (hidden column)
            c.alignment = _align("left", "center", wrap=True)
        elif col == 2:  # Property
            c.alignment = _align("center", "center", wrap=False)
            c.font = _font(_RED if is_subject else _DARK, bold=True, sz=12)
        elif dtype in ("int", "float") and isinstance(val, (int, float)):
            c.alignment = _align("right", "center", wrap=False)
            c.number_format = fmt
        else:
            c.alignment = _align("center", "center", wrap=True)

    ws.row_dimensions[row].height = 48


def _avg_row(ws, row: int, first_r: int, last_r: int, schema: list):
    avg_keys = {"rental_rate", "site_area"}
    for col, entry in enumerate(schema, 1):
        key, dtype, fmt = entry[1], entry[2], entry[3]
        c = ws.cell(row=row, column=col)
        c.fill = _fill(_AVGBG); c.border = _border()
        if key == "property":
            c.value = "Average"
            c.font = _font(_DARK, bold=True, sz=9)
            c.alignment = _align("left", "center", wrap=False)
        elif key in avg_keys:
            col_ltr = get_column_letter(col)
            c.value = f'=IFERROR(AVERAGEIF({col_ltr}{first_r}:{col_ltr}{last_r},">0"),"—")'
            c.number_format = fmt
            c.font = _font(_DARK, bold=True, sz=9)
            c.alignment = _align("right", "center", wrap=False)
        else:
            c.value = ""; c.font = _font(_DARK, bold=False, sz=9)
            c.alignment = _align("center", "center", wrap=False)
    ws.row_dimensions[row].height = 20


# ═══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════

def build_workbook(subject_row: dict, comp_rows: list,
                   subject_cfg: dict, output_path: str):
    schema    = get_global_rent_schema(subject_cfg)
    cur       = get_deal_currency(subject_cfg)
    area_unit = get_deal_area_unit(subject_cfg)
    nc        = len(schema)
    deal_name = subject_cfg.get("deal_name", subject_cfg.get("property_name", "Deal"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rental Comparables"
    ws.sheet_view.showGridLines = False

    for i, (*_, w) in enumerate(schema, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions["A"].hidden = True   # Type column hidden

    # ── Title block ──────────────────────────────────────────────────────────
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    t = ws.cell(row=r, column=1,
                value=f"{deal_name.upper()}  —  RENTAL COMPARABLES")
    t.fill = _fill(_NAVY); t.font = _font(_WHITE, bold=True, sz=13)
    t.alignment = _align("center", "center", wrap=False)
    ws.row_dimensions[r].height = 24

    r = 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    s = ws.cell(row=r, column=1, value="Confidential — For Discussion Purposes Only")
    s.fill = _fill(_NAVYL); s.font = _font(_WHITE, bold=False, sz=9, italic=True)
    s.alignment = _align("center", "center", wrap=False)
    ws.row_dimensions[r].height = 14

    # ── Table 1: Subject Rental ───────────────────────────────────────────────
    r = 4
    _section_header(ws, r, "  Subject Property — Based on Underwriting", nc)
    r = 5; _col_headers(ws, r, schema)
    r = 6; _data_row(ws, r, subject_row, schema, bold=True, is_subject=True)

    ws.row_dimensions[7].height = 6

    # ── Table 2: Comparable Rentals ───────────────────────────────────────────
    r = 8
    _section_header(ws, r, "  Comparable Rental Transactions / Market Evidence", nc)
    r = 9; _col_headers(ws, r, schema)
    for i, crow in enumerate(comp_rows):
        r = 10 + i
        _data_row(ws, r, crow, schema, alt=(i % 2 == 1))

    if comp_rows:
        _avg_row(ws, 10 + len(comp_rows), 10, 9 + len(comp_rows), schema)

    # ── Notes footer ──────────────────────────────────────────────────────────
    r = 10 + len(comp_rows) + 2
    notes = (
        f"Notes:  (1) ★ = {deal_name}; markers 1–{len(comp_rows)} = comparable transactions.  "
        f"(2) Net Rental Rate in {cur}/{area_unit}/mth as reported in source.  "
        "Source: JLL, Savills, Colliers, CBRE, local market announcements."
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    n = ws.cell(row=r, column=1, value=notes)
    n.fill = _fill(_NOTE); n.font = _font("FF404040", bold=False, sz=8, italic=True)
    n.alignment = _align("left", "top", wrap=True)
    ws.row_dimensions[r].height = 60

    wb.save(output_path)
    print(f"  Saved → {output_path}")
