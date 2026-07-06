#!/usr/bin/env python3
"""
search_online_comps.py
======================
Searches online for comparable office investment transactions near the subject
property using OpenAI's web search capability.

Search strategy (proximity-first, sub-market fallback):
  Level 1 — Proximity  : within proximity_km of subject (default 1km)
  Level 2 — Sub-market : same location tier, within submarket_km (default 10km)
  Level 3 — Broader    : full Singapore office market, no distance cap

Each level is tried in order; the next level activates only when the current
level returns fewer than min_results confirmed transactions.

Usage
-----
    python3 search_online_comps.py                         # deal_config.json
    python3 search_online_comps.py --config my_deal.json
    python3 search_online_comps.py --map                   # also generate map PNG

Config additions (deal_config.json)
------------------------------------
    "openai": {
        "api_key": "sk-...",
        "search_model": "gpt-4o-mini-search-preview",
        "extract_model": "gpt-4o-mini"
    },
    "online_search": {
        "proximity_km":  1.0,
        "submarket_km":  10.0,
        "min_results":   3,
        "max_results":   10,
        "years_back":    2
    }

Output
------
    Online_Comparables_<DealName>.xlsx   — formatted 13-column table
    Online_Comparables_<DealName>_map.png  — Mapbox map (if --map passed)
"""

import sys as _sys
from pathlib import Path as _Path
_sys.path.insert(0, str(_Path(__file__).parent))
# ── Windows UTF-8 fix ─────────────────────────────────────────────────────────
for _stream in (_sys.stdout, _sys.stderr):
    try:
        if getattr(_stream, "encoding", "utf-8").lower().replace("-", "") != "utf8":
            _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, OSError):
        pass
import argparse
import hashlib
import json
import math
import os
import re
import subprocess
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from generate_sales_comps_table import (
    OUTPUT_SCHEMA, _STAKE_COL, bala_factor,
    _fill, _font, _border, _align,
    _section_header, _col_headers, _data_row, _write_formulas, _build_params_sheet,
    subject_to_row, get_output_schema,
)
from generate_sales_comps_map import geocode_with_fallbacks, _parse_property_text


def _shared_mapbox_token() -> str:
    """Mapbox token fallback: shared_settings.json (single source of truth) → env.
    Lets Online Search find the token set in Shared Settings / cloud secrets even
    when the deal config has no mapbox.token."""
    try:
        p = Path(__file__).parent.parent / "configs" / "shared_settings.json"
        if p.exists():
            tok = (json.loads(p.read_text(encoding="utf-8")) or {}).get("mapbox_token", "")
            if tok:
                return tok
    except Exception:
        pass
    return os.environ.get("MAPBOX_TOKEN", "") or os.environ.get("MAPBOX_API_KEY", "")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — SUB-MARKET KNOWLEDGE
#
# Maps the subject property's location tier → keyword lists used in queries.
# Add more tiers / countries here as the pipeline is used for new deals.
# ═══════════════════════════════════════════════════════════════════════════════

_SUBMARKET_KEYWORDS = {
    # ── Singapore Office ──────────────────────────────────────────────────────
    "Prime Core CBD (Marina Bay)":        ["Marina Bay", "Marina Boulevard", "Marina One"],
    "Core CBD (Raffles Place / Shenton)": ["Raffles Place", "Shenton Way", "Robinson Road", "Market Street"],
    "Core CBD (City Hall / Beach Rd)":    ["City Hall", "Beach Road", "Suntec City", "Temasek Boulevard"],
    "Orchard / City Fringe":              ["Orchard Road", "Somerset", "Penang Road", "Dhoby Ghaut"],
    "CBD Fringe (Tanjong Pagar)":         ["Tanjong Pagar", "Anson Road", "McCallum Street", "Cecil Street"],
    "Decentralised / JLD":                ["Jurong Lake District", "Jurong East", "JLD", "Gateway Drive", "Jurong Gateway"],
    "Decentralised (Paya Lebar)":         ["Paya Lebar", "Paya Lebar Quarter", "PLQ", "Geylang"],
    "Suburban":                           ["Yishun", "Woodlands", "Tampines", "Novena", "Bishan"],
    # ── South Korea Logistics ─────────────────────────────────────────────────
    "Busan Area / Gimhae City":           ["Gimhae", "Busan", "Gyeongnam", "Gyeongsangnam-do", "Sangdong"],
    "Seoul Metropolitan Logistics":       ["Seoul", "Incheon", "Gyeonggi", "Icheon", "Anseong", "Yongin"],
    "Chungcheong Logistics Hub":          ["Cheonan", "Pyeongtaek", "Chungbuk", "Ochang", "Anseong"],
    "Daegu / Gyeongbuk Logistics":        ["Daegu", "Gyeongbuk", "Gyeongsan", "Chilgok"],
    # ── Japan Logistics ───────────────────────────────────────────────────────
    "Greater Tokyo Logistics":            ["Tokyo", "Kanagawa", "Chiba", "Saitama", "Zama", "Atsugi"],
    "Greater Osaka Logistics":            ["Osaka", "Hyogo", "Amagasaki", "Itami", "Maishima"],
}

# Broader market fallback per tier
_BROADER_MARKET = {
    # Singapore Office
    "Prime Core CBD (Marina Bay)":        "Singapore Grade A office CBD investment sale",
    "Core CBD (Raffles Place / Shenton)": "Singapore Grade A office CBD investment sale",
    "Core CBD (City Hall / Beach Rd)":    "Singapore Grade A office CBD investment sale",
    "Orchard / City Fringe":              "Singapore Grade A office city fringe investment sale",
    "CBD Fringe (Tanjong Pagar)":         "Singapore office CBD fringe investment sale",
    "Decentralised / JLD":                "Singapore decentralised office investment sale",
    "Decentralised (Paya Lebar)":         "Singapore decentralised office investment sale",
    "Suburban":                           "Singapore suburban office investment sale",
    # South Korea Logistics
    "Busan Area / Gimhae City":           "South Korea Busan Gyeongnam logistics warehouse investment sale",
    "Seoul Metropolitan Logistics":       "South Korea Seoul metropolitan logistics park investment transaction",
    "Chungcheong Logistics Hub":          "South Korea logistics warehouse investment sale Chungcheong",
    "Daegu / Gyeongbuk Logistics":        "South Korea Daegu Gyeongbuk logistics investment sale",
    # Japan Logistics
    "Greater Tokyo Logistics":            "Japan Tokyo logistics warehouse investment sale J-REIT",
    "Greater Osaka Logistics":            "Japan Osaka logistics warehouse investment sale J-REIT",
}


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — QUERY BUILDING
# ═══════════════════════════════════════════════════════════════════════════════

def _year_window(years_back: int) -> str:
    """Return "2024 OR 2025 OR 2026" style string for the current + past N years."""
    now = datetime.now().year
    return " OR ".join(str(now - i) for i in range(years_back + 1))


def build_queries(subject_cfg: dict, level: str, years_back: int = 2) -> list:
    """
    Return 2–3 search queries for the given expansion level.
    level: "proximity" | "submarket" | "market"
    Fully generic: uses country_name, asset_class, currency from subject_cfg.
    """
    location     = subject_cfg.get("location", "")
    prop_name    = subject_cfg["property_name"]
    address      = subject_cfg.get("address", "")
    asset_class  = subject_cfg.get("asset_class", "office")
    currency     = subject_cfg.get("currency", "SGD")
    country_name = subject_cfg.get("country_name", "Singapore")
    yrs          = _year_window(years_back)

    # Asset keyword: config > fallback to asset_class value itself
    asset_kw = subject_cfg.get("asset_search_keyword", asset_class)

    # Submarket keywords: config > hardcoded dict (SG office legacy) > location string
    kws_cfg  = subject_cfg.get("submarket_keywords")
    kws      = kws_cfg or _SUBMARKET_KEYWORDS.get(location, [location] if location else [country_name])
    primary  = kws[0] if kws else country_name

    # Broader market query: config > hardcoded dict (SG office legacy) > generic
    broader_cfg = subject_cfg.get("broader_market_query")

    if level == "proximity":
        parts    = [p.strip() for p in address.split(",") if p.strip()]
        precinct = parts[-3] if len(parts) >= 3 else (parts[-2] if len(parts) >= 2 else primary)
        return [
            f'{country_name} "{precinct}" {asset_kw} investment sale ({yrs}) {currency}',
            f'{country_name} "{primary}" {asset_kw} transaction ({yrs}) whole block',
        ]

    elif level == "submarket":
        kw_expr = " OR ".join(f'"{k}"' for k in kws[:3])
        broader = broader_cfg or _BROADER_MARKET.get(location, f"{country_name} {asset_kw} investment")
        return [
            f'{country_name} ({kw_expr}) {asset_kw} investment sale ({yrs}) {currency}',
            f'{broader} ({yrs}) block sale OR "partial stake"',
        ]

    else:  # market
        broader = broader_cfg or _BROADER_MARKET.get(location, f"{country_name} {asset_kw} investment sale")
        return [
            f'{broader} ({yrs}) cap rate transaction {currency}',
            f'{country_name} {asset_kw} investment ({yrs}) JLL Savills CBRE Colliers {currency}',
        ]


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — OPENAI WEB SEARCH + EXTRACTION  (single call per query)
# ═══════════════════════════════════════════════════════════════════════════════

def _build_search_system(country: str = "Singapore", asset_desc: str = "commercial real estate") -> str:
    """Build a dynamic search system prompt (country + asset class aware)."""
    return (
        f"You are a {country} commercial real estate research analyst. "
        f"Use web search to find confirmed {asset_desc} investment transactions. "
        "Only extract real completed sales — not listings, asking prices, or unsold tender bids. "
        "Return your answer as a JSON object with a 'transactions' array."
    )


_EXTRACT_SYSTEM = (
    "You are a real estate data extraction assistant. "
    "Extract structured transaction data from the provided article text. "
    "Return your answer as a JSON object with a 'transactions' array."
)

_EXTRACT_PROMPT = """\
Extract all confirmed real estate investment transactions from this text:
---
{text}
---

For each transaction return a JSON object with:
  property_name, address, sale_date (e.g. "Q3 2025"), price_sgd_m (float),
  gfa_sf (int or null), remaining_yrs (int, 0=freehold, null if unknown),
  cap_rate_pct (float % or null), stake_pct (float or null — null means 100%),
  sale_type ("Block Sale"|"Partial Stake"|"En Bloc"|"Strata Sale"),
  land_zoning (or null), buyer (or null), seller (or null), country (default "Singapore")

Only include entries where property_name AND price_sgd_m are known.
Return valid JSON only: {{"transactions": [...]}}
"""


def _parse_json_from_text(text: str) -> list:
    """Extract a JSON transactions list from free-form model output."""
    text = re.sub(r"^```[a-z]*\n?", "", text.strip())
    text = re.sub(r"\n?```$", "", text)
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        # Find the first {...} block
        m = re.search(r"\{.*\}", text, re.DOTALL)
        if not m:
            return []
        try:
            parsed = json.loads(m.group())
        except json.JSONDecodeError:
            return []
    if isinstance(parsed, dict):
        parsed = parsed.get("transactions") or next(
            (v for v in parsed.values() if isinstance(v, list)), []
        )
    return parsed if isinstance(parsed, list) else []


def search_and_extract(query: str, client,
                       search_model: str  = "gpt-4o-mini-search-preview",
                       extract_model: str = "gpt-4o-mini",
                       subject_cfg: dict  = None) -> tuple:
    """
    Two-step pipeline.
    Returns (transactions: list, sources: list[dict])
      sources = [{"title": "...", "url": "..."}] extracted from response annotations.
    subject_cfg is used to tailor the prompts for currency, gfa_unit, asset_class, country.
    """
    cfg         = subject_cfg or {}
    currency    = cfg.get("currency", "SGD")
    gfa_unit    = cfg.get("gfa_unit", "sf").lower()
    country     = cfg.get("country_name", "Singapore")
    asset_class = cfg.get("asset_class", "office")
    gfa_desc    = "sqm" if gfa_unit == "sqm" else "sq ft"
    asset_desc  = {"logistics": "logistics / warehouse", "office": "office",
                   "industrial": "industrial"}.get(asset_class, asset_class)

    # Step 1: web search → raw article text + URL citations
    search_resp = client.chat.completions.create(
        model=search_model,
        messages=[
            {"role": "system", "content":
             f"You are a {country} commercial real estate research analyst. "
             f"Search the web and summarise all confirmed {asset_desc} investment "
             f"transactions you find. Include property name, address, sale date, "
             f"price in {currency} millions, GFA in {gfa_desc}, remaining leasehold years, "
             "NPI/NOI yield, stake sold, buyer, and seller for each transaction. "
             "Be thorough — list every transaction found."},
            {"role": "user", "content": f"Search for: {query}"},
        ],
    )
    msg          = search_resp.choices[0].message
    article_text = msg.content or ""

    # Capture URL citations from annotations (gpt-4o-mini-search-preview feature)
    sources = []
    for ann in getattr(msg, "annotations", None) or []:
        if getattr(ann, "type", "") == "url_citation":
            uc = getattr(ann, "url_citation", None)
            if uc:
                sources.append({
                    "title": getattr(uc, "title", "") or "",
                    "url":   getattr(uc, "url",   "") or "",
                })
    # Deduplicate by URL
    seen_urls = set()
    sources = [s for s in sources if s["url"] not in seen_urls
               and not seen_urls.add(s["url"])]

    if not article_text or len(article_text) < 50:
        return [], sources

    # Step 2: extract structured JSON from article text
    extract_prompt = (
        f"Extract all confirmed {asset_desc} investment transactions from this text:\n"
        f"---\n{{text}}\n---\n\n"
        f"IMPORTANT: Only include {asset_desc} properties. "
        f"Exclude residential, hotel, serviced apartment, and other non-{asset_desc} assets.\n\n"
        f"For each transaction return a JSON object with:\n"
        f"  property_name, address, sale_date (e.g. 'Q3 2025'), "
        f"price_sgd_m (price in {currency} millions, float),\n"
        f"  gfa_sf (GFA in {gfa_desc}, int or null), remaining_yrs (int, 0=freehold, null if unknown),\n"
        f"  cap_rate_pct (float % or null), stake_pct (float or null — null means 100%%),\n"
        f"  sale_type ('Block Sale'|'Partial Stake'|'En Bloc'|'Strata Sale'),\n"
        f"  asset_type (e.g. 'Block Sale ({asset_desc.title()})'),\n"
        f"  land_zoning (or null), buyer (or null), seller (or null), "
        f"country (default '{country}')\n\n"
        f"Only include entries where property_name AND price_sgd_m are known.\n"
        f"Return valid JSON only: {{{{\"transactions\": [...]}}}}"
    )
    extract_resp = client.chat.completions.create(
        model=extract_model,
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content":
             "You are a real estate data extraction assistant. "
             "Extract structured transaction data from the provided article text. "
             "Return your answer as a JSON object with a 'transactions' array."},
            {"role": "user", "content": extract_prompt.format(text=article_text[:4000])},
        ],
    )
    return _parse_json_from_text(extract_resp.choices[0].message.content), sources


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — VALIDATE, DEDUP, DISTANCE FILTER
# ═══════════════════════════════════════════════════════════════════════════════

# ── Asset-type guard ──────────────────────────────────────────────────────────
# Confirm: keywords that positively identify a particular asset class.
_ASSET_ACCEPT_KWS = {
    "office":      ["office"],
    "logistics":   ["logistic", "warehouse", "cold storage", "distribution centre",
                    "distribution center", "fulfillment", "fulfilment"],
    "industrial":  ["industrial", "business park", "factory"],
    "retail":      ["retail", "shophouse", "mall"],
    "residential": ["residential", "apartment", "condominium", "condo", "landed"],
}
# Reject: keywords that clearly indicate a DIFFERENT, incompatible asset class.
# Only applied when no accept keyword matched (so "Mixed Use (Office + Retail)"
# stays accepted for an office subject because "office" is in accept first).
_ASSET_REJECT_KWS = {
    "office":      ["residential", "apartment", "condominium", "condo",
                    "hotel", "serviced apartment", "hospitality", "columbarium"],
    "logistics":   ["residential", "apartment", "condominium", "condo",
                    "hotel", "serviced apartment", "office"],
    "industrial":  ["residential", "apartment", "condominium", "condo",
                    "hotel", "serviced apartment"],
    "retail":      ["residential", "apartment", "condominium", "condo",
                    "hotel", "serviced apartment",
                    "logistics park", "warehouse", "industrial estate"],
    "residential": ["office", "logistics park", "logistics centre",
                    "warehouse", "industrial estate"],
}


def _asset_type_matches(record: dict, asset_class: str) -> bool:
    """
    Return True if the record's property type is compatible with asset_class.
    Conservative — only rejects clear mismatches; keeps ambiguous records.

    Inspects: land_zoning, sale_type, asset_type, lease_type, property_type, comment.
    Logic:
      1. If any ACCEPT keyword is found → True (handles "Mixed Use (Office + Retail)")
      2. If any REJECT keyword is found (and no accept match) → False
      3. Otherwise (no type info, or ambiguous) → True
    """
    if not asset_class:
        return True

    combined = " ".join(str(record.get(f) or "") for f in (
        "land_zoning", "sale_type", "asset_type",
        "lease_type", "property_type", "comment",
    )).lower()

    if not combined.strip():
        return True  # no type information — keep

    ac = asset_class.lower()
    accept_kws = _ASSET_ACCEPT_KWS.get(ac, [])
    reject_kws = _ASSET_REJECT_KWS.get(ac, [])

    if any(kw in combined for kw in accept_kws):
        return True   # explicit match → accept

    if any(kw in combined for kw in reject_kws):
        return False  # explicit mismatch (no accept keyword saved it) → reject

    return True       # ambiguous → keep


def _haversine_km(lon1: float, lat1: float, lon2: float, lat2: float) -> float:
    R = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2
         + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2))
         * math.sin(dlon / 2) ** 2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def validate_dedup(records: list, subject_name: str = "",
                   subject_country: str = "",
                   subject_asset_class: str = "") -> list:
    """
    Keep only confirmed investment transactions matching the subject country and asset class.
    Drop strata sales, wrong-type assets, non-matching country, duplicates, and subject itself.

    Dedup strategy (catches bilingual duplicates like Korean + English same property):
      1. Name-based key  : normalized ASCII name[:24] + price → existing approach
      2. Coordinate key  : (lon rounded 2dp, lat rounded 2dp) + price within 1%
         → catches same property reported in two languages / by two sources
    """
    seen_name_keys  = set()
    seen_coord_keys = set()   # (lon2, lat2, price_bucket)
    subject_key  = re.sub(r"\W+", "", subject_name.lower())[:24] if subject_name else ""
    sc_tokens    = subject_country.lower().split() if subject_country else []
    out          = []
    for r in records:
        name  = str(r.get("property_name") or "").strip()
        price = r.get("price_sgd_m")
        if not name or not price:
            continue
        # Country filter
        rec_country = str(r.get("country") or "").lower().strip()
        if rec_country and sc_tokens:
            if not any(tok in rec_country for tok in sc_tokens):
                continue
        stype = str(r.get("sale_type") or "")
        if "Strata" in stype:
            continue
        # Asset-type filter: skip if clearly a different property type
        if subject_asset_class and not _asset_type_matches(r, subject_asset_class):
            print(f"    [type-filter] skipped '{name}' "
                  f"(not {subject_asset_class})")
            continue
        # Skip subject property itself
        norm_key = re.sub(r"\W+", "", name.lower())[:24]
        if subject_key and subject_key in norm_key:
            continue
        # Dedup 1: name-based
        name_key = norm_key + f"_{float(price):.0f}"
        if name_key in seen_name_keys:
            continue
        # Dedup 2: coordinate-based (catches bilingual dupes with same location + price)
        lon, lat = r.get("lon"), r.get("lat")
        if lon is not None and lat is not None:
            price_bucket = round(float(price) / max(float(price) * 0.05, 1))
            coord_key    = (round(lon, 2), round(lat, 2), price_bucket)
            if coord_key in seen_coord_keys:
                continue
            seen_coord_keys.add(coord_key)
        seen_name_keys.add(name_key)
        out.append(r)
    return out


def geocode_records(records: list, mapbox_token: str,
                    country_code: str = "", country_name: str = "") -> list:
    """Add lon/lat to each record. Only geocodes records that have a real address.
    Property names and descriptions are NOT used as geocoding queries — they
    produce unreliable or wrong coordinates."""
    suffix = f", {country_name}" if country_name else ""
    out = []
    for r in records:
        name = str(r.get("property_name") or "")
        addr = str(r.get("address") or "").strip()
        if not addr:
            out.append({**r, "lon": None, "lat": None})
            continue
        query = f"{addr}{suffix}" if suffix not in addr else addr
        try:
            lon, lat, _ = geocode_with_fallbacks([query], mapbox_token, country_code)
            out.append({**r, "lon": lon, "lat": lat})
        except Exception:
            out.append({**r, "lon": None, "lat": None})
    return out


def filter_by_distance(records: list,
                        subject_lon: float, subject_lat: float,
                        max_km: float) -> list:
    """Keep records where (lon, lat) is within max_km of subject."""
    return [
        r for r in records
        if r.get("lon") is not None
        and _haversine_km(r["lon"], r["lat"], subject_lon, subject_lat) <= max_km
    ]


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — CLASSIFY  (Location / Quality / Asset Type via OpenAI)
# ═══════════════════════════════════════════════════════════════════════════════

def classify_records(records: list, subject_cfg: dict,
                     client, extract_model: str = "gpt-4o-mini") -> list:
    if not records:
        return []
    country    = subject_cfg.get("country_name", "Singapore")
    asset_cls  = subject_cfg.get("asset_class", "office")
    qual_opts  = (
        '"Grade A+" | "Grade A" | "Grade B" | "Grade A (Mixed-Use)"'
        if asset_cls == "office" else
        '"Grade A (Modern)" | "Grade B (Older)" | "Cold Storage" | "Dry Warehouse"'
        if asset_cls == "logistics" else
        '"Grade A" | "Grade B" | "Grade C"'
    )
    slim = [{"index": i,
             "property_name": r.get("property_name"),
             "address":       r.get("address"),
             "sale_type":     r.get("sale_type"),
             "gfa_sf":        r.get("gfa_sf"),
             "remaining_yrs": r.get("remaining_yrs")}
            for i, r in enumerate(records)]
    classify_system = (
        f"You are a senior {country} commercial real estate analyst. "
        "Return your answer as a JSON object."
    )
    classify_prompt = (
        f"Subject property:\n{json.dumps(subject_cfg, indent=2)}\n\n"
        f"For each comparable below assign:\n"
        f"  location        — market sub-region / location of the property\n"
        f"  quality         — {qual_opts}\n"
        f"  asset_type      — \"<Sale Structure> (<Use>)\" e.g. \"Block Sale ({asset_cls.title()})\"\n"
        f"  relevance_score — 1–10 vs the subject (same sub-market + GFA + tenure = higher)\n\n"
        f"Comparables:\n{json.dumps(slim, indent=2)}\n\n"
        'Return valid JSON only: {"results": [{"index":0,"location":"...","quality":"...","asset_type":"...","relevance_score":5},...]}'
    )
    resp = client.chat.completions.create(
        model=extract_model,
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": classify_system},
            {"role": "user",   "content": classify_prompt},
        ],
    )
    parsed = json.loads(resp.choices[0].message.content)
    cls_list = parsed.get("results") or next(
        (v for v in parsed.values() if isinstance(v, list)), []
    )
    out = [dict(r) for r in records]
    for cls in cls_list:
        idx = cls.get("index")
        if idx is not None and idx < len(out):
            out[idx].update({
                "location":        cls.get("location", ""),
                "quality":         cls.get("quality", ""),
                "asset_type":      cls.get("asset_type", ""),
                "relevance_score": int(cls.get("relevance_score") or 5),
            })
    out.sort(key=lambda x: -x.get("relevance_score", 0))
    for i, r in enumerate(out, 1):
        r["map_marker"] = str(i)
    return out


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — CONVERT TO OUTPUT ROW DICT
# ═══════════════════════════════════════════════════════════════════════════════

def record_to_row(r: dict, subject_cfg: dict, bala_yield: float) -> dict:
    subj_yrs = subject_cfg["remaining_leasehold_yrs"]
    price_m  = float(r.get("price_sgd_m") or 0)
    stake    = float(r.get("stake_pct") or 100) / 100.0
    gfa      = int(r.get("gfa_sf") or 0)
    rem_yrs  = int(r.get("remaining_yrs") or 0)
    cr_raw   = r.get("cap_rate_pct")
    cap_rate = float(cr_raw) / 100.0 if cr_raw is not None else None

    psf    = round((price_m / stake) * 1e6 / gfa) if gfa and price_m else None
    adj_cr = (cap_rate * bala_factor(rem_yrs, bala_yield) / bala_factor(subj_yrs, bala_yield)
              if cap_rate else None)

    prop_name = str(r.get("property_name") or "")
    address   = str(r.get("address")       or "")
    prop_text = f"{prop_name}\n{address}" if address else prop_name

    stype     = str(r.get("sale_type")   or "Block Sale")
    use       = str(r.get("asset_type")  or f"{stype} (Office)")

    return {
        "property":      prop_text,
        "map_marker":    str(r.get("map_marker", "")),
        "sale_date":     str(r.get("sale_date") or ""),
        "land_zoning":   str(r.get("land_zoning") or subject_cfg.get("land_zoning") or ""),
        "remaining_yrs": rem_yrs,
        "gfa_sf":        gfa      if gfa      else None,
        "price_sgd_m":   price_m  if price_m  else None,
        "price_psf_gfa": psf      if psf      else None,
        "ftm_cap_rate":  cap_rate,
        "adj_cap_rate":  adj_cr,
        "location":      str(r.get("location") or ""),
        "quality":       str(r.get("quality")  or ""),
        "asset_type":    use,
        "stake_pct":     stake,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 7 — EXCEL OUTPUT  (same visual style as main pipeline)
# ═══════════════════════════════════════════════════════════════════════════════

_NAVY  = "FF1F3864"
_NAVYL = "FF2E4C7E"
_NOTE  = "FFEBF3FB"


def _build_sources_sheet(wb, classified: list):
    """
    Add a 'Sources' sheet listing each comp's verification URLs.
    classified : the classified records list (each has a 'sources' key).
    """
    ws = wb.create_sheet("Sources")
    ws.sheet_view.showGridLines = False

    NAVY  = "FF1F3864"
    WHITE = "FFFFFFFF"
    LGRAY = "FFF2F2F2"

    # Header
    headers = ["#", "Property", "Source Title", "URL"]
    widths  = [4, 38, 40, 70]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(col)].width = w
        c = ws.cell(row=1, column=col, value=h)
        c.fill = _fill(NAVY)
        c.font = _font(WHITE, bold=True, sz=10)
        c.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 18

    row = 2
    for rec in classified:
        marker = str(rec.get("map_marker", ""))
        name   = str(rec.get("property_name", ""))
        srcs   = rec.get("sources") or []
        if not srcs:
            # No sources captured — show placeholder
            srcs = [{"title": "No URL captured (try --refresh)", "url": ""}]
        for i, s in enumerate(srcs):
            alt = (row % 2 == 0)
            bg  = LGRAY if alt else "FFFFFFFF"
            ws.cell(row=row, column=1, value=marker if i == 0 else "").fill = _fill(bg)
            ws.cell(row=row, column=1).alignment = _align("center", "center")
            ws.cell(row=row, column=2, value=name if i == 0 else "").fill = _fill(bg)
            title_cell = ws.cell(row=row, column=3, value=s.get("title") or "")
            title_cell.fill = _fill(bg)
            url = s.get("url") or ""
            url_cell = ws.cell(row=row, column=4, value=url)
            url_cell.fill = _fill(bg)
            if url:
                url_cell.hyperlink = url
                url_cell.font = _font("FF1155CC", sz=9, bold=False)
                url_cell.style = "Hyperlink"
            for col in range(1, 5):
                ws.cell(row=row, column=col).alignment = _align("left", "center", wrap=False)
            row += 1


def build_workbook_online(subject_row: dict, comp_rows: list,
                          subject_cfg: dict, output_path: str,
                          bala_yield: float, search_levels_used: list,
                          classified: list = None):
    schema   = get_output_schema(subject_cfg)
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Online Comparables"
    ws.sheet_view.showGridLines = False

    nc       = len(schema)
    subj_yrs = subject_cfg["remaining_leasehold_yrs"]
    deal     = subject_cfg.get("deal_name", subject_cfg["property_name"])
    levels   = " + ".join(search_levels_used)

    for i, (*_, w) in enumerate(schema, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions[_STAKE_COL].width  = 0
    ws.column_dimensions[_STAKE_COL].hidden = True

    # ── Title block ──────────────────────────────────────────────────────────
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    t = ws.cell(row=r, column=1,
                value=f"{deal.upper()}  —  ONLINE TRANSACTION COMPARABLES (ASSET SALES)")
    t.fill = _fill(_NAVY); t.font = _font("FFFFFFFF", bold=True, sz=13)
    t.alignment = _align("center", "center", wrap=False)
    ws.row_dimensions[r].height = 24

    r = 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    s = ws.cell(row=r, column=1,
                value=f"AI web search: {levels}  |  Verify all figures before use  |  Confidential — For Discussion Purposes Only")
    s.fill = _fill(_NAVYL); s.font = _font("FFFFFFFF", bold=False, sz=9, italic=True)
    s.alignment = _align("center", "center", wrap=False)
    ws.row_dimensions[r].height = 14

    # ── Table 1: Subject ─────────────────────────────────────────────────────
    r = 4;  _section_header(ws, r, "  Subject Sale — Based on Underwriting", nc)
    r = 5;  _col_headers(ws, r, subj_yrs, schema=schema)
    r = 6;  _data_row(ws, r, subject_row, bold=True, schema=schema)
    _write_formulas(ws, r, is_subject=True, schema=schema)
    ws.row_dimensions[7].height = 6

    # ── Table 2: Online Comps ─────────────────────────────────────────────────
    r = 8;  _section_header(ws, r, "  Online Comparable Asset Sales (AI-Sourced)", nc)
    r = 9;  _col_headers(ws, r, subj_yrs, schema=schema)
    for i, crow in enumerate(comp_rows):
        r = 10 + i
        _data_row(ws, r, crow, alt=(i % 2 == 1), schema=schema)
        _write_formulas(ws, r, is_subject=False, schema=schema)

    # ── Notes footer ─────────────────────────────────────────────────────────
    r = 10 + len(comp_rows) + 1
    currency = subject_cfg.get("currency", "SGD")
    gfa_unit = subject_cfg.get("gfa_unit", "sf").lower()
    area_lbl = "psm" if gfa_unit == "sqm" else "psf"
    notes = (
        "Notes:  Data sourced via OpenAI web search and extracted by AI — "
        "verify all figures against primary sources (JLL, Savills, CBRE, local market "
        "announcements) before use in formal documents.  "
        f"Price per area shown as {currency} {area_lbl} GFA.  "
        "Adj. Cap Rate uses Bala Table; see 'Params' sheet for parameters.  "
        f"Search levels applied: {levels}."
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    n = ws.cell(row=r, column=1, value=notes)
    n.fill = _fill(_NOTE); n.font = _font("FF404040", bold=False, sz=8, italic=True)
    n.alignment = _align("left", "top", wrap=True)
    ws.row_dimensions[r].height = 56

    _build_params_sheet(wb, subject_cfg, bala_yield)
    if classified:
        _build_sources_sheet(wb, classified)
    wb.save(output_path)
    print(f"  Saved → {output_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 8 — SEARCH CACHE
#
# Saves raw geocoded records to  <deal>_search_cache.json  after the first run.
# Subsequent runs load from cache (stable, reproducible results).
# Pass  --refresh  on the CLI to discard the cache and re-search.
#
# Cache is automatically invalidated when subject_property or online_search
# parameters change (detected via a hash of those config sections).
# ═══════════════════════════════════════════════════════════════════════════════

def _cache_key(subject_cfg: dict, sc_cfg: dict) -> str:
    """Short hash of the config sections that affect search results."""
    blob = json.dumps({"s": subject_cfg, "p": sc_cfg}, sort_keys=True)
    return hashlib.md5(blob.encode()).hexdigest()[:12]


def load_cache(cache_path: str, expected_key: str) -> tuple:
    """
    Return (records, levels_used) if a valid cache exists, else (None, None).
    Cache is invalid if the config key has changed.
    Sources are embedded in each record's 'sources' field.
    """
    try:
        with open(cache_path, encoding="utf-8") as f:
            data = json.load(f)
        if data.get("cache_key") != expected_key:
            print("  Cache key mismatch (config changed) — running fresh search.")
            return None, None
        ts = data.get("timestamp", "unknown")
        print(f"  Loaded from cache (saved {ts}). Pass --refresh to re-search.")
        return data["records"], data["levels_used"]
    except FileNotFoundError:
        return None, None


def save_cache(cache_path: str, records: list, levels_used: list, cache_key: str):
    data = {
        "cache_key":   cache_key,
        "timestamp":   datetime.now().strftime("%Y-%m-%d %H:%M"),
        "levels_used": levels_used,
        "records":     records,
    }
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)
    print(f"  Cache saved → {cache_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 9 — MAIN  (proximity → sub-market → market expansion)
# ═══════════════════════════════════════════════════════════════════════════════

def run(config_path: str = "configs/deal_config.json", generate_map: bool = False,
        refresh: bool = False):
    with open(config_path, encoding="utf-8") as f:
        cfg = json.load(f)

    subject_cfg = cfg["subject_property"]
    params      = cfg.get("parameters", {})
    bala_yield  = params.get("bala_yield", 0.06)
    mb_cfg      = cfg.get("mapbox", {})
    mapbox_tok  = mb_cfg.get("token", "") or _shared_mapbox_token()
    oa_cfg      = cfg.get("openai", {})
    api_key      = oa_cfg.get("api_key") or os.environ.get("OPENAI_API_KEY", "")
    search_model  = oa_cfg.get("search_model",  "gpt-4o-mini-search-preview")
    extract_model = oa_cfg.get("extract_model", "gpt-4o-mini")

    sc_cfg          = cfg.get("online_search", {})
    proximity_km    = sc_cfg.get("proximity_km",    1.0)
    submarket_km    = sc_cfg.get("submarket_km",    3.0)
    market_km       = sc_cfg.get("market_km",       submarket_km)  # Level 3 cap; defaults to submarket_km
    min_results     = sc_cfg.get("min_results",     3)
    max_results     = sc_cfg.get("max_results",     10)
    years_back      = sc_cfg.get("years_back",      2)
    years_back_max  = sc_cfg.get("years_back_max",  8)
    years_back_step = sc_cfg.get("years_back_step", 2)
    max_level       = sc_cfg.get("max_level",       3)   # 1=proximity only, 2=+submarket, 3=+broader
    # Grounded data sources to combine with (or instead of) OpenAI web search.
    # Defaults to web-search only → identical behaviour to before.
    sources_cfg     = sc_cfg.get("sources") or ["web_search"]

    if not api_key:
        raise ValueError(
            "OpenAI API key not found.\n"
            "  Set openai.api_key in deal_config.json  OR  export OPENAI_API_KEY=sk-..."
        )
    if not mapbox_tok:
        raise ValueError("No Mapbox token found (needed for geocoding). Set it in "
                         "Shared Settings / MAPBOX_TOKEN secret, or mapbox.token in the deal config.")

    try:
        from openai import OpenAI
    except ImportError:
        raise ImportError("Run: pip3 install openai")

    client    = OpenAI(api_key=api_key)
    deal_name = subject_cfg.get("deal_name", subject_cfg["property_name"])
    deal_slug = deal_name.replace(' ', '_')

    # Derive output directory from output_file path (supports subfolders)
    output_file = cfg.get("output_file", f"Transaction_Comparables_{deal_slug}.xlsx")
    out_dir     = Path(output_file).parent
    out_dir.mkdir(parents=True, exist_ok=True)

    out_excel  = str(out_dir / f"Online_Comparables_{deal_slug}.xlsx")
    out_map    = str(out_dir / f"Online_Comparables_{deal_slug}_map.png")
    cache_path = str(out_dir / f"Online_Comparables_{deal_slug}_search_cache.json")

    print(f"\n{'='*62}\n  Online Comps Search : {deal_name}\n{'='*62}")

    # ── Geocode subject ───────────────────────────────────────────────────────
    address      = subject_cfg.get("address", "")
    prop_name    = subject_cfg["property_name"]
    country_name = subject_cfg.get("country_name", "")
    # country_code must be set explicitly in config (e.g. "sg", "kr", "jp").
    # No address-sniffing heuristic — if missing, geocoding proceeds without country filter.
    country_code = cfg.get("country_code", "")

    print(f"\n[0/5] Geocoding subject property")
    s_lon, s_lat, _ = geocode_with_fallbacks(
        [f"{prop_name}, {address}", address, prop_name],
        mapbox_tok, country_code,
    )
    print(f"      {prop_name} → ({s_lon}, {s_lat})")

    # ── Search loop (proximity → submarket → market), with caching ───────────
    c_key = _cache_key(subject_cfg, sc_cfg)

    if not refresh:
        print(f"\n[Search] Checking cache…")
        cached_records, cached_levels = load_cache(cache_path, c_key)
    else:
        print(f"\n[Search] --refresh: discarding any existing cache.")
        cached_records, cached_levels = None, None

    if cached_records is not None:
        all_records = cached_records
        levels_used = cached_levels
    else:
        GEO_LEVELS = [
            ("proximity", proximity_km, f"Proximity ≤{proximity_km}km"),
            ("submarket", submarket_km, f"Sub-market ≤{submarket_km}km"),
        ]
        # Level 3 uses market_km (defaults to submarket_km if not set in config)
        BROAD_LEVEL = ("market", market_km, f"Broader market ≤{market_km}km")

        all_records:     list = []
        levels_used:     list = []
        seen_keys:       dict = {}   # name+price key → index in all_records
        seen_coord_keys: set  = set()  # (lon2dp, lat2dp, price_bucket) — bilingual dedup

        yrs = years_back   # will increment on temporal expansion

        def _merge_geocoded(geocoded, srcs, max_km, source_name=""):
            """Distance-filter + (coordinate & name) dedup-merge a batch of geocoded
            records into all_records; returns count added. Shared by web search and
            grounded connectors so every source is treated identically."""
            added = 0
            srcs  = srcs or []
            for r in geocoded:
                price = float(r.get('price_sgd_m') or 0)
                key   = re.sub(r"\W+", "", str(r.get("property_name", "")).lower())[:24] \
                        + f"_{price:.0f}"
                if max_km is not None and r.get("lon") is not None:
                    if _haversine_km(r["lon"], r["lat"], s_lon, s_lat) > max_km:
                        continue
                _srcs = [({**s, "source_name": source_name} if source_name else s) for s in srcs]
                if r.get("lon") is not None and price > 0:
                    p_bucket  = round(price / max(price * 0.05, 0.5))
                    coord_key = (round(r["lon"], 2), round(r["lat"], 2), p_bucket)
                    if coord_key in seen_coord_keys:
                        if key in seen_keys:
                            idx = seen_keys[key]
                            existing = {s.get("url") for s in all_records[idx].get("sources", [])}
                            for s in _srcs:
                                if s.get("url") and s["url"] not in existing:
                                    all_records[idx].setdefault("sources", []).append(s)
                        continue
                    seen_coord_keys.add(coord_key)
                if key in seen_keys:
                    idx = seen_keys[key]
                    existing = {s.get("url") for s in all_records[idx].get("sources", [])}
                    for s in _srcs:
                        if s.get("url") and s["url"] not in existing:
                            all_records[idx].setdefault("sources", []).append(s)
                    continue
                r["sources"] = list(_srcs)
                seen_keys[key] = len(all_records)
                all_records.append(r)
                added += 1
            return added

        def _run_level(level_id, max_km, level_label, yrs_used):
            """Run one search level; return count of new records added."""
            queries   = build_search_queries_for_level(subject_cfg, level_id, yrs_used)
            level_new = 0
            print(f"\n[Search] {level_label}  (years back: {yrs_used})")
            for q in queries:
                print(f"  Query: {q[:80]}…" if len(q) > 80 else f"  Query: {q}")
                try:
                    raw, q_sources = search_and_extract(q, client, search_model, extract_model,
                                                        subject_cfg=subject_cfg)
                except Exception as e:
                    print(f"  ✗ {e.__class__.__name__}: {e}")
                    continue
                if q_sources:
                    print(f"  ✦ {len(q_sources)} source(s) found")
                cleaned  = validate_dedup(raw, subject_name=prop_name,
                                          subject_country=country_name,
                                          subject_asset_class=subject_cfg.get("asset_class", ""))
                geocoded = geocode_records(cleaned, mapbox_tok, country_code,
                                           country_name=country_name)
                level_new += _merge_geocoded(geocoded, q_sources, max_km)
            label_str = f"{level_label} (yrs:{yrs_used})"
            if label_str not in levels_used:
                levels_used.append(label_str)
            print(f"  → {level_new} new  |  total: {len(all_records)}")
            return level_new

        if "web_search" in sources_cfg:
            # ── Phase 1: geo levels (proximity + submarket) with temporal expansion ──
            active_levels = GEO_LEVELS[:max_level]
            if max_level < 3:
                print(f"  (Level 3 'Broader market' disabled — max_level={max_level})")

            while len(all_records) < min_results:
                for level_id, max_km, level_label in active_levels:
                    _run_level(level_id, max_km, level_label, yrs)
                    if len(all_records) >= min_results:
                        break

                if len(all_records) >= min_results:
                    print(f"  ✓ min_results ({min_results}) reached.")
                    break

                # Temporal expansion
                if yrs < years_back_max:
                    yrs += years_back_step
                    print(f"\n  ↩  Sparse results ({len(all_records)}). "
                          f"Extending search window to {yrs} years back…")
                else:
                    print(f"\n  ↩  Temporal limit ({years_back_max} yrs) reached. "
                          f"{len(all_records)} record(s) found in geo levels.")
                    break

            # ── Phase 2: broader market (only if max_level=3 and still sparse) ───
            if len(all_records) < min_results and max_level >= 3:
                print(f"\n  Falling back to broader market search (geo levels exhausted).")
                _run_level(*BROAD_LEVEL, yrs)
                if len(all_records) < min_results:
                    print(f"  ⚠  All levels exhausted — {len(all_records)} record(s) total.")

        # ── Grounded connectors (URA PMI, …): fetch once, ingest via same pipeline ──
        from sources.registry import get_grounded
        _conn_params = {"country_code": country_code, "country_name": country_name,
                        "years_back": yrs, "s_lon": s_lon, "s_lat": s_lat,
                        "proximity_km": proximity_km, "submarket_km": submarket_km,
                        "market_km": market_km, "comp_type": "sales",
                        "client": client, "extract_model": extract_model,
                        "ura_max_rows": sc_cfg.get("ura_max_rows", 60),
                        "broker_pages": sc_cfg.get("broker_pages"),
                        "broker_max_pdfs": sc_cfg.get("broker_max_pdfs", 4)}
        for _conn in get_grounded((country_code or "sg").lower(), "sales",
                                  sources_cfg, _conn_params):
            print(f"\n[Source] {_conn.label or _conn.name}")
            try:
                _raw, _srcs = _conn.fetch(subject_cfg, _conn_params)
            except Exception as e:
                print(f"  ✗ {e.__class__.__name__}: {e}")
                continue
            if not _raw:
                print("  → 0 records")
                continue
            _cleaned = validate_dedup(_raw, subject_name=prop_name,
                                      subject_country=country_name,
                                      subject_asset_class=subject_cfg.get("asset_class", ""))
            # ── Comparability rules on grounded records ──────────────────────
            # Recency: keep only recent sales (default 12 months; set recency_months
            # in the search config to tighten, e.g. 6). Unparseable dates are kept.
            from sources.base import months_ago as _months_ago
            _rec_m = int(sc_cfg.get("recency_months", 12) or 12)
            _before = len(_cleaned)
            _cleaned = [r for r in _cleaned
                        if (_months_ago(str(r.get("sale_date") or "")) or 0) <= _rec_m]
            if _before != len(_cleaned):
                print(f"  · recency ≤{_rec_m}mo: kept {len(_cleaned)}/{_before}")
            _geo = geocode_records(_cleaned, mapbox_tok, country_code,
                                   country_name=country_name)
            # Location: keep only same sub-market (tight comp radius, not city-wide).
            _added = _merge_geocoded(
                _geo, _srcs or [{"title": _conn.label or _conn.name, "url": ""}],
                submarket_km, source_name=_conn.name)
            _lbl = _conn.label or _conn.name
            if _lbl not in levels_used:
                levels_used.append(_lbl)
            print(f"  → {_added} new  |  total: {len(all_records)}")

        save_cache(cache_path, all_records, levels_used, c_key)

    # ── Trim to max_results ───────────────────────────────────────────────────
    records = all_records[:max_results]
    print(f"\n[1/5] SEARCH complete — {len(records)} record(s) for processing")

    if not records:
        print("\n  No comparable transactions found. "
              "Consider widening years_back or checking your OpenAI model supports web search.\n")
        return

    # ── Classify ──────────────────────────────────────────────────────────────
    print(f"\n[2/5] CLASSIFY  ({extract_model})")
    classified = classify_records(records, subject_cfg, client, extract_model)
    print(f"      → {len(classified)} records classified")

    # ── Location competitiveness (SG, URA proximity vs subject) ───────────────
    # Overrides the LLM free-text location with Superior/Comparable/Inferior using
    # the SAME logic as the internal pipeline (tools.location_score), reusing the
    # map-resolved lon/lat already on each comp + the subject. SG-only; others blank.
    try:
        from tools.location_score import apply_location as _apply_loc
        classified = _apply_loc(classified,
                                subject_cfg.get("property_name", ""),
                                subject_cfg.get("address", ""),
                                subject_cfg.get("asset_class", ""),
                                subj_lonlat=(s_lon, s_lat))
    except Exception as _le:
        print(f"  [location] skipped: {_le}")

    # ── Compute metrics ───────────────────────────────────────────────────────
    print(f"\n[3/5] CALCULATE  (Bala y = {bala_yield*100:.1f}%)")
    subj_row  = subject_to_row(subject_cfg)
    comp_rows = [record_to_row(r, subject_cfg, bala_yield) for r in classified]

    print(f"\n  {'#':<3} {'Property':<42} {'km':>5} {'Yrs':>4} {'FTM':>7} {'AdjCR':>7}")
    print("  " + "─" * 70)
    for r, crow in zip(classified, comp_rows):
        km   = _haversine_km(r["lon"], r["lat"], s_lon, s_lat) if r.get("lon") else 0
        ftm  = float(crow.get("ftm_cap_rate") or 0) * 100
        adj  = float(crow.get("adj_cap_rate") or 0) * 100
        name = str(r.get("property_name", ""))[:40]
        print(f"  {r.get('map_marker',''):<3} {name:<42} {km:>5.1f} "
              f"{int(crow.get('remaining_yrs') or 0):>4} "
              f"{ftm:>6.2f}% {adj:>6.2f}%")

    # ── Source summary ────────────────────────────────────────────────────────
    total_sources = sum(len(r.get("sources") or []) for r in classified)
    if total_sources:
        print(f"\n  Sources ({total_sources} URL(s) across {len(classified)} comp(s)):")
        for r in classified:
            srcs = r.get("sources") or []
            if srcs:
                mk   = r.get("map_marker", "?")
                name = str(r.get("property_name", ""))[:38]
                print(f"  {mk}. {name}")
                for s in srcs:
                    title = (s.get("title") or "")[:60]
                    url   = s.get("url", "")
                    print(f"       ✦ {title}")
                    if url:
                        print(f"         {url}")
    else:
        print("\n  (No source URLs captured — search model may not have returned annotations)")

    # ── Write Excel ───────────────────────────────────────────────────────────
    print(f"\n[4/5] RENDER   {out_excel}")
    build_workbook_online(subj_row, comp_rows, subject_cfg, out_excel,
                          bala_yield, levels_used, classified=classified)

    # ── Optionally generate map ───────────────────────────────────────────────
    if generate_map:
        print(f"\n[5/5] MAP   {out_map}")
        _generate_map(classified, subject_cfg, subj_row, s_lon, s_lat,
                      mapbox_tok, out_map, mb_cfg)
    else:
        print(f"\n[5/5] MAP   skipped (pass --map to generate)")

    print("\nDone.\n")


def build_search_queries_for_level(subject_cfg: dict, level: str, years_back: int) -> list:
    """Thin wrapper so run() stays readable."""
    return build_queries(subject_cfg, level, years_back)


def _generate_map(classified: list, subject_cfg: dict, subj_row: dict,
                  s_lon: float, s_lat: float,
                  mapbox_tok: str, out_map: str, mb_cfg: dict):
    from generate_sales_comps_map import render_map
    comps_geo = [
        (r["map_marker"], r["lon"], r["lat"])
        for r in classified
        if r.get("lon") is not None
    ]
    render_map(
        subject_lonlat = (s_lon, s_lat),
        comps          = comps_geo,
        token          = mapbox_tok,
        output_path    = out_map,
        style          = mb_cfg.get("style",    "streets-v12"),
        width          = mb_cfg.get("width",    1200),
        height         = mb_cfg.get("height",   900),
        padding        = mb_cfg.get("padding",  100),
        pin_size       = mb_cfg.get("pin_size", "xl"),
    )


# ═══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Search online for comparable transactions.")
    parser.add_argument("--config", default="configs/deal_config.json")
    parser.add_argument("--map",     action="store_true",
                        help="Also generate a Mapbox map PNG after the Excel")
    parser.add_argument("--refresh", action="store_true",
                        help="Discard cached search results and re-search the web")
    args = parser.parse_args()
    run(args.config, generate_map=args.map, refresh=args.refresh)
