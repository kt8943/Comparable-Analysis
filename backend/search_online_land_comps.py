#!/usr/bin/env python3
"""
search_online_land_comps.py
============================
Searches online for land sale comparables near the subject property using
OpenAI's web search capability.  Targets GLS tenders, en bloc land deals,
and private land transactions — NOT building/asset sales.

Output schema  (13 columns — company template)
-----------------------------------------------
  Property | Map Marker | Date of Launch | Land Zoning | Land Tenure (Y)
  Site Area (SF) | Max GFA (SF) | Price (SGD M) | Price (SGD psf ppr)
  Adj. Price (SGD psf ppr) | Location | Quality | Comment

Search strategy (proximity-first, sub-market fallback)
-------------------------------------------------------
  Level 1 — Proximity  : within proximity_km of subject (default 1km)
  Level 2 — Sub-market : same location tier, within submarket_km (default 5km)
  Level 3 — Broader    : full market, no distance cap

Usage
-----
    python3 search_online_land_comps.py --config configs/deal_config_88_Cecil.json
    python3 search_online_land_comps.py --config configs/deal_config_88_Cecil.json --map
    python3 search_online_land_comps.py --config configs/deal_config_88_Cecil.json --map --refresh

Config keys used
----------------
    openai.api_key          : OpenAI API key (or set OPENAI_API_KEY env var)
    land_search             : optional dict to override search parameters
      proximity_km, submarket_km, min_results, max_results, years_back, max_level
    output_file             : used to derive the output folder
"""

import sys as _sys
from pathlib import Path as _Path
_sys.path.insert(0, str(_Path(__file__).parent))
# ── Windows UTF-8 fix ─────────────────────────────────────────────────────────
for _stream in (_sys.stdout, _sys.stderr):
    try:
        if getattr(_stream, "encoding", "utf-8").lower().replace("-", "") != "utf8":
            _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, OSError):
        pass

import argparse
import hashlib
import json
import math
import os
import re
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl

from generate_land_comps_map import geocode_with_fallbacks, render_map
from generate_land_comps_table import (
    get_land_schema, bala_factor,
    subject_to_row, comp_to_row,
    _fill, _font, _border, _align,
    _NAVY, _NAVYL, _NOTE, _WHITE, _LGRAY, _DARK,
    _section_header, _col_headers, _data_row, _write_formulas,
    _build_params_sheet,
)
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# SEARCH QUERIES  (land-specific)
# ─────────────────────────────────────────────────────────────────────────────

_SUBMARKET_KEYWORDS = {
    "Prime Core CBD (Marina Bay)":        ["Marina Bay", "Marina Boulevard", "Central Boulevard"],
    "Core CBD (Raffles Place / Shenton)": ["Raffles Place", "Shenton Way", "Robinson Road"],
    "Core CBD (City Hall / Beach Rd)":    ["City Hall", "Beach Road", "Suntec"],
    "Orchard / City Fringe":              ["Orchard Road", "Somerset", "Penang Road"],
    "CBD Fringe (Tanjong Pagar)":         ["Tanjong Pagar", "Anson Road", "Cecil Street"],
    "Decentralised / JLD":                ["Jurong Lake District", "JLD", "Jurong East"],
    "Decentralised (Paya Lebar)":         ["Paya Lebar", "Geylang"],
    "one-north / Buona Vista":            ["one-north", "Fusionopolis", "Buona Vista"],
    "Suburban":                           ["Woodlands", "Yishun", "Tampines"],
}


def _year_window(years_back: int) -> str:
    now = datetime.now().year
    return " OR ".join(str(now - i) for i in range(years_back + 1))


def build_land_queries(subject_cfg: dict, level: str, years_back: int = 3) -> list:
    """
    Build search queries for land sale comps at the given expansion level.
    Uses land-specific keywords: GLS, land tender, site area, psf ppr.
    """
    location     = subject_cfg.get("location", "")
    address      = subject_cfg.get("address", "")
    asset_class  = subject_cfg.get("asset_class", "office")
    country_name = subject_cfg.get("country_name", "Singapore")
    currency     = subject_cfg.get("currency", "SGD")
    yrs          = _year_window(years_back)

    # Submarket keywords
    kws_cfg = subject_cfg.get("submarket_keywords")
    kws     = kws_cfg or _SUBMARKET_KEYWORDS.get(location, [location] if location else [country_name])
    primary = kws[0] if kws else country_name

    # Asset keyword for land context
    asset_kw = subject_cfg.get("asset_search_keyword", asset_class)

    # Broader market query
    broader_cfg = subject_cfg.get("broader_market_query", "")
    broader_land = (broader_cfg.replace("investment sale", "land sale GLS tender")
                    if broader_cfg else
                    f"{country_name} {asset_kw} land sale GLS tender {currency}")

    if level == "proximity":
        parts    = [p.strip() for p in address.split(",") if p.strip()]
        precinct = parts[-3] if len(parts) >= 3 else (parts[-2] if len(parts) >= 2 else primary)
        return [
            f'{country_name} "{precinct}" land sale GLS tender site ({yrs})',
            f'{country_name} "{primary}" {asset_kw} land tender award ({yrs}) {currency} psf ppr',
        ]

    elif level == "submarket":
        kw_expr = " OR ".join(f'"{k}"' for k in kws[:3])
        return [
            f'{country_name} ({kw_expr}) land sale GLS award site area ({yrs}) {currency}',
            f'{country_name} ({kw_expr}) land tender {asset_kw} psf ppr ({yrs})',
        ]

    else:  # market
        return [
            f'{broader_land} ({yrs}) site area GFA tenure',
            f'{country_name} {asset_kw} land GLS tender award ({yrs}) URA JLL CBRE Savills',
        ]


# ─────────────────────────────────────────────────────────────────────────────
# OPENAI WEB SEARCH + EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def _parse_json_from_text(text: str) -> list:
    text = re.sub(r"^```[a-z]*\n?", "", text.strip())
    text = re.sub(r"\n?```$", "", text)
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", text, re.DOTALL)
        if not m:
            return []
        try:
            parsed = json.loads(m.group())
        except json.JSONDecodeError:
            return []
    if isinstance(parsed, dict):
        parsed = parsed.get("transactions") or next(
            (v for v in parsed.values() if isinstance(v, list)), [])
    return parsed if isinstance(parsed, list) else []


def search_and_extract_land(query: str, client,
                             search_model: str  = "gpt-4o-mini-search-preview",
                             extract_model: str = "gpt-4o-mini",
                             subject_cfg: dict  = None) -> tuple:
    """
    Two-step pipeline for land sale transactions.
    Returns (transactions: list, sources: list[dict]).
    """
    cfg          = subject_cfg or {}
    currency     = cfg.get("currency", "SGD")
    gfa_unit     = cfg.get("gfa_unit", "sf").lower()
    country      = cfg.get("country_name", "Singapore")
    asset_class  = cfg.get("asset_class", "office")
    gfa_desc     = "sqm" if gfa_unit == "sqm" else "sq ft"

    # Step 1: Web search
    search_resp = client.chat.completions.create(
        model=search_model,
        messages=[
            {"role": "system", "content":
             f"You are a {country} commercial real estate research analyst. "
             f"Search the web for confirmed land sale transactions and GLS tender results. "
             f"Focus on LAND parcels (not building sales). Extract site name, location, "
             f"transaction date, land tenure, site area, max GFA, price in {currency} millions, "
             f"land rate in {currency} psf ppr (per square foot per plot ratio), and zoning."},
            {"role": "user", "content": f"Search for: {query}"},
        ],
    )
    msg          = search_resp.choices[0].message
    article_text = msg.content or ""

    # Capture URL citations
    sources = []
    for ann in getattr(msg, "annotations", None) or []:
        if getattr(ann, "type", "") == "url_citation":
            uc = getattr(ann, "url_citation", None)
            if uc:
                sources.append({
                    "title": getattr(uc, "title", "") or "",
                    "url":   getattr(uc, "url",   "") or "",
                })
    seen_urls = set()
    sources = [s for s in sources if s["url"] not in seen_urls
               and not seen_urls.add(s["url"])]

    if not article_text or len(article_text) < 50:
        return [], sources

    # Step 2: Extract structured data
    asset_desc = {"logistics": "logistics / industrial", "office": "commercial / office",
                  "industrial": "industrial"}.get(asset_class, asset_class)
    extract_prompt = (
        f"Extract all confirmed {asset_desc} LAND SALE transactions from this text:\n---\n"
        f"{{text}}\n---\n\n"
        f"IMPORTANT: Only include land parcels intended for {asset_desc} use. "
        f"Exclude residential, hotel, and purely retail land transactions.\n\n"
        f"For each land transaction return a JSON object with:\n"
        f"  site_name     : name of the site or development\n"
        f"  address       : street address or location\n"
        f"  launch_date   : award date or transaction date (e.g. 'Mar 2024')\n"
        f"  land_zoning   : planning zone (e.g. Commercial, White, Business Park)\n"
        f"  tenure        : land tenure (e.g. '99-year Leasehold', 'Freehold')\n"
        f"  site_area_sf  : site area in {gfa_desc} (integer or null)\n"
        f"  max_gfa_sf    : maximum GFA in {gfa_desc} (integer or null)\n"
        f"  price_sgd_m   : price in {currency} millions (float or null)\n"
        f"  price_psf_ppr : land rate in {currency} psf ppr (integer or null)\n"
        f"  sale_type     : transaction type (e.g. 'GLS Confirmed List', 'GLS Reserve List', 'En Bloc', 'Private Sale')\n"
        f"  asset_type    : primary intended use of the land (e.g. 'Office', 'Logistics', 'Industrial')\n"
        f"  country       : country of the asset (default '{country}')\n\n"
        f"Include only entries where site_name AND price_sgd_m are known.\n"
        f"Return valid JSON only: {{{{\"transactions\": [...]}}}}"
    )
    extract_resp = client.chat.completions.create(
        model=extract_model,
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content":
             "You are a real estate data extraction assistant. Extract land sale transaction "
             "data from the provided article text. Return JSON with a 'transactions' array."},
            {"role": "user", "content": extract_prompt.format(text=article_text[:4000])},
        ],
    )
    return _parse_json_from_text(extract_resp.choices[0].message.content), sources


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _haversine_km(lon1: float, lat1: float, lon2: float, lat2: float) -> float:
    R = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2
         + math.cos(math.radians(lat1))
         * math.cos(math.radians(lat2))
         * math.sin(dlon / 2) ** 2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def _parse_tenure_yrs(val):
    if val is None:
        return None
    s = str(val).strip()
    if "freehold" in s.lower():
        return 0
    m = re.search(r"(\d+)", s)
    if m:
        n = int(m.group(1))
        return 0 if n >= 999 else n
    return None


def _num(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(",", "").replace("$", "").strip())
    except Exception:
        return None


# Shared asset-type guard (same logic as search_online_sales_comps.py)
_ASSET_ACCEPT_KWS = {
    "office":      ["office"],
    "logistics":   ["logistic", "warehouse", "cold storage", "distribution centre",
                    "distribution center", "fulfillment", "fulfilment"],
    "industrial":  ["industrial", "business park", "factory"],
    "retail":      ["retail", "shophouse", "mall"],
    "residential": ["residential", "apartment", "condominium", "condo", "landed"],
}
_ASSET_REJECT_KWS = {
    "office":      ["residential", "apartment", "condominium", "condo",
                    "hotel", "serviced apartment", "hospitality", "columbarium"],
    "logistics":   ["residential", "apartment", "condominium", "condo",
                    "hotel", "serviced apartment", "office"],
    "industrial":  ["residential", "apartment", "condominium", "condo",
                    "hotel", "serviced apartment"],
    "retail":      ["residential", "apartment", "condominium", "condo",
                    "hotel", "serviced apartment",
                    "logistics park", "warehouse", "industrial estate"],
    "residential": ["office", "logistics park", "logistics centre",
                    "warehouse", "industrial estate"],
}


def _asset_type_matches(record: dict, asset_class: str) -> bool:
    """Return True if the record's land zoning/type is compatible with asset_class."""
    if not asset_class:
        return True
    combined = " ".join(str(record.get(f) or "") for f in (
        "land_zoning", "sale_type", "asset_type", "comment",
    )).lower()
    if not combined.strip():
        return True
    ac = asset_class.lower()
    if any(kw in combined for kw in _ASSET_ACCEPT_KWS.get(ac, [])):
        return True
    if any(kw in combined for kw in _ASSET_REJECT_KWS.get(ac, [])):
        return False
    return True


def validate_dedup_land(records: list, subject_name: str = "",
                         subject_country: str = "",
                         subject_asset_class: str = "") -> list:
    """Deduplicate and filter land transaction records by country, asset type, and name."""
    seen_keys  = set()
    sc_tokens  = subject_country.lower().split() if subject_country else []
    subj_key   = re.sub(r"\W+", "", subject_name.lower())[:24] if subject_name else ""
    out        = []
    for r in records:
        name  = str(r.get("site_name") or r.get("property_name") or "").strip()
        price = _num(r.get("price_sgd_m"))
        if not name or not price:
            continue
        # Country filter
        rec_country = str(r.get("country") or "").lower()
        if rec_country and sc_tokens:
            if not any(tok in rec_country for tok in sc_tokens):
                continue
        # Asset-type filter: skip if clearly a different property type
        if subject_asset_class and not _asset_type_matches(r, subject_asset_class):
            print(f"    [type-filter] skipped '{name}' "
                  f"(not {subject_asset_class})")
            continue
        # Skip if appears to be subject
        norm = re.sub(r"\W+", "", name.lower())[:24]
        if subj_key and subj_key in norm:
            continue
        # Dedup
        key = norm + f"_{price:.0f}"
        if key in seen_keys:
            continue
        seen_keys.add(key)
        # Normalise field names
        r["property_name"] = name
        r["address"]       = str(r.get("address") or "").strip()
        r["launch_date"]   = str(r.get("launch_date") or "").strip()
        r["land_zoning"]   = str(r.get("land_zoning") or "").strip()
        r["tenure_yrs"]    = _parse_tenure_yrs(r.get("tenure"))
        r["site_area_sf"]  = (int(_num(r.get("site_area_sf")))
                               if _num(r.get("site_area_sf")) else None)
        r["max_gfa_sf"]    = (int(_num(r.get("max_gfa_sf")))
                               if _num(r.get("max_gfa_sf")) else None)
        r["price_sgd_m"]   = price
        r["price_psf_ppr"] = (int(_num(r.get("price_psf_ppr")))
                               if _num(r.get("price_psf_ppr")) else None)
        # If price_psf_ppr missing, compute from price + max_gfa
        if r["price_psf_ppr"] is None and price and r["max_gfa_sf"]:
            r["price_psf_ppr"] = round(price * 1_000_000 / r["max_gfa_sf"])
        r["comment"]       = str(r.get("sale_type") or "").strip()
        r["location"]      = ""
        r["quality"]       = ""
        out.append(r)
    return out


def geocode_land_records(records: list, mapbox_token: str,
                          country_code: str = "",
                          country_name: str = "") -> list:
    """Add lon/lat to each record. Only geocodes records that have a real address.
    Property names and descriptions are NOT used as geocoding queries — they
    produce unreliable or wrong coordinates."""
    suffix = f", {country_name}" if country_name else ""
    out = []
    for r in records:
        name  = str(r.get("property_name") or "").strip()
        addr  = str(r.get("address") or "").strip()
        if not addr:
            out.append({**r, "lon": None, "lat": None})
            continue
        query = f"{addr}{suffix}" if suffix.lower() not in addr.lower() else addr
        try:
            lon, lat, _ = geocode_with_fallbacks([query], mapbox_token, country_code)
            out.append({**r, "lon": lon, "lat": lat})
        except Exception:
            out.append({**r, "lon": None, "lat": None})
    return out


def _classify_rules(records: list):
    """Keyword rules fallback when OpenAI classification fails."""
    _LOC = [
        (["Marina Bay"],                                   "Prime Core CBD (Marina Bay)"),
        (["Raffles", "Robinson", "Shenton"],               "Core CBD (Raffles Place / Shenton)"),
        (["City Hall", "Beach Rd", "Suntec"],              "Core CBD (City Hall / Beach Rd)"),
        (["Orchard", "Somerset"],                          "Orchard / City Fringe"),
        (["Tanjong Pagar", "Cecil", "Anson", "McCallum"],  "CBD Fringe (Tanjong Pagar)"),
        (["Jurong", "JLD"],                                "Decentralised / JLD"),
        (["Paya Lebar"],                                   "Decentralised (Paya Lebar)"),
        (["one-north", "Fusionopolis", "Buona Vista"],     "one-north / Buona Vista"),
        (["Woodlands", "Yishun", "Tampines"],              "Suburban"),
    ]
    for r in records:
        desc   = f"{r.get('property_name', '')} {r.get('address', '')}".lower()
        loc    = "Decentralised"
        for kws, label in _LOC:
            if any(k.lower() in desc for k in kws):
                loc = label; break
        r.setdefault("location", loc)

        zoning = r.get("land_zoning", "").lower()
        if "white" in zoning:
            r.setdefault("quality", "Mixed-Use (White Site)")
        elif "business park" in zoning:
            r.setdefault("quality", "Grade B (Business Park)")
        elif "commercial" in zoning:
            r.setdefault("quality",
                         "Prime (Core CBD)" if "core" in loc.lower()
                         else "Grade A (CBD Fringe)")
        else:
            r.setdefault("quality", "")


def classify_land_online(records: list, subject_cfg: dict,
                          client, extract_model: str = "gpt-4o-mini") -> list:
    """Use OpenAI to classify location + quality for online-searched land comps."""
    if not records:
        return []
    slim = [{"index": i,
             "site_name":   r.get("property_name", ""),
             "address":     r.get("address", ""),
             "land_zoning": r.get("land_zoning", ""),
             "tenure_yrs":  r.get("tenure_yrs"),
             "max_gfa_sf":  r.get("max_gfa_sf"),
             "price_sgd_m": r.get("price_sgd_m")}
            for i, r in enumerate(records)]
    classify_prompt = (
        f"Subject property:\n{json.dumps(subject_cfg, indent=2)}\n\n"
        "For each LAND SALE comparable below assign:\n"
        "  location  — Singapore location tier (e.g. 'Core CBD (Raffles Place / Shenton)', "
        "'CBD Fringe (Tanjong Pagar)', 'Decentralised / JLD', etc.)\n"
        "  quality   — 'Prime (Core CBD)' | 'Grade A (CBD Fringe)' | 'Grade A (Decentralised)' "
        "| 'Mixed-Use (White Site)' | 'Grade B (Business Park)' | 'Grade B (Suburban)'\n\n"
        f"Comparables:\n{json.dumps(slim, indent=2)}\n\n"
        'Return valid JSON only: {"results": [{"index":0,"location":"...","quality":"..."},...]}'
    )
    try:
        resp = client.chat.completions.create(
            model=extract_model,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content":
                 "Senior real estate analyst. Return JSON classification."},
                {"role": "user", "content": classify_prompt},
            ],
        )
        parsed   = json.loads(resp.choices[0].message.content)
        cls_list = parsed.get("results") or next(
            (v for v in parsed.values() if isinstance(v, list)), [])
        out = [dict(r) for r in records]
        for cls in cls_list:
            idx = cls.get("index")
            if idx is not None and idx < len(out):
                out[idx]["location"] = cls.get("location", "")
                out[idx]["quality"]  = cls.get("quality", "")
        return out
    except Exception as exc:
        print(f"      OpenAI classify failed ({exc}). Using keyword rules.")
        _classify_rules(records)
        return records


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT WORKBOOK  (online variant — adds Sources sheet + AI disclaimer)
# ─────────────────────────────────────────────────────────────────────────────

def _build_sources_sheet(wb, records: list):
    ws = wb.create_sheet("Sources")
    ws.sheet_view.showGridLines = False
    headers = ["#", "Site / Property", "Source Title", "URL"]
    widths  = [4, 38, 44, 70]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(col)].width = w
        c = ws.cell(row=1, column=col, value=h)
        c.fill = _fill(_NAVY)
        c.font = _font(_WHITE, bold=True, sz=10)
        c.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 18

    row = 2
    for rec in records:
        marker = str(rec.get("map_marker", ""))
        name   = str(rec.get("property_name", ""))
        srcs   = rec.get("sources") or [{"title": "No URL captured", "url": ""}]
        for i, s in enumerate(srcs):
            alt = (row % 2 == 0)
            bg  = _LGRAY if alt else _WHITE
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = _fill(bg)
                ws.cell(row=row, column=col).alignment = _align("left", "center", wrap=False)
            ws.cell(row=row, column=1, value=marker if i == 0 else "")
            ws.cell(row=row, column=2, value=name   if i == 0 else "")
            ws.cell(row=row, column=3, value=s.get("title") or "")
            url = s.get("url") or ""
            uc  = ws.cell(row=row, column=4, value=url)
            if url:
                uc.hyperlink = url
                uc.font = _font("FF1155CC", sz=9, bold=False)
                uc.style = "Hyperlink"
            row += 1


def build_workbook_online(subject_cfg: dict, subject_row_d: dict,
                          comp_rows: list, out_excel: str,
                          schema: list, bala_yield: float,
                          levels_used: list, classified: list = None):
    """Online-specific workbook: adds AI disclaimer row and Sources sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Online Land Comps"
    ws.sheet_view.showGridLines = False

    nc        = len(schema)
    deal_name = subject_cfg.get("deal_name", subject_cfg["property_name"])
    levels    = " + ".join(levels_used) if levels_used else "web search"

    for i, (*_, w) in enumerate(schema, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    t = ws.cell(row=r, column=1,
                value=f"{deal_name.upper()}  —  ONLINE LAND SALE COMPARABLES (AI-SOURCED)")
    t.fill = _fill(_NAVY); t.font = _font(_WHITE, bold=True, sz=13)
    t.alignment = _align("center", "center", wrap=False)
    ws.row_dimensions[r].height = 24

    r = 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    s = ws.cell(row=r, column=1,
                value=f"AI web search: {levels}  |  Verify all figures before use  |  "
                       "Confidential — For Discussion Purposes Only")
    s.fill = _fill(_NAVYL); s.font = _font(_WHITE, bold=False, sz=9, italic=True)
    s.alignment = _align("center", "center", wrap=False)
    ws.row_dimensions[r].height = 14

    # Table 1: Subject
    r = 4
    _section_header(ws, r, "  Subject Property — Site / Land Reference", nc)
    r = 5; _col_headers(ws, r, schema)
    r = 6; _data_row(ws, r, subject_row_d, schema, bold=True)
    _write_formulas(ws, r, schema, is_subject=True)
    ws.row_dimensions[7].height = 6

    # Table 2: Online Land Comps
    r = 8
    _section_header(ws, r, "  Online Land Sale Comparables (AI-Sourced)", nc)
    r = 9; _col_headers(ws, r, schema)
    for i, crow in enumerate(comp_rows):
        r = 10 + i
        _data_row(ws, r, crow, schema, alt=(i % 2 == 1))
        _write_formulas(ws, r, schema, is_subject=False)

    # Notes footer
    r = 10 + len(comp_rows) + 1
    currency = subject_cfg.get("currency", "SGD")
    area_lbl = "psm" if subject_cfg.get("gfa_unit","sf").lower()=="sqm" else "psf"
    notes = (
        "Notes:  Data sourced via OpenAI web search — verify all figures against "
        "URA, JLL, CBRE, Savills before use in formal documents.  "
        f"(2) Price ({currency} {area_lbl} ppr) = Sale Price × 1,000,000 / Max GFA (live formula).  "
        f"(3) Adj. Price ({currency} {area_lbl} ppr) uses Bala Table; see Params sheet.  "
        f"Search levels applied: {levels}."
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    n = ws.cell(row=r, column=1, value=notes)
    n.fill = _fill(_NOTE); n.font = _font("FF404040", bold=False, sz=8, italic=True)
    n.alignment = _align("left", "top", wrap=True)
    ws.row_dimensions[r].height = 56

    _build_params_sheet(wb, subject_cfg, bala_yield)
    if classified:
        _build_sources_sheet(wb, classified)
    wb.save(out_excel)
    print(f"  Saved → {out_excel}")


# ─────────────────────────────────────────────────────────────────────────────
# SEARCH CACHE
# ─────────────────────────────────────────────────────────────────────────────

def _cache_key(subject_cfg: dict, sc_cfg: dict) -> str:
    blob = json.dumps({"s": subject_cfg, "p": sc_cfg}, sort_keys=True)
    return hashlib.md5(blob.encode()).hexdigest()[:12]


def load_cache(cache_path: str, expected_key: str):
    try:
        with open(cache_path, encoding="utf-8") as f:
            data = json.load(f)
        if data.get("cache_key") != expected_key:
            print("  Cache key mismatch (config changed) — running fresh search.")
            return None, None
        print(f"  Loaded from cache (saved {data.get('timestamp','?')}). "
              "Pass --refresh to re-search.")
        return data["records"], data["levels_used"]
    except FileNotFoundError:
        return None, None


def save_cache(cache_path: str, records: list, levels_used: list, cache_key: str):
    data = {"cache_key":   cache_key,
            "timestamp":   datetime.now().strftime("%Y-%m-%d %H:%M"),
            "levels_used": levels_used,
            "records":     records}
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)
    print(f"  Cache saved → {cache_path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def run(config_path: str = "configs/deal_config.json",
        generate_map: bool = False, refresh: bool = False):

    with open(config_path, encoding="utf-8") as f:
        cfg = json.load(f)

    subject_cfg   = cfg["subject_property"]
    params        = cfg.get("parameters", {})
    bala_yield    = params.get("bala_yield",   0.06)
    mb_cfg        = cfg.get("mapbox", {})
    mapbox_tok    = mb_cfg.get("token", "")
    oa_cfg        = cfg.get("openai", {})
    api_key       = oa_cfg.get("api_key") or os.environ.get("OPENAI_API_KEY", "")
    search_model  = oa_cfg.get("search_model",  "gpt-4o-mini-search-preview")
    extract_model = oa_cfg.get("extract_model", "gpt-4o-mini")

    # Search config (prefer land_search, fall back to online_search)
    sc_cfg        = cfg.get("land_search") or cfg.get("online_search") or {}
    proximity_km  = sc_cfg.get("proximity_km",   1.0)
    submarket_km  = sc_cfg.get("submarket_km",   5.0)
    market_km     = sc_cfg.get("market_km",      submarket_km)
    min_results   = sc_cfg.get("min_results",    3)
    max_results   = sc_cfg.get("max_results",    10)
    years_back    = sc_cfg.get("years_back",     3)
    years_back_max  = sc_cfg.get("years_back_max",   10)
    years_back_step = sc_cfg.get("years_back_step",   2)
    max_level     = sc_cfg.get("max_level",      3)
    # Grounded data sources to combine with (or instead of) OpenAI web search.
    # Defaults to web-search only → identical behaviour to before.
    sources_cfg   = sc_cfg.get("sources") or ["web_search"]

    if not api_key:
        raise ValueError("OpenAI API key not found.  Set openai.api_key in config or "
                         "export OPENAI_API_KEY=sk-...")
    if not mapbox_tok:
        raise ValueError("Add mapbox.token to deal config (needed for geocoding).")

    try:
        from openai import OpenAI
    except ImportError:
        raise ImportError("Run: pip3 install openai")

    client       = OpenAI(api_key=api_key)
    prop_name    = subject_cfg["property_name"]
    deal_name    = subject_cfg.get("deal_name", prop_name)
    deal_slug    = deal_name.replace(" ", "_")
    country_code = cfg.get("country_code", "")
    country_name = subject_cfg.get("country_name", "")

    output_file = cfg.get("output_file", f"output/{deal_slug}/{deal_slug}.xlsx")
    out_dir     = Path(output_file).parent
    out_dir.mkdir(parents=True, exist_ok=True)
    out_excel   = str(out_dir / f"Online_Land_Comps_{deal_slug}.xlsx")
    out_map     = str(out_dir / f"Online_Land_Comps_{deal_slug}_map.png")
    cache_path  = str(out_dir / f"Online_Land_Comps_{deal_slug}_search_cache.json")

    print(f"\n{'='*64}")
    print(f"  Online Land Comps Search : {deal_name}")
    print(f"{'='*64}")

    # Geocode subject
    address = subject_cfg.get("address", "")
    print(f"\n[0/5] Geocoding subject property")
    s_lon, s_lat, _ = geocode_with_fallbacks(
        [f"{prop_name}, {address}", address, prop_name],
        mapbox_tok, country_code,
    )
    print(f"      {prop_name}  →  ({s_lon:.5f}, {s_lat:.5f})")

    # Cache handling
    c_key = _cache_key(subject_cfg, sc_cfg)
    if not refresh:
        print(f"\n[Search] Checking cache …")
        all_records, levels_used = load_cache(cache_path, c_key)
    else:
        print(f"\n[Search] --refresh: discarding existing cache.")
        all_records, levels_used = None, None

    if all_records is None:
        GEO_LEVELS   = [
            ("proximity", proximity_km, f"Proximity ≤{proximity_km}km"),
            ("submarket", submarket_km, f"Sub-market ≤{submarket_km}km"),
        ]
        BROAD_LEVEL  = ("market", market_km, f"Broader market ≤{market_km}km")
        all_records  = []
        levels_used  = []
        seen_keys    = {}
        yrs          = years_back

        def _merge_geocoded(geocoded, srcs, max_km, source_name=""):
            """Distance-filter + dedup-merge a batch of geocoded records into
            all_records; returns count added. Shared by web search and grounded
            connectors so every source is treated identically."""
            added = 0
            srcs  = srcs or []
            for r in geocoded:
                price = float(r.get("price_sgd_m") or 0)
                key   = (re.sub(r"\W+", "", str(r.get("property_name", "")).lower())[:24]
                         + f"_{price:.0f}")
                if (max_km is not None and r.get("lon") is not None
                        and _haversine_km(r["lon"], r["lat"], s_lon, s_lat) > max_km):
                    continue
                _srcs = [({**s, "source_name": source_name} if source_name else s) for s in srcs]
                if key in seen_keys:
                    idx = seen_keys[key]
                    existing = {s.get("url") for s in all_records[idx].get("sources", [])}
                    for s in _srcs:
                        if s.get("url") and s["url"] not in existing:
                            all_records[idx].setdefault("sources", []).append(s)
                    continue
                r["sources"]   = list(_srcs)
                seen_keys[key] = len(all_records)
                all_records.append(r)
                added += 1
            return added

        def _run_level(level_id, max_km, level_label, yrs_used):
            queries   = build_land_queries(subject_cfg, level_id, yrs_used)
            level_new = 0
            print(f"\n[Search] {level_label}  (years back: {yrs_used})")
            for q in queries:
                print(f"  Query: {q[:90]}…" if len(q) > 90 else f"  Query: {q}")
                try:
                    raw, q_sources = search_and_extract_land(
                        q, client, search_model, extract_model,
                        subject_cfg=subject_cfg)
                except Exception as e:
                    print(f"  ✗ {e.__class__.__name__}: {e}")
                    continue
                if q_sources:
                    print(f"  ✦ {len(q_sources)} source(s) found")
                cleaned  = validate_dedup_land(raw, subject_name=prop_name,
                                               subject_country=country_name,
                                               subject_asset_class=subject_cfg.get("asset_class", ""))
                geocoded = geocode_land_records(cleaned, mapbox_tok, country_code,
                                                country_name=country_name)
                level_new += _merge_geocoded(geocoded, q_sources, max_km)
            lbl = f"{level_label} (yrs:{yrs_used})"
            if lbl not in levels_used:
                levels_used.append(lbl)
            print(f"  → {level_new} new  |  total: {len(all_records)}")
            return level_new

        if "web_search" in sources_cfg:
            active_levels = GEO_LEVELS[:max_level]
            while len(all_records) < min_results:
                for level_id, max_km, label in active_levels:
                    _run_level(level_id, max_km, label, yrs)
                    if len(all_records) >= min_results:
                        break
                if len(all_records) >= min_results:
                    print(f"  ✓ min_results ({min_results}) reached.")
                    break
                if yrs < years_back_max:
                    yrs += years_back_step
                    print(f"\n  ↩  Sparse results ({len(all_records)}). "
                          f"Extending to {yrs} years back …")
                else:
                    print(f"\n  ↩  Temporal limit reached. {len(all_records)} record(s) found.")
                    break

            if len(all_records) < min_results and max_level >= 3:
                print(f"\n  Falling back to broader market search.")
                _run_level(*BROAD_LEVEL, yrs)

        # ── Grounded connectors (URA GLS, …): fetch once, ingest via same pipeline ──
        from sources.registry import get_grounded
        _conn_params = {"country_code": country_code, "country_name": country_name,
                        "years_back": yrs, "s_lon": s_lon, "s_lat": s_lat,
                        "proximity_km": proximity_km, "submarket_km": submarket_km,
                        "market_km": market_km}
        for _conn in get_grounded((country_code or "sg").lower(), "land",
                                  sources_cfg, _conn_params):
            print(f"\n[Source] {_conn.label or _conn.name}")
            try:
                _raw, _srcs = _conn.fetch(subject_cfg, _conn_params)
            except Exception as e:
                print(f"  ✗ {e.__class__.__name__}: {e}")
                continue
            if not _raw:
                print("  → 0 records")
                continue
            _cleaned = validate_dedup_land(_raw, subject_name=prop_name,
                                           subject_country=country_name,
                                           subject_asset_class=subject_cfg.get("asset_class", ""))
            _geo = geocode_land_records(_cleaned, mapbox_tok, country_code,
                                        country_name=country_name)
            _added = _merge_geocoded(
                _geo, _srcs or [{"title": _conn.label or _conn.name, "url": ""}],
                market_km, source_name=_conn.name)
            _lbl = _conn.label or _conn.name
            if _lbl not in levels_used:
                levels_used.append(_lbl)
            print(f"  → {_added} new  |  total: {len(all_records)}")

        save_cache(cache_path, all_records, levels_used, c_key)

    records = all_records[:max_results]
    print(f"\n[1/5] SEARCH complete — {len(records)} record(s)")

    if not records:
        print("\n  No land sale transactions found. "
              "Consider widening years_back or checking OpenAI model supports web search.\n")
        return

    # Classify
    print(f"\n[2/5] CLASSIFY  ({extract_model})")
    classified = classify_land_online(records, subject_cfg, client, extract_model)
    print(f"      → {len(classified)} records classified")

    # Sort by distance
    print(f"\n[3/5] SORT by distance")
    classified = [r for r in classified if r.get("lon") is not None]
    classified.sort(key=lambda r: _haversine_km(r["lon"], r["lat"], s_lon, s_lat))
    for i, r in enumerate(classified):
        r["map_marker"] = str(i + 1)

    # Print summary
    currency = subject_cfg.get("currency_symbol", subject_cfg.get("currency", "S$"))
    print(f"\n  {'#':<3} {'Site':<40} {'km':>5}  {'Price M':>9}  {'Yrs':>5}")
    print("  " + "─" * 68)
    for r in classified:
        km   = _haversine_km(r["lon"], r["lat"], s_lon, s_lat)
        pm   = float(r.get("price_sgd_m") or 0)
        yrs_s = "FH" if not r.get("tenure_yrs") or r["tenure_yrs"] == 0 else str(r["tenure_yrs"])
        name  = str(r.get("property_name") or "")[:40]
        print(f"  {r['map_marker']:<3} {name:<40} {km:>5.2f}  "
              f"{currency}{pm:>7.1f}M  {yrs_s:>5}")

    # Render
    print(f"\n[4/5] RENDER   {out_excel}")
    schema      = get_land_schema(subject_cfg)
    subj_row    = subject_to_row(subject_cfg)
    comp_rows   = [comp_to_row(r) for r in classified]
    build_workbook_online(subject_cfg, subj_row, comp_rows, out_excel,
                          schema, bala_yield, levels_used, classified=classified)

    if generate_map:
        print(f"\n[5/5] MAP   {out_map}")
        comps_geo = [(r["map_marker"], r["lon"], r["lat"]) for r in classified]
        render_map(
            subject_lonlat = (s_lon, s_lat),
            comps          = comps_geo,
            token          = mapbox_tok,
            output_path    = out_map,
            style          = mb_cfg.get("style",    "streets-v12"),
            width          = mb_cfg.get("width",    1200),
            height         = mb_cfg.get("height",   900),
            padding        = mb_cfg.get("padding",  100),
            pin_size       = mb_cfg.get("pin_size", "l"),
        )
    else:
        print(f"\n[5/5] MAP   skipped (pass --map to generate)")

    print(f"\n  Done — {len(classified)} comp(s) written to {out_excel}\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Search online for land sale comparables (GLS, en bloc, private land)."
    )
    parser.add_argument("--config",  default="configs/deal_config.json")
    parser.add_argument("--map",     action="store_true",
                        help="Generate Mapbox map PNG")
    parser.add_argument("--refresh", action="store_true",
                        help="Discard cached search results and re-search")
    args = parser.parse_args()
    run(args.config, generate_map=args.map, refresh=args.refresh)
