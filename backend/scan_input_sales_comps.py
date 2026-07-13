#!/usr/bin/env python3
"""
scan_input_sales_comps.py
=========================
Reads a user-provided Excel of asset sales comparables, uses Ollama to
intelligently map columns and classify records, then produces the formatted
13-column output Excel (company template) and an optional Mapbox map.

No web search — all data comes from the input file specified by
``input_file`` in the deal config.

Pipeline  (5 stages)
--------------------
  1  GEOCODE    Geocode the subject property via Mapbox (skipped if no token)
  2  PARSE      Read input Excel; Ollama auto-detects column mapping
  3  CLASSIFY   Ollama assigns location, quality, asset type, relevance score
  4  CALCULATE  Price psf GFA + Bala-adjusted cap rate; geocode + sort comps
  5  RENDER     Formatted 13-column Excel; optional Mapbox map PNG

Usage
-----
    python3 scan_input_sales_comps.py --config configs/deal_config_88_Cecil.json
    python3 scan_input_sales_comps.py --config configs/deal_config_88_Cecil.json --map

Config keys used
----------------
    input_file             : path to the input Excel with sales comp data (required)
    output_file            : destination Excel path
    parameters.max_comps   : (removed — all comps are kept)
    parameters.bala_yield  : (removed — official Bala Table is used directly)
"""

import sys as _sys
from pathlib import Path as _Path
_sys.path.insert(0, str(_Path(__file__).parent))
# ── Corporate proxy TLS fix (trust OS cert store; no-op without truststore) ────
from tools import corp_ssl  # noqa: F401  — must import before any HTTPS call
# ── Windows UTF-8 fix ─────────────────────────────────────────────────────────
for _stream in (_sys.stdout, _sys.stderr):
    try:
        if getattr(_stream, "encoding", "utf-8").lower().replace("-", "") != "utf8":
            _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, OSError):
        pass
import argparse
import json
import math
import re
import urllib.request
from pathlib import Path

import openpyxl

from generate_sales_comps_table import (
    bala_factor, subject_to_row, comp_to_row, build_workbook,
)
from generate_sales_comps_map import render_map
import generate_global_sales_comps_table as _global_sales_tbl
from generate_comps_map_base import geocode_any as geocode_with_fallbacks, build_geocode_queries as _build_geocode_queries, near_country_centroid as _near_country_centroid, country_code_from_name as _cc_from_name, clean_property_name as _clean_name
from tools.calculations import (
    haversine_km as _haversine_km,
    parse_num as _num,
    parse_remaining_yrs as _parse_remaining_yrs,
    parse_sale_date as _parse_sale_date,
)
from tools.json_utils import (
    fix_json as _fix_json,
    split_json_arrays as _split_json_arrays,
)
from tools.llm_client import ollama_post as _ollama_post, apply_refinement as _apply_refinement
from tools.excel_reader import find_best_sheet as _find_best_sheet, find_header_row as _find_header_row, sheet_keywords as _sheet_keywords, split_tables as _split_tables
from tools.vision_llm import call_vision_llm as _call_vision_llm
from tools.column_mapper import map_columns as _map_columns
from tools.geo_utils import write_geo_sidecar
from tools.ura_zone import resolve_ura_zone as _resolve_ura_zone, zone_from_coords as _zone_from_coords


# ─────────────────────────────────────────────────────────────────────────────
# FIELD REQUIREMENTS  —  flip the switch here, no other changes needed
# ─────────────────────────────────────────────────────────────────────────────

# Set True  → GFA (sf) must be present; records without it are dropped.
# Set False → GFA is optional; price_psf_gfa will be blank when absent.
_REQUIRE_GFA = False

# "Property name only" mode (set at runtime via run(name_only=...) / --name-only).
# When True, a comp qualifies on its property name alone — the price gate is skipped.
_NAME_ONLY = False


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL PARSING  (Ollama-driven column mapping)
# ─────────────────────────────────────────────────────────────────────────────

# Output column names drive column detection — DO NOT rename these; they are
# the exact column names that appear in the output Excel report.
# The description (3rd element) tells Ollama how to find the matching input column.
_OUTPUT_FIELDS = [
    # (output_col_name,            internal_key,    description_for_ollama)
    ("Property",                "property_name",  "Name of the specific building or asset — NOT land zoning or property type — e.g. 'Property', 'Property Name', 'Asset', 'Name', 'Building'"),
    ("Sale Date",               "sale_date_raw",  "Transaction or sale date — e.g. 'Transaction Date', 'Date', 'Period'"),
    ("Land Zoning",             "land_zoning",    "Land use or planning zone for the site — e.g. 'Zoning', 'Land Use', 'Sector', 'Property Type', 'Asset Class', 'Use', 'Type of Development Allowed', 'Permitted Use', 'Development Type'. Do NOT use the tenure or leasehold years column for this."),
    ("Remaining Leasehold (Y)", "remaining_yrs",  "Lease duration information — extract the raw cell value exactly as shown, whether it is already the remaining years (e.g. 77) or a full tenure string (e.g. '99 years from 2004', '99-year lease wef 2010', 'Freehold', 'FH'). Do NOT map land-zoning or property-type columns here. Columns to look for: 'Tenure', 'Leasehold', 'Remaining Lease', 'Lease Term'."),
    ("GFA (SF)",                "gfa_sf",         "Floor area (size of the building/unit) — e.g. 'GFA', 'NLA', 'Net Lettable Area', 'Transacted Area (sqm)', 'Transaction Area'. This is a SIZE field, NOT a price."),
    ("Price (SGD M)",           "price_sgd_m",    "Transaction price in millions (any currency) — e.g. 'Sale Price', 'Transaction Price'"),
    ("Price (SGD psf GFA)",     "price_psf_gfa",  "Unit price per sq ft of GFA — e.g. 'Price psf GFA', 'Price (SGD psf GFA)', 'Price psf'"),
    ("FTM NOI Capitalisation Rate", "npi_yield",   "Forward-looking stabilised NOI/NPI yield at time of transaction — e.g. 'FTM NOI Cap Rate', 'NPI Yield', 'Cap Rate', 'Net Yield', 'Capitalisation Rate'. Do NOT use the adjusted cap rate column."),
    ("Adj. Capitalisation Rate",    "adj_npi_yield", "Leasehold-adjusted or Bala-adjusted cap rate — e.g. 'Adj. Cap Rate', 'Adjusted Cap Rate', 'Adj NPI Yield', 'Adj. Capitalisation Rate'. This is DIFFERENT from FTM NOI Cap Rate."),
    ("Sale Type",               "sale_type",      "Sale structure — e.g. 'Strata', 'Whole Bldg', 'Block Sale', 'En Bloc'"),
    ("Buyer",                   "buyer",          "Purchaser or acquiring entity — e.g. 'Buyer', 'Purchaser', 'Acquirer'"),
]
# Address is an internal helper field (not a standalone output column)
# — used to enrich the Property field with a full address line.
_EXTRA_FIELDS = [
    ("address", "Street address of the property — e.g. 'Address', 'Street'"),
    ("location", "Location-competitiveness label already provided in the input — values "
                 "are 'Superior', 'Comparable' or 'Inferior'. NOT a street address or "
                 "building name; do not use for geocoding."),
]
_OUTPUT_COL_TO_KEY = {col: key for col, key, _ in _OUTPUT_FIELDS}

# All transaction types are kept exactly as extracted from the input —
# no rows are filtered out based on sale type.

_sheet_kws  = _sheet_keywords(_OUTPUT_FIELDS)
_best_sheet = lambda wb: _find_best_sheet(wb, _sheet_kws)



def parse_input_excel(input_file: str, base_url: str, model: str,
                      subject_name: str = "", llm_cfg: dict = None,
                      area_unit: str = "sf", _segment: tuple = None) -> list:
    """
    Read any Excel of asset sales comps.  Uses LLM (GPT or Ollama) to map columns.
    Returns list of dicts with standardised field names ready for
    classify_comps() and metrics calculation.
    All transaction types are kept exactly as extracted from the input.
    area_unit: target output unit for area fields — "sf" (default, SG) or "sqm" (global).

    A single sheet may contain several STACKED tables with different column layouts;
    each is detected and parsed with its OWN header row (via ``_segment`` recursion),
    so a table below the first is never mapped using the first table's columns.
    """
    if _segment is not None:
        best, headers, data_rows = _segment
        print(f"  Headers: {[h for h in headers if h]}")
    else:
        wb   = openpyxl.load_workbook(input_file, data_only=True)
        best = _best_sheet(wb)
        ws   = wb[best]
        rows = [tuple(c.value for c in row) for row in ws.iter_rows()]

        segments = _split_tables(rows, _sheet_kws)
        if len(segments) > 1:
            print(f"  Sheet {best!r}: {len(segments)} stacked tables detected — "
                  "parsing each with its own header row.")
            _all, _cm0, _hd0, _pu0 = [], None, None, "M"
            for _k, (_hi, _sh, _sd) in enumerate(segments, 1):
                print(f"\n  ══ Table {_k}/{len(segments)} — header row {_hi + 1}, "
                      f"{len(_sd)} data row(s) ══")
                _recs, _cm, _hd, _pu = parse_input_excel(
                    input_file, base_url, model, subject_name=subject_name,
                    llm_cfg=llm_cfg, area_unit=area_unit, _segment=(best, _sh, _sd))
                _all += _recs
                if _cm0 is None:
                    _cm0, _hd0, _pu0 = _cm, _hd, _pu
            return _all, _cm0 or {}, _hd0 or [], _pu0

        _hi, headers, data_rows = segments[0]
        print(f"  Sheet: {best!r}  |  Header row: {_hi + 1}  |  Data rows: {len(data_rows)}")
        print(f"  Headers: {[h for h in headers if h]}")

    # Pick sample rows for Ollama: prefer rows with >= 3 non-empty cells
    # (avoids section sub-header rows that only have 1–2 cells filled)
    sample_rows = [r for r in data_rows
                   if sum(1 for c in r if c not in (None, "")) >= 3][:3]
    if not sample_rows:
        sample_rows = data_rows[:3]

    # Tiered column mapping: exact → embedding → LLM (GPT/Ollama)
    print(f"  Mapping columns …")
    _is_global = area_unit.lower() != "sf"
    col_map, unit_map = _map_columns(
        headers, sample_rows, _OUTPUT_FIELDS, _OUTPUT_COL_TO_KEY,
        base_url, model, extra_fields=_EXTRA_FIELDS, llm_cfg=llm_cfg,
        passthrough_units=_is_global,
    )
    # For global comps passthrough_units=True means no conversion was applied.
    # Detect the price unit from the column header directly.
    if _is_global:
        _date_col_idx_p = col_map.get("price_sgd_m")
        _price_header = headers[_date_col_idx_p] if _date_col_idx_p is not None and _date_col_idx_p < len(headers) else ""
        _price_unit_detected = "B" if re.search(r"\bbillion\b|\bbn\b|\bbil\b|\bb\b", _price_header.lower()) else "M"
    else:
        _price_unit_detected = "B" if unit_map.get("price_sgd_m", 1.0) >= 1000 else "M"

    # Extract year from the sale date column header (e.g. "Sale Date 2025").
    # Used as fallback when individual cells have no year (e.g. "Jan", "Q1").
    # Extract year from sheet name first (e.g. tab "2025" or "Sales 2025"),
    # then fall back to the sale date column header if the sheet name has no year.
    _date_header_year = None
    _sheet_ym = re.search(r"\b(20\d{2})\b", best)
    if _sheet_ym:
        _date_header_year = _sheet_ym.group(1)
        print(f"  [date-year] {_date_header_year!r} extracted from sheet name {best!r}")
    else:
        _date_col_idx = col_map.get("sale_date_raw")
        if _date_col_idx is not None and _date_col_idx < len(headers):
            _col_ym = re.search(r"\b(20\d{2})\b", headers[_date_col_idx])
            if _col_ym:
                _date_header_year = _col_ym.group(1)
                print(f"  [date-year] {_date_header_year!r} extracted from column header {headers[_date_col_idx]!r}")
    if not _date_header_year:
        print(f"  [date-year] no year found in sheet name {best!r} or sale date column header")

    def _get(row, field):
        idx = col_map.get(field)
        if idx is not None and idx < len(row):
            return row[idx]
        return None

    def _get_num(row, field):
        """Get numeric value with unit conversion applied."""
        val = _num(_get(row, field))
        if val is None:
            return None
        return val * unit_map.get(field, 1.0)

    subj_tokens = set(re.sub(r"\W+", " ", subject_name.lower()).split()) if subject_name else set()

    records = []
    for row in data_rows:
        # Build description from separate Property Name + Address columns
        name = str(_get(row, "property_name") or "").strip()
        addr = str(_get(row, "address")       or "").strip()
        desc = f"{name}\n{addr}" if addr else name
        label = name[:45] if name else "(no name)"
        if not desc:
            print(f"    SKIP {label!r:<47} — no property name or address")
            continue

        # Skip subject property.
        # Filter pure numbers (e.g. "88") before comparing — street numbers
        # appear in many addresses and cause false positives.  Require 80% of
        # the remaining tokens to match so one-word address differences (e.g.
        # "Cecil" vs "Market") are enough to distinguish two buildings.
        if subj_tokens:
            meaningful_subj = {t for t in subj_tokens if not t.isdigit()}
            if meaningful_subj:
                desc_tokens = set(re.sub(r"\W+", " ", desc.lower()).split())
                needed = max(1, math.ceil(len(meaningful_subj) * 0.75))
                if len(desc_tokens & meaningful_subj) >= needed:
                    print(f"    SKIP {label!r:<47} — matches subject property name")
                    continue

        # Skip totals / averages rows
        if re.search(r"\b(total|average|avg|summary|subtotal)\b", name, re.I):
            print(f"    SKIP {label!r:<47} — totals/summary row")
            continue

        gfa     = _get_num(row, "gfa_sf")
        price_m = _get_num(row, "price_sgd_m")
        # Safety net: if header had no unit marker, detect from value magnitude
        if price_m is not None and price_m > 100_000:
            price_m = round(price_m / 1_000_000, 3)
        if not price_m and not _NAME_ONLY:
            raw_price = _get(row, "price_sgd_m")
            print(f"    SKIP {label!r:<47} — no price (raw cell: {raw_price!r})")
            continue
        if _REQUIRE_GFA and not gfa:  # flip _REQUIRE_GFA at top of file to enforce
            print(f"    SKIP {label!r:<47} — no GFA (required)")
            continue

        sale_type = str(_get(row, "sale_type") or "").strip()

        # Derive stake_pct from sale_type or description text.
        # Whole building / block / en-bloc → 100%.
        # Partial stakes: look for "49%", "(1/3 stake)", or similar patterns.
        combined = f"{sale_type} {desc}"
        pct_m    = re.search(r"(\d+(?:\.\d+)?)\s*%", combined)
        frac_m   = re.search(r"\((\d+)\s*/\s*(\d+)\s*stake", combined, re.I)
        if any(kw in sale_type.lower() for kw in ("whole", "block sale", "en bloc")):
            stake_pct = 1.0
        elif pct_m:
            pct = float(pct_m.group(1))
            stake_pct = pct / 100.0 if pct <= 100 else 1.0
        elif frac_m:
            stake_pct = float(frac_m.group(1)) / float(frac_m.group(2))
        else:
            stake_pct = 1.0

        # Parse sale date from a single date column (or fallback to empty)
        sale_date = _parse_sale_date(_get(row, "sale_date_raw"), fallback_year=_date_header_year)

        records.append({
            "raw_description": desc,
            "property_name":   name,
            "address":         addr,
            "gfa_sf":          int(gfa) if gfa else None,
            "price_sgd_m":     price_m,
            "price_psf_gfa":   _get_num(row, "price_psf_gfa"),
            "remaining_yrs":   round(float(_parse_remaining_yrs(_get(row, "remaining_yrs")) or 0), 1),
            "adj_npi_yield":   _num(_get(row, "adj_npi_yield")),
            "npi_yield":       _num(_get(row, "npi_yield")),
            "sale_type":       sale_type,
            "buyer":           str(_get(row, "buyer") or "").strip(),
            "land_zoning":     str(_get(row, "land_zoning") or "").strip(),
            # Location-competitiveness label if the input already supplies one
            # (Superior/Comparable/Inferior). Respected by apply_location.
            "location":        str(_get(row, "location") or "").strip(),
            "sale_date":       sale_date,
            "stake_pct":       stake_pct,
            "_source":         "excel",
            "_price_unit":     _price_unit_detected,
            "_area_unit":      area_unit,
        })

    return records, col_map, headers, _price_unit_detected


# ─────────────────────────────────────────────────────────────────────────────
# PDF INPUT PARSING  (delegates to shared pdf_extractor.py pipeline)
# ─────────────────────────────────────────────────────────────────────────────

# Section keywords for Stage 1 page discovery — covers common section titles
# in IMs, broker reports, and appraisal reports.
_PDF_SECTION_KEYWORDS = [
    "Transaction Comparables", "Additional Transaction Comparables",
    "Sales Comparables", "Additional Sales Comparables",
    "Comparable Sales", "Comparable Transactions",
    "Capital Market Transactions", "Investment Sales",
    "Market Transactions", "Recent Sales",
    "Recent Transactions", "Sale Evidence",
    "Significant Sales", "Significant Transactions",
    "Notable Sales", "Notable Transactions",
    "Key Transactions", "Key Sales Transactions",
    "Investment Activity", "Investment Transactions",
    "Private Transactions",
    # Broker market-report tables (e.g. CBRE "TABLE 2: Major Transactions, Q4 2025")
    "Major Transactions", "Major Deals", "Major Sales",
    "Selected Transactions", "Selected Deals", "Selected Sales",
    "Notable Deals", "Key Deals", "Significant Deals",
]

# Full extraction schema for PDFs: same as _OUTPUT_FIELDS but with an
# extended property_name description that covers GLS tables where the site
# identifier column is labelled "LOCATION" or "SITE" rather than "Property".
_PDF_FIELD_SCHEMA = [
    (col, key,
     desc + ", Location, Site (GLS tables label the site as LOCATION)"
     if key == "property_name" else desc)
    for col, key, desc in _OUTPUT_FIELDS
] + [
    ("Address", "address", "Street address — e.g. 'Address', 'Street'"),
]


def _parse_pdf_records(pdf_path: str, llm_cfg: dict,
                       subject_name: str = "") -> list:
    """
    Extract asset sales comp records from a PDF using the shared 4-stage
    pdf_extractor pipeline (pdfplumber page discovery → table detection →
    field mapping → record assembly).

    Returns the same record format as parse_input_excel() so downstream
    classify_comps() and compute_metrics() work unchanged.
    """
    from pdf_extractor import extract_pdf_records

    # Header keywords that mark a table as GLS / government land-tender (a LAND
    # sale, not an asset/building sale). Such tables belong in the Land comps
    # tab and are excluded from the asset-sales analysis.
    _LAND_TABLE_MARKERS = [
        "type of development allowed", "development allowed",
        "successful tender", "tenderer", "psf ppr", "per plot ratio",
    ]

    raw_records = extract_pdf_records(
        pdf_path, _PDF_SECTION_KEYWORDS, _PDF_FIELD_SCHEMA,
        llm_cfg, subject_name=subject_name,
        reject_table_headers=_LAND_TABLE_MARKERS,
        extra_exclusion_note=(
            "Decide what a table is from its TITLE and COLUMN HEADERS only — never "
            "from the property use/sector of individual rows. (Both residential AND "
            "commercial land sales exist, so a row marked 'Commercial' does not make "
            "a land-tender table in scope.) "
            "EXTRACT from a table whose title/columns indicate the SALE or INVESTMENT "
            "TRANSACTION of built properties (e.g. title 'Investment Sales' / 'Notable "
            "Transactions', with a Buyer or Purchaser column). "
            "SKIP the whole table when its title or columns indicate a Government Land "
            "Sales (GLS) land tender — e.g. a title like 'Successful Tender' or columns "
            "like 'Successful Tenderer', 'Date of Award', or 'psf ppr' — or Leasing / "
            "Rental activity. Those are land or lease tables, not building sales, "
            "regardless of whether individual rows say Residential or Commercial."
        ),
    )
    if not raw_records:
        return []

    subj_tokens = set(re.sub(r"\W+", " ", subject_name.lower()).split()) if subject_name else set()

    records = []
    for item in raw_records:
        if not isinstance(item, dict):
            continue

        name  = str(item.get("property_name") or "").strip()
        addr  = str(item.get("address")        or "").strip()

        # GLS / land-tender tables use "LOCATION" as the site description —
        # pdfplumber maps it to `address` (no separate property_name column).
        # Fall back to address so these records are not silently dropped.
        if not name and addr:
            name = addr
        if not name:
            print(f"      SKIP (no name)  raw={str(item)[:80]!r}")
            continue

        if subj_tokens:
            meaningful_subj = {t for t in subj_tokens if not t.isdigit()}
            if meaningful_subj:
                desc_tokens = set(re.sub(r"\W+", " ", (name + " " + addr).lower()).split())
                needed = max(1, math.ceil(len(meaningful_subj) * 0.75))
                if len(desc_tokens & meaningful_subj) >= needed:
                    print(f"      SKIP (subject match)  {name!r:.60}")
                    continue

        gfa           = _num(item.get("gfa_sf"))
        price_raw_str = str(item.get("price_sgd_m") or "").strip()
        price_m       = _num(price_raw_str)
        if price_m is not None and price_m > 100_000:
            price_m = round(price_m / 1_000_000, 3)
        if not price_m:
            print(f"      SKIP (no price)  {name!r:.60}  price_raw={price_raw_str!r}")
            continue
        if _REQUIRE_GFA and not gfa:  # flip _REQUIRE_GFA at top of file to enforce
            continue

        # Date sanity: non-empty date with no 4-digit year is a sentence fragment.
        sale_date_str = str(item.get("sale_date_raw") or "").strip()
        if sale_date_str and not re.search(r"\b(?:19|20)\d{2}\b", sale_date_str):
            print(f"      SKIP (bad date)  {name!r:.60}  date={sale_date_str!r}")
            continue

        _range_m = re.search(r"\d+(?:\.\d+)?\s*[-–]\s*\d+", price_raw_str)
        price_display = re.sub(r"[*†#]+", "", price_raw_str).strip() if _range_m else None

        sale_type = str(item.get("sale_type") or "").strip()
        combined  = f"{sale_type} {name} {addr}"
        pct_m     = re.search(r"(\d+(?:\.\d+)?)\s*%", combined)
        frac_m    = re.search(r"\((\d+)\s*/\s*(\d+)\s*stake", combined, re.I)
        if any(kw in sale_type.lower() for kw in ("whole", "block sale", "en bloc")):
            stake_pct = 1.0
        elif pct_m:
            pct = float(pct_m.group(1))
            stake_pct = pct / 100.0 if pct <= 100 else 1.0
        elif frac_m:
            stake_pct = float(frac_m.group(1)) / float(frac_m.group(2))
        else:
            stake_pct = 1.0

        records.append({
            "raw_description":     f"{name}\n{addr}" if addr else name,
            "property_name":       name,
            "address":             addr,
            "gfa_sf":              int(gfa) if gfa else None,
            "price_sgd_m":         price_m,
            "price_psf_gfa":       _num(item.get("price_psf_gfa")),
            "price_sgd_m_display": price_display,
            "remaining_yrs":   round(float(_parse_remaining_yrs(item.get("remaining_yrs")) or 0), 1),
            "adj_npi_yield":   _num(item.get("adj_npi_yield")),
            "npi_yield":       _num(item.get("npi_yield")),
            "sale_type":       sale_type,
            "buyer":           str(item.get("buyer") or "").strip(),
            "land_zoning":     str(item.get("land_zoning") or "").strip(),
            # Preserve original PDF date text (e.g. "Feb 2026", "Q1 2026").
            # _parse_sale_date() is only for Excel where dates arrive as Python
            # datetime objects that need rendering as human-readable text.
            "sale_date":       str(item.get("sale_date_raw") or "").strip(),
            "stake_pct":       stake_pct,
            "_source":         "pdf",
        })

    print(f"  → {len(records)} valid records after filtering")
    return records


# ─────────────────────────────────────────────────────────────────────────────
# IMAGE / SCREENSHOT INPUT PARSING
# ─────────────────────────────────────────────────────────────────────────────

_IMAGE_EXTRACT_PROMPT = """\
You are extracting real estate asset sales comparable data from a table screenshot.

IMPORTANT: Count every data row visible in the table and extract ALL of them — do not stop after the first row.

Extract ALL rows from the table in the image. For each row return these fields:
{field_list}

Rules:
- Return ONLY a single valid JSON array containing ALL rows. No preamble, no explanation, no markdown fences.
- ALL rows must be in ONE array: [{{}}, {{}}, ...] — do NOT return separate arrays per row.
- If a field is not present for a row, use null.
- Numbers as numbers (not strings). Dates as strings.
- Do not invent or estimate values — only extract what is visible in the image.
- Skip header rows and average/total rows.
- Price ranges like "600-630" should be returned as the string "600-630" (not split).
- SCOPE — this is an ASSET SALES / INVESTMENT analysis. Decide what the table is from
  its TITLE and COLUMN HEADERS only — never from the property use/sector of individual
  rows (both residential AND commercial land sales exist, so a 'Commercial' row does
  not make a land-tender table in scope). EXTRACT when the title/columns indicate the
  SALE or INVESTMENT TRANSACTION of built properties (e.g. 'Investment Sales' / 'Notable
  Transactions', with a Buyer/Purchaser column). SKIP the whole table (return []) when
  the title or columns indicate a Government Land Sales (GLS) land tender — e.g. title
  'Successful Tender' or columns 'Successful Tenderer' / 'Date of Award' / 'psf ppr' —
  or Leasing / Rental activity.
- Extract raw numeric values exactly as shown — do NOT convert units. Python handles
  all unit conversion (sqm to sqft, SGD million to billion) after extraction.
"""


def _parse_image_records(image_path: str, llm_cfg: dict, openai_key: str = "",
                         subject_name: str = "") -> list:
    """
    Extract asset sales comp records from a table screenshot using a vision LLM.
    Returns the same record format as parse_input_excel().
    """
    print(f"  Reading image: {Path(image_path).name} ...")

    field_list = "\n".join(
        f'  "{key}": {desc}'
        for _, key, desc in _OUTPUT_FIELDS
    ) + '\n  "address": Street address of the property'

    prompt = _IMAGE_EXTRACT_PROMPT.format(field_list=field_list)

    # Retry up to 3 times; keep whichever attempt returns the most JSON objects.
    # All attempts always run — there is no early-stop threshold, because the
    # input may legitimately have fewer rows than any fixed threshold.
    _MAX_ATTEMPTS = 3
    best_raw: str = ""
    best_count: int = -1
    for attempt in range(1, _MAX_ATTEMPTS + 1):
        print(f"  Extracting via vision LLM (attempt {attempt}/{_MAX_ATTEMPTS}) ...")
        try:
            raw_attempt = _call_vision_llm(image_path, prompt, llm_cfg, openai_key)
        except Exception as e:
            print(f"  [warning] Vision LLM call failed on attempt {attempt}: {e}")
            continue
        # Count how many JSON objects are in this response (rough heuristic)
        obj_count = raw_attempt.count('"property_name"')
        print(f"  [debug]   attempt {attempt}: {len(raw_attempt)} chars, "
              f"~{obj_count} record(s) detected")
        if obj_count > best_count:
            best_count = obj_count
            best_raw   = raw_attempt

    if not best_raw:
        print("  [warning] All vision LLM attempts failed — no response.")
        return []

    raw = best_raw
    print(f"  [debug]   Best response ({len(raw)} chars, ~{best_count} record(s)):\n{raw[:1500]}")

    # Parse JSON array from response
    m = re.search(r"```(?:json)?\s*([\s\S]*?)```", raw)
    if m:
        raw = m.group(1).strip()
    m2 = re.search(r"\[[\s\S]*\]", raw)
    if not m2:
        print("  [warning] No JSON array found in vision LLM response.")
        print(f"  [debug]   Raw response: {raw[:400]}")
        return []
    raw_json = m2.group(0)

    def _try_parse(s: str):
        """Try JSON → fixed JSON → Python literal eval, return list or None."""
        import ast as _ast
        # Pre-convert JSON keywords → Python equivalents for ast.literal_eval.
        # minicpm-v often mixes single-quoted strings (Python) with null/true/false
        # (JSON), so neither parser handles it as-is without this normalisation.
        _py = re.sub(r'\bnull\b', 'None',
               re.sub(r'\btrue\b', 'True',
               re.sub(r'\bfalse\b', 'False', s)))
        for attempt in (
            lambda: json.loads(s),
            lambda: json.loads(_fix_json(s)),
            lambda: _ast.literal_eval(_py),  # single-quoted strings + null/true/false
        ):
            try:
                r = attempt()
                if isinstance(r, list):
                    return r
            except Exception:
                pass
        return None

    extracted = _try_parse(raw_json)
    if extracted is None:
        # Last resort — vision LLM may have returned one array per record on
        # separate lines ("Extra data" error).  Find every top-level array and merge.
        all_arrays = _split_json_arrays(raw)
        if len(all_arrays) > 1:
            print(f"  [info]    Found {len(all_arrays)} separate JSON arrays — merging …")
            merged = []
            for arr_str in all_arrays:
                arr = _try_parse(arr_str)
                if arr:
                    merged.extend(arr)
            if merged:
                extracted = merged
            else:
                print(f"  [warning] Could not parse vision LLM JSON (all strategies failed).")
                print(f"  [debug]   Raw response (first 500 chars):\n{raw[:500]}")
                return []
        else:
            print(f"  [warning] Could not parse vision LLM JSON (all strategies failed).")
            print(f"  [debug]   Raw response (first 500 chars):\n{raw[:500]}")
            return []
    if not isinstance(extracted, list):
        print("  [warning] Vision LLM did not return a JSON array.")
        return []

    print(f"  → {len(extracted)} raw records extracted from image")

    subj_tokens = set(re.sub(r"\W+", " ", subject_name.lower()).split()) if subject_name else set()

    records = []
    for item in extracted:
        if not isinstance(item, dict):
            continue
        name  = str(item.get("property_name") or "").strip()
        addr  = str(item.get("address")        or "").strip()
        if not name:
            continue
        if subj_tokens:
            desc_tokens = set(re.sub(r"\W+", " ", (name + " " + addr).lower()).split())
            if len(desc_tokens & subj_tokens) >= max(2, len(subj_tokens) * 0.75):
                continue
        gfa           = _num(item.get("gfa_sf"))
        price_raw_str = str(item.get("price_sgd_m") or "").strip()
        price_m       = _num(price_raw_str)
        if price_m is not None and price_m > 100_000:
            price_m = round(price_m / 1_000_000, 3)
        if not price_m:
            continue
        if _REQUIRE_GFA and not gfa:  # flip _REQUIRE_GFA at top of file to enforce
            continue
        # Preserve range strings like "600-630*" for Excel display.
        # Strip footnote markers (*/**) but keep the low–high range.
        # The numeric midpoint (price_m) is stored separately for calculations.
        _range_m = re.search(r"\d+(?:\.\d+)?\s*[-–]\s*\d+", price_raw_str)
        price_display = re.sub(r"[*†#]+", "", price_raw_str).strip() if _range_m else None

        sale_type = str(item.get("sale_type") or "").strip()
        combined = f"{sale_type} {name} {addr}"
        pct_m  = re.search(r"(\d+(?:\.\d+)?)\s*%", combined)
        frac_m = re.search(r"\((\d+)\s*/\s*(\d+)\s*stake", combined, re.I)
        if any(kw in sale_type.lower() for kw in ("whole", "block sale", "en bloc")):
            stake_pct = 1.0
        elif pct_m:
            pct = float(pct_m.group(1))
            stake_pct = pct / 100.0 if pct <= 100 else 1.0
        elif frac_m:
            stake_pct = float(frac_m.group(1)) / float(frac_m.group(2))
        else:
            stake_pct = 1.0
        records.append({
            "raw_description":  f"{name}\n{addr}" if addr else name,
            "property_name":    name,
            "address":          addr,
            "gfa_sf":           int(gfa) if gfa else None,
            "price_sgd_m":      price_m,
            "price_psf_gfa":    _num(item.get("price_psf_gfa")),
            "price_sgd_m_display": price_display,  # "600-630" range string for Excel
            "remaining_yrs":   round(float(_parse_remaining_yrs(item.get("remaining_yrs")) or 0), 1),
            "adj_npi_yield":   _num(item.get("adj_npi_yield")),
            "npi_yield":       _num(item.get("npi_yield")),
            "sale_type":       sale_type,
            "buyer":           str(item.get("buyer") or "").strip(),
            "land_zoning":     str(item.get("land_zoning") or "").strip(),
            # Preserve original text from image (e.g. "Feb 2026", "Q1 2026")
            "sale_date":       str(item.get("sale_date_raw") or "").strip(),
            "stake_pct":       stake_pct,
            "_source":         "image",
        })

    print(f"  → {len(records)} valid records after filtering")
    return records


# ─────────────────────────────────────────────────────────────────────────────
# COLUMN MAPPING DOCUMENTATION TAB
# ─────────────────────────────────────────────────────────────────────────────

def _write_mapping_tab(excel_path: str, col_map: dict, input_headers: list,
                       input_file: str):
    """
    Add (or replace) a 'Column Mapping' sheet in the output Excel documenting
    which input column was used for each output column, and any substitutions
    (e.g. NLA used in place of GFA).
    """
    # Build list of (output_col, input_col, note) rows
    rows = []

    # Output fields
    for out_col, int_key, _ in _OUTPUT_FIELDS:
        idx = col_map.get(int_key)
        if idx is not None and idx < len(input_headers):
            in_col = input_headers[idx]
        else:
            in_col = "(not found in input)"

        note = ""
        # Flag name mismatches that may imply a proxy / substitution
        if idx is not None and in_col != "(not found in input)":
            out_tokens = set(re.sub(r"[^a-z0-9]", " ", out_col.lower()).split())
            in_tokens  = set(re.sub(r"[^a-z0-9]", " ", in_col.lower()).split())
            if not (out_tokens & in_tokens):          # no word in common
                note = f"Input column name differs — '{in_col}' used as proxy for '{out_col}'"
            elif out_col == "Sale Date":
                note = "Value converted to Q+Year format (e.g. Q1 2024)"
        else:
            note = "Column not mapped — field left blank in output"

        rows.append((out_col, in_col, note))

    # Address (extra field combined into Property)
    addr_idx = col_map.get("address")
    if addr_idx is not None and addr_idx < len(input_headers):
        rows.append(("Property (address line)",
                     input_headers[addr_idx],
                     "Address appended to Property name in output"))

    # Write tab
    wb = openpyxl.load_workbook(excel_path)
    if "Column Mapping" in wb.sheetnames:
        del wb["Column Mapping"]
    ws = wb.create_sheet("Column Mapping")

    # Header row
    ws.append(["Output Column (Report)", "Source Input Column", "Notes"])
    for r in rows:
        ws.append(list(r))
    ws.append([])
    ws.append(["Source file", input_file])

    # Basic styling
    from openpyxl.styles import Font, PatternFill, Alignment
    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 60

    wb.save(excel_path)
    print(f"  Column Mapping tab → {excel_path}")


# ─────────────────────────────────────────────────────────────────────────────
# CLASSIFY  (Ollama with rules fallback)
# ─────────────────────────────────────────────────────────────────────────────

def _system_msg(country_name: str) -> str:
    return (
        f"You are a senior {country_name} commercial real estate analyst supporting "
        "institutional deal underwriting. You classify comparable investment "
        "transactions and assess how relevant each is to a given subject property. "
        'Return your answer as a JSON object with key "comparables" containing an array — no markdown fences.'
    )


_CLASSIFY_PROMPT = """\
SUBJECT PROPERTY:
{subject_json}

RAW COMPARABLES (0-indexed):
{comps_json}

Classify each comparable in the list above. For each one assign:
  asset_type     — "<Sale Structure> (<Use>)" ONLY if both the sale structure and
                   property use can be determined from the input data.
                   e.g. "Block Sale (Office)", "En Bloc (Mixed-Use)".
                   Return "" if the sale structure or property use is not stated.
  relevance_score — integer 1–10; higher = more comparable to subject

Return ONLY: {{"comparables": [{{"index":<int>,"asset_type":"...","relevance_score":<int>}},...]}}\
"""


_LOC_LABELS = {"superior", "comparable", "inferior"}


def _provided_location(rc: dict) -> str:
    """Return the input-provided location-competitiveness label
    (Superior/Comparable/Inferior) if the source file supplied one, else "".
    Used so a user-provided Location column survives classification/validation
    instead of being blanked (Location is otherwise never inferred for sales)."""
    v = str(rc.get("location") or "").strip()
    return v.capitalize() if v.lower() in _LOC_LABELS else ""


def _parse_classify_response(text: str, raw_comps: list) -> list:
    text   = re.sub(r"^```[a-z]*\n?", "", text.strip())
    text   = re.sub(r"\n?```$", "", text)
    parsed = json.loads(text)
    if isinstance(parsed, dict):
        parsed = parsed.get("comparables") or next(
            (v for v in parsed.values() if isinstance(v, list)), [])
    if not isinstance(parsed, list):
        raise ValueError(f"Expected list, got {type(parsed).__name__}")

    # Index the LLM's classifications so we can attach them to the original
    # comps.  The output always has one entry per input comp — comps the model
    # skipped (small models often drop the tail of a list) keep blank
    # classification fields rather than vanishing from the result.
    by_idx = {}
    for cls in parsed:
        idx = cls.get("index")
        if isinstance(idx, int) and 0 <= idx < len(raw_comps):
            by_idx[idx] = cls

    augmented = []
    for i, rc in enumerate(raw_comps):
        cls = by_idx.get(i, {})
        c = dict(rc)
        c.update({
            "location":        _provided_location(rc) or cls.get("location", ""),
            "quality":         cls.get("quality", ""),
            "asset_type":      cls.get("asset_type", ""),
            "relevance_score": int(cls.get("relevance_score") or 5),
        })
        augmented.append(c)

    for i, c in enumerate(augmented, 1):
        c["map_marker"] = str(i)
    return augmented


def _classify_llm(raw_comps: list, subject_cfg: dict, llm_cfg: dict) -> list:
    slim   = [{k: c.get(k) for k in ("raw_description", "sale_date", "gfa_sf",
               "price_sgd_m", "remaining_yrs", "land_zoning", "sale_type", "stake_pct")}
              for c in raw_comps]
    prompt = _CLASSIFY_PROMPT.format(
        subject_json = json.dumps(subject_cfg, indent=2),
        comps_json   = json.dumps(slim, indent=2),
    )
    country  = subject_cfg.get("country_name", "Singapore")
    messages = [{"role": "system", "content": _system_msg(country)},
                {"role": "user",   "content": prompt}]
    provider = (llm_cfg or {}).get("provider", "ollama")
    if provider == "openai":
        from tools.llm_client import openai_chat as _openai_chat
        text = _openai_chat(llm_cfg, messages, json_mode=True)
    else:
        ocfg     = (llm_cfg or {}).get("ollama", {})
        base_url = ocfg.get("base_url", "http://localhost:11434")
        model    = ocfg.get("model",    "qwen2.5:3b")
        text     = _ollama_post(base_url, model, messages)
    return _parse_classify_response(text, raw_comps)


def _classify_rules(raw_comps: list, subject_cfg: dict = None) -> list:
    """
    Fallback when Ollama is unavailable.
    Location and Quality are never inferred — they are always left blank and
    filled in manually by the analyst.  Asset Type is built from sale_type
    (source data) + asset_class (deal config) when sale_type is present.
    """
    asset_class = (subject_cfg or {}).get("asset_class", "office").title()

    augmented = []
    for c in raw_comps:
        stype  = str(c.get("sale_type")   or "")
        zoning = str(c.get("land_zoning") or "")

        # Asset Type: only when sale_type is explicitly in the source record
        base = re.sub(r"\s*\(.*\)", "", stype).strip()
        if base:
            if "Residential" in zoning:
                use = f"Mixed Use ({asset_class} + Retail)"
            elif "Hospitality" in zoning:
                use = f"Mixed Use ({asset_class} + Hotel)"
            else:
                use = asset_class
            asset_type = f"{base} ({use})"
        else:
            asset_type = ""

        d = dict(c)
        d.update(location=_provided_location(c), quality="",
                 asset_type=asset_type, relevance_score=5)
        augmented.append(d)

    for i, c in enumerate(augmented, 1):
        c["map_marker"] = str(i)
    return augmented


def classify_comps(raw_comps: list, subject_cfg: dict, llm_cfg: dict) -> list:
    """Try LLM (GPT or Ollama based on provider) first; fall back to keyword rules on any error."""
    provider = (llm_cfg or {}).get("provider", "ollama")
    if provider in ("none", "rules"):
        print("  [LLM] Rule-based mode — classifying comps with keyword rules (no LLM).")
        return _classify_rules(raw_comps, subject_cfg)
    if provider == "openai":
        model_label = (llm_cfg or {}).get("openai_model", "gpt-4o-mini")
    else:
        model_label = (llm_cfg or {}).get("ollama", {}).get("model", "qwen2.5:3b")
    try:
        print(f"  [LLM] {'GPT' if provider == 'openai' else 'Ollama'} ({model_label}) …")
        result = _classify_llm(raw_comps, subject_cfg, llm_cfg)
        print(f"  [LLM] OK — {len(result)} comps classified.")
        return result
    except Exception as exc:
        print(f"  [LLM] Failed ({exc.__class__.__name__}: {exc}). Falling back to rules.")
        return _classify_rules(raw_comps, subject_cfg)


# ─────────────────────────────────────────────────────────────────────────────
# CALCULATE  (Bala-adjusted cap rate)
# ─────────────────────────────────────────────────────────────────────────────

def compute_metrics(comps: list, subject_cfg: dict) -> list:
    """Add price_psf_gfa, ftm_cap_rate, adj_cap_rate to every comp dict.
    Adj Cap Rate uses the official Singapore Bala Table lookup.
    price_sgd_m is always stored in millions (M-normalized), so we always use 1e6.
    """
    # The Bala Table is Singapore-specific (SLA/SISV leasehold valuation). It does
    # not apply to foreign markets, so skip the adjustment for non-SG deals.
    _global  = str(subject_cfg.get("country_name", "")).lower() not in ("", "singapore")
    subj_yrs = subject_cfg.get("remaining_leasehold_yrs", 0) or 0
    out = []
    for c in comps:
        d       = dict(c)
        price_m = float(d.get("price_sgd_m") or 0)
        stake   = float(d.get("stake_pct")   or 1.0)
        gfa     = d.get("gfa_sf")            # may be None when _REQUIRE_GFA = False
        rem_yrs = int(d.get("remaining_yrs") or 0)
        ftm_cr  = float(d.get("npi_yield") or 0)

        d["price_psf_gfa"] = round((price_m / stake) * 1e6 / gfa) if gfa else d.get("price_psf_gfa")
        d["ftm_cap_rate"]  = ftm_cr
        if _global or not bala_factor(subj_yrs):
            d["adj_cap_rate"] = ftm_cr       # no SG Bala adjustment for foreign deals
        else:
            d["adj_cap_rate"] = (ftm_cr
                                 * bala_factor(rem_yrs)
                                 / bala_factor(subj_yrs))
        out.append(d)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# GEOCODING
# ─────────────────────────────────────────────────────────────────────────────

# Generic asset-class / sector words that sometimes leak into the address column
# (e.g. a "PROPERTY TYPE" column with values Office/Logistics/Hospitality). These
# are NOT geocodable addresses — geocoding must ignore them and fall back to the
# property name, otherwise every comp of the same type stacks on one point.
_NON_ADDRESS_WORDS = {
    "office", "logistics", "retail", "industrial", "hospitality", "hotel",
    "residential", "commercial", "mixed", "mixed-use", "mixed use",
    "warehouse", "business park", "data centre", "data center", "others",
    # Placeholder / "no address" values — geocoding these returns the country
    # centroid, stacking every such comp on one point. Fall back to the name.
    "n/a", "na", "n.a.", "n.a", "nil", "none", "tbd", "-", "--", "—",
    "not available", "not applicable", "not appl.", "unknown", ".",
}


def _geocode_comps(records: list, mapbox_tok: str,
                   country_code: str, country_name: str,
                   s_lon: float, s_lat: float) -> list:
    """Geocode each comp; attach lon / lat / distance_km.

    Tries address first (more precise), then property name as fallback.

    Records that already have lon/lat (loaded from _records.json on a
    --from-records re-run) are skipped — distance is recalculated but no
    Mapbox API call is made.  Only new/ungeocoded rows call the API.
    """
    suffix = f", {country_name}" if country_name else ""
    for r in records:
        name = str(r.get("property_name") or "").strip()

        # ── Already geocoded — reuse saved coordinates ────────────────────────
        if r.get("lon") is not None and r.get("lat") is not None:
            dist = _haversine_km(r["lon"], r["lat"], s_lon, s_lat)
            r["distance_km"] = dist
            # Preserve existing geo status fields from cached records
            r.setdefault("_geo_provider", "cached")
            r.setdefault("_geo_note", "cached")
            print(f"      {name[:50]:<50}  {dist:>5.2f} km  (cached)")
            continue

        # ── Not yet geocoded — call geocoder ─────────────────────────────────
        addr = str(r.get("address") or "").strip()
        # Clean newlines (can appear when stake/descriptor is on a second line)
        name = name.replace("\n", " ").replace("\r", " ").strip()
        addr = addr.replace("\n", " ").replace("\r", " ").strip()

        if not name and not addr:
            r["lon"], r["lat"], r["distance_km"] = None, None, 9999.0
            r["_geo_provider"] = "failed"
            r["_geo_note"]     = "no address or name"
            print(f"      {'(no name)':<50}  NOT PLOTTED — no address or name")
            continue

        # Geocoding priority: if an Address is present (the analyst can fill it in
        # the preview when a property name is too rough to geocode), use it — it
        # overrides the name. Address-first, with the name kept as a fallback.
        # If Address is blank, geocode by the property name as before.
        # Country suffix helps foreign-address geocoding (Mapbox especially).
        _sfx = (f", {country_name}"
                if country_name and country_name.strip().lower() not in ("", "singapore")
                else "")
        # Use the address only if it looks like a real location — a bare
        # asset-class word (Office/Logistics/…) that slipped into the address
        # column must NOT be geocoded, or all same-type comps stack on one point.
        if addr and addr.strip().lower() not in _NON_ADDRESS_WORDS:
            # Address present → geocode by ADDRESS ONLY (no name fallback).
            # Falling back to the name collapses distinct properties that share a
            # brand: "Weave Place – Hoegi" and "Weave Place – Gangnam Station" both
            # strip to "Weave Place" → same pin. Address-only keeps them distinct.
            queries = _build_geocode_queries("", addr, _sfx)
            source  = "address"
        else:
            # Geocode by NAME, adding the submarket/district as a locality HINT
            # (backend only — never shown) so a weak building name still resolves
            # to the right area instead of failing or hitting the country centroid.
            _hint = str(r.get("submarket") or "").strip()
            queries = _build_geocode_queries(name, _hint, _sfx)
            source  = "name"

        try:
            lon, lat, geo_note = geocode_with_fallbacks(queries, mapbox_tok, country_code)
            dist = _haversine_km(lon, lat, s_lon, s_lat)
            r["lon"], r["lat"], r["distance_km"] = lon, lat, dist
            r["_geo_provider"] = geo_note
            r["_geo_note"]     = geo_note
            tag = " (by name)" if source == "name" else ""
            # Flag likely-invalid geocodes that landed on the country centroid.
            if _near_country_centroid(lon, lat, country_code):
                r["_geo_suspect"] = True
                print(f"      {name[:50]:<50}  {dist:>5.2f} km{tag}  "
                      f"⚠ ON COUNTRY CENTROID — likely invalid, VERIFY (flagged for review)")
            else:
                print(f"      {name[:50]:<50}  {dist:>5.2f} km{tag}")
        except Exception as exc:
            r["lon"], r["lat"], r["distance_km"] = None, None, 9999.0
            r["_geo_provider"] = "failed"
            r["_geo_note"]     = str(exc)
            print(f"      {name[:50]:<50}  NOT PLOTTED — {exc}")
    return records


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def run(config_path: str = "configs/deal_config.json",
        generate_map: bool = False,
        from_records: str = None,
        refinement_file: str = None,
        name_only: bool = False,
        refine_engine: str = None):

    global _NAME_ONLY
    _NAME_ONLY = name_only

    with open(config_path, encoding="utf-8") as f:
        cfg = json.load(f)

    subject_cfg  = cfg["subject_property"]
    mb_cfg       = cfg.get("mapbox", {})

    # Mapbox token — always read from shared_settings.json (single source of truth).
    # Fall back to the deal config's mapbox.token only if shared_settings is absent.
    _ss_path = Path(__file__).parent.parent / "configs" / "shared_settings.json"
    try:
        _ss = json.loads(_ss_path.read_text(encoding="utf-8"))
        mapbox_tok   = _ss.get("mapbox_token", "")
        onemap_token = _ss.get("onemap_token", "")
    except Exception:
        mapbox_tok   = mb_cfg.get("token", "")
        onemap_token = ""
    llm_cfg      = cfg.get("llm", {"provider": "ollama",
                                    "ollama": {"base_url": "http://localhost:11434",
                                               "model": "qwen2.5:3b"}})
    # Per-run refinement engine override (from the frontend radio / CLI).
    if refine_engine:
        llm_cfg = {**llm_cfg, "refine_engine": refine_engine}
    ollama_cfg   = llm_cfg.get("ollama", {})
    base_url     = ollama_cfg.get("base_url", "http://localhost:11434")
    model        = ollama_cfg.get("model",    "qwen2.5:3b")
    country_name = subject_cfg.get("country_name", "")
    # country_code drives the geocoder's country restriction. Prefer an explicit code,
    # else DERIVE it from country_name (so a deal with only the name still geocodes to
    # the right country instead of a same-named place abroad). No Singapore hardcode.
    country_code = (cfg.get("country_code") or subject_cfg.get("country_code")
                    or _cc_from_name(country_name) or "")
    prop_name    = subject_cfg["property_name"]
    deal_name    = subject_cfg.get("deal_name", prop_name)
    # Strip characters Windows forbids in file paths (< > : " / \ | ? * and control
    # chars) — an unsanitised name causes [Errno 22] Invalid argument on Windows.
    deal_slug    = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", deal_name).strip(" .").replace(" ", "_") or "deal"

    # ── Paths ─────────────────────────────────────────────────────────────────
    # Supports Excel (.xlsx), PDF (.pdf), and/or Image (.png/.jpg) for one deal.
    # input_file       → Excel source  (config key: "input_file")
    # input_pdf_file   → PDF source    (config key: "input_pdf_file")
    # input_image_file → Image/screenshot source (config key: "input_image_file")
    # Records from all sources are merged before classification.
    _xl_cfg  = cfg.get("input_file")
    input_excel_files = ([_xl_cfg] if isinstance(_xl_cfg, str) else list(_xl_cfg)) if _xl_cfg else []
    input_file        = input_excel_files[0] if input_excel_files else None
    _pdf_cfg          = cfg.get("input_pdf_file", [])
    input_pdf_files   = [_pdf_cfg] if isinstance(_pdf_cfg, str) else list(_pdf_cfg)
    input_pdf_file    = input_pdf_files[0] if input_pdf_files else None
    _img_cfg  = cfg.get("input_image_file")
    input_image_files = ([_img_cfg] if isinstance(_img_cfg, str) else list(_img_cfg)) if _img_cfg else []
    input_image_file  = input_image_files[0] if input_image_files else None

    if not from_records:
        if not input_excel_files and not input_pdf_files and not input_image_files:
            raise ValueError(
                "No input file found in config.\n"
                "Add:  \"input_file\": \"Input_files/your_sales_comps.xlsx\"  (Excel)\n"
                "  or  \"input_pdf_file\": \"Input_files/your_report.pdf\"    (PDF)\n"
                "  or  \"input_image_file\": \"Input_files/table_screenshot.png\"  (Image)\n"
                "  or any combination for a mixed-source deal."
            )
        for _xf in input_excel_files:
            if not Path(_xf).exists():
                raise FileNotFoundError(f"Input Excel not found: {_xf}")
        for _pf in input_pdf_files:
            if not Path(_pf).exists():
                raise FileNotFoundError(f"Input PDF not found: {_pf}")
        for _imf in input_image_files:
            if not Path(_imf).exists():
                raise FileNotFoundError(f"Input image not found: {_imf}")

    output_file = cfg.get("output_file",
                          f"output/{deal_slug}/Transaction_Comparables_{deal_slug}.xlsx")
    out_dir    = Path(output_file).parent
    out_dir.mkdir(parents=True, exist_ok=True)
    # Derive map + geo sidecar names from the actual Excel stem so they always
    # match regardless of whether the config has a custom output_file path.
    excel_stem = Path(output_file).stem                          # e.g. "Transaction_Comparables_88_Cecil"
    out_map    = str(out_dir / f"{excel_stem}_map.png")
    out_geo    = str(out_dir / f"{excel_stem}_geo.json")

    print(f"\n{'='*64}")
    print(f"  Asset Sales Comps : {deal_name}")
    print(f"{'='*64}")
    for _i, _xf in enumerate(input_excel_files, 1):
        _lbl = f" {_i}" if len(input_excel_files) > 1 else ""
        print(f"  Excel{_lbl} → {_xf}")
    for _i, _pf in enumerate(input_pdf_files, 1):
        _lbl = f" {_i}" if len(input_pdf_files) > 1 else ""
        print(f"  PDF{_lbl}   → {_pf}")
    for _i, _imf in enumerate(input_image_files, 1):
        _lbl = f" {_i}" if len(input_image_files) > 1 else ""
        print(f"  Image{_lbl} → {_imf}")
    print(f"  Output → {output_file}")

    # ── 1. Geocode subject (only if mapbox token present) ─────────────────────
    s_lon = s_lat = None
    if mapbox_tok:
        address = subject_cfg.get("address", "")
        print(f"\n[1/5] Geocoding subject property")
        try:
            s_lon, s_lat, _ = geocode_with_fallbacks(
                [f"{prop_name}, {address}", address, prop_name],
                mapbox_tok, country_code,
            )
            print(f"      {prop_name}  →  ({s_lon:.5f}, {s_lat:.5f})")
        except Exception as e:
            print(f"      Geocoding failed: {e}  (distance sorting skipped)")
            s_lon = s_lat = None
    else:
        print(f"\n[1/5] No Mapbox token — geocoding skipped (comps ranked by relevance)")

    # ── 2. Parse input files (Excel, PDF, and/or Image) ──────────────────────
    out_records = str(out_dir / f"{excel_stem}_records.json")
    records = []
    _col_map = {}
    _input_headers = []

    if from_records:
        # ── Short-circuit: load previously saved records JSON ─────────────────
        # Skip all extraction (Ollama column mapping, PDF/image parsing) and
        # jump straight to Stage 3 classification.  This is the fast path for:
        #   • Manually adding or editing comps after the first run
        #   • Re-running with a different LLM model without re-parsing files
        print(f"\n[2/5] Loading extracted records from  {from_records}")
        try:
            with open(from_records, encoding="utf-8") as rf:
                records = json.load(rf)
            if not isinstance(records, list):
                raise ValueError("Expected a JSON array of records.")
            print(f"      → {len(records)} records loaded from JSON")
        except Exception as e:
            raise ValueError(f"Cannot load records JSON ({from_records}): {e}") from e
    else:
        # ── Normal path: parse all configured input files ─────────────────────
        print(f"\n[2/5] Parsing input files  (exact → embedding → LLM column mapping)")
        openai_key = llm_cfg.get("openai_api_key", "")

        for _xl_i, _xl_path in enumerate(input_excel_files, 1):
            _src_label = f"excel_{_xl_i}" if len(input_excel_files) > 1 else "excel"
            _tag = f" {_xl_i}" if len(input_excel_files) > 1 else ""
            print(f"  [Excel{_tag}] {Path(_xl_path).name}")
            _is_global_early = country_name.lower() not in ("", "singapore")
            _area_unit = subject_cfg.get("area_unit", "sqm") if _is_global_early else "sf"
            _excel_records, _xl_col_map, _xl_headers, _xl_pu = parse_input_excel(
                _xl_path, base_url, model, subject_name=prop_name,
                llm_cfg=llm_cfg, area_unit=_area_unit)
            # Auto-propagate detected price unit (e.g. "B") into subject config
            # so _m_to_display uses the right scale when rendering the global table.
            if _xl_pu and not subject_cfg.get("price_unit"):
                subject_cfg["price_unit"] = _xl_pu
            for _r in _excel_records:
                _r["_source"] = _src_label
            if not _col_map:
                _col_map, _input_headers = _xl_col_map, _xl_headers
            records += _excel_records
            print(f"      → {len(_excel_records)} records from Excel{_tag}")

        for _pdf_i, _pdf_path in enumerate(input_pdf_files, 1):
            _src_label = f"pdf_{_pdf_i}" if len(input_pdf_files) > 1 else "pdf"
            _tag = f" {_pdf_i}" if len(input_pdf_files) > 1 else ""
            print(f"  [PDF{_tag}] {Path(_pdf_path).name}")
            try:
                _pdf_records = _parse_pdf_records(
                    _pdf_path, llm_cfg, subject_name=prop_name)
            except Exception as _e:
                print(f"      [PDF{_tag}] extraction failed: {_e}  — skipping this file")
                _pdf_records = []
            for _r in _pdf_records:
                _r["_source"] = _src_label
            records += _pdf_records
            print(f"      → {len(_pdf_records)} records from PDF{_tag}")

        for _img_i, _img_path in enumerate(input_image_files, 1):
            _src_label = f"image_{_img_i}" if len(input_image_files) > 1 else "image"
            _tag = f" {_img_i}" if len(input_image_files) > 1 else ""
            print(f"  [Image{_tag}] {Path(_img_path).name}")
            _img_records = _parse_image_records(
                _img_path, llm_cfg, openai_key=openai_key, subject_name=prop_name)
            for _r in _img_records:
                _r["_source"] = _src_label
            records += _img_records
            print(f"      → {len(_img_records)} records from Image{_tag}")

        print(f"      → {len(records)} eligible transactions found (combined).")

        # Submarket hygiene: it is a BACKEND geocoding hint only — never shown in the
        # Property, Address or Land-Zoning columns. Strip a stacked 'Name⏎Submarket'
        # off the name, and blank an address / zoning cell that merely repeats the
        # submarket. The submarket value is kept on the record (r["submarket"]) for the
        # geocoder below.
        for _r in records:
            _sub = str(_r.get("submarket") or "").strip()
            for _nk in ("property_name", "property"):
                if _r.get(_nk):
                    _r[_nk] = _clean_name(_r[_nk])
            for _bad in ("address", "land_zoning"):
                _v = str(_r.get(_bad) or "").strip()
                if _sub and _v and _v.lower() == _sub.lower():
                    _r[_bad] = ""

        # Backfill blank dates from each input file's report period (title / cover page),
        # e.g. a comp with no row-level date in a "…Q2 2025…" report → "~Q2 2025".
        try:
            from tools.report_period import backfill_missing_dates
            _nfill = backfill_missing_dates(records, "sale_date",
                                            input_excel_files, input_pdf_files, input_image_files)
            if _nfill:
                print(f"      → filled {_nfill} blank sale date(s) from the report period")
        except Exception as _e:
            print(f"      [date-backfill] skipped: {_e}")

        # ── Save records to JSON sidecar ──────────────────────────────────────
        if records:
            with open(out_records, "w", encoding="utf-8") as rf:
                json.dump(records, rf, indent=2, default=str)
            print(f"  Records → {out_records}")

    if not records:
        print("  No eligible transactions found (need property description + price).")
        return

    # ── Auto-detect price unit from input column headers ──────────────────────
    # Prefer "B" if any source detected billions; fall back to "M".
    # This overrides the deal config so the output column header matches the input.
    _detected_pu = next(
        (r["_price_unit"] for r in records if r.get("_price_unit") == "B"), "M"
    )
    if _detected_pu != subject_cfg.get("price_unit", "M"):
        subject_cfg["price_unit"] = _detected_pu
        print(f"  [Price unit] Auto-detected '{_detected_pu}' from input column headers")

    # ── Apply analyst refinement instructions (only when --refinement-file given) ──
    if refinement_file:
        print(f"\n  Applying refinement instructions from {refinement_file} ...")
        instructions = Path(refinement_file).read_text(encoding="utf-8").strip()
        # Augment records with current map_marker from geo sidecar so the LLM
        # can match instructions like "remove map marker 11" or "keep only 1-5".
        _geo_path = Path(out_geo)
        if _geo_path.exists():
            try:
                _geo = json.loads(_geo_path.read_text(encoding="utf-8"))
                _marker_by_prop = {
                    c["property"].lower().strip(): c["map_marker"]
                    for c in _geo.get("comps", [])
                    if c.get("property") and c.get("map_marker")
                }
                for rec in records:
                    pn = str(rec.get("property_name") or "").lower().strip()
                    if pn in _marker_by_prop:
                        rec["_map_marker"] = _marker_by_prop[pn]
            except Exception:
                pass
        # Ensure distance_km exists so "within X km" instructions can be applied.
        # Refinement runs before geocoding, and --from-records rows carry lon/lat
        # but not a persisted distance_km — so recompute it from the subject coords.
        if s_lon is not None and s_lat is not None:
            for _rec in records:
                if (_rec.get("distance_km") in (None, "", 9999.0)
                        and _rec.get("lon") not in (None, "")
                        and _rec.get("lat") not in (None, "")):
                    try:
                        _rec["distance_km"] = round(_haversine_km(
                            float(_rec["lon"]), float(_rec["lat"]), s_lon, s_lat), 2)
                    except Exception:
                        pass
        _ci_subject = {k: subject_cfg.get(k) for k in
                       ("property_name", "deal_name", "asset_class", "address",
                        "gfa_sf", "price_sgd_m", "land_zoning", "country_name")
                       if subject_cfg.get(k) not in (None, "")}
        records = _apply_refinement(records, instructions, llm_cfg, subject=_ci_subject)
        for rec in records:
            rec.pop("_map_marker", None)
        if not records:
            print("  Refinement resulted in 0 records — nothing to output.")
            return
        with open(out_records, "w", encoding="utf-8") as _rf:
            json.dump(records, _rf, indent=2, default=str)
        print(f"  Refined records saved → {out_records}  ({len(records)} kept)")

    # ── 3. Classify via LLM ──────────────────────────────────────────────────
    _provider_label = "GPT" if llm_cfg.get("provider") == "openai" else "Ollama"
    print(f"\n[3/5] Classifying via {_provider_label}  (asset type / relevance)")
    classified = classify_comps(records, subject_cfg, llm_cfg)
    print(f"      → {len(classified)} comparables selected.")

    # ── Post-classification field validation ─────────────────────────────────
    # Location and Quality are NEVER derived from input data in sales comps —
    # the source columns are Property, Address, GFA, Price, Sale Type, etc.
    # There is no Location column and no Quality column.  Any value the LLM
    # assigned came from world knowledge, not the user's file.
    # Asset Type is only meaningful when sale_type was explicitly in the source.
    for c in classified:
        # Keep a user-provided Superior/Comparable/Inferior label; otherwise Location
        # is not inferred from sales input data (apply_location computes it from coords).
        c["location"] = _provided_location(c)
        c["quality"]  = ""
        # Asset Type: only keep when sale_type was explicitly in the source record.
        # The LLM builds asset_type as "<sale_type> (<use>)"; without a sale_type
        # the LLM has no factual basis for the asset structure.
        if not str(c.get("sale_type") or "").strip():
            c["asset_type"] = ""

    # ── 4. Calculate metrics + geocode + sort ─────────────────────────────────
    _is_global = country_name.lower() not in ("", "singapore")
    print(f"\n[4/5] Calculating metrics")
    processed = compute_metrics(classified, subject_cfg)

    if mapbox_tok and s_lon is not None:
        print(f"      Geocoding comparables …")
        processed = _geocode_comps(processed, mapbox_tok, country_code,
                                   country_name, s_lon, s_lat)
        for i, r in enumerate(processed, 1):
            r["map_marker"] = str(i)

        # ── Location competitiveness (SG, URA proximity vs subject) ───────────
        # Sets the Location column to Superior/Comparable/Inferior for comps
        # geocoded by OneMap; others left blank. SG-only (URA Master Plan).
        try:
            from tools.location_score import apply_location as _apply_loc
            processed = _apply_loc(processed,
                                   subject_cfg.get("property_name", ""),
                                   subject_cfg.get("address", ""),
                                   subject_cfg.get("asset_class", ""),
                                   subj_lonlat=(s_lon, s_lat))
        except Exception as _le:
            print(f"  [location] skipped: {_le}")

        # ── Resolve URA land use zone codes (Singapore only) ─────────────────
        # URA Master Plan zoning is Singapore-specific; skip it for foreign deals.
        if not _is_global:
            print(f"      Resolving URA land use zones …")
            for r in processed:
                raw_zone = str(r.get("land_zoning") or "").strip()
                if raw_zone:
                    code = _resolve_ura_zone(
                        raw_zone,
                        lon=r.get("lon"), lat=r.get("lat"),
                        onemap_token=onemap_token,
                    )
                    if code and code != raw_zone:
                        r["land_zoning"] = code
                else:
                    # No zoning in the source PDF — derive it from the URA Master
                    # Plan land use at the comp's parcel (local, token-free).
                    code = _zone_from_coords(lon=r.get("lon"), lat=r.get("lat"))
                    if code:
                        r["land_zoning"] = code
                        _nm = str(r.get("property_name") or "")[:46]
                        print(f"        {_nm:<46}  zoning ← {code} (from land use)")

        # ── Write lon/lat + map_marker back into _records.json ───────────────
        # lon/lat: re-runs via --from-records skip already-geocoded rows.
        # map_marker: needed so _sync_records_json in the frontend can match
        #             existing rows by marker instead of treating them as new.
        try:
            _rj_path = out_dir / f"{excel_stem}_records.json"
            if _rj_path.exists():
                with open(_rj_path, encoding="utf-8") as _rf:
                    _saved_records = json.load(_rf)
                _meta_by_name = {
                    str(r.get("property_name", "")): {
                        "lon":          r.get("lon"),
                        "lat":          r.get("lat"),
                        "map_marker":   r.get("map_marker"),
                        "_geo_provider": r.get("_geo_provider"),
                        "_geo_note":    r.get("_geo_note"),
                        # location competitiveness label — sync so the preview
                        # shows it, not the raw extracted location text.
                        "location":     r.get("location"),
                        # resolved/derived URA zoning — sync so the preview shows
                        # the code (and any zoning filled in from land use).
                        "land_zoning":  r.get("land_zoning"),
                    }
                    for r in processed
                }
                _any_updated = False
                for sr in _saved_records:
                    _name = str(sr.get("property_name", ""))
                    _meta = _meta_by_name.get(_name)
                    if _meta:
                        for _fld in ("lon", "lat", "map_marker",
                                     "_geo_provider", "_geo_note", "location",
                                     "land_zoning"):
                            if _meta.get(_fld) is not None and sr.get(_fld) != _meta[_fld]:
                                sr[_fld] = _meta[_fld]
                                _any_updated = True
                if _any_updated:
                    with open(_rj_path, "w", encoding="utf-8") as _rf:
                        json.dump(_saved_records, _rf, indent=2, default=str)
                    print(f"  Records (+ coords + markers) → {_rj_path}")
        except Exception:
            pass   # geocoding worked — don't abort over a JSON write failure

    # Console summary
    _is_global = country_name.lower() not in ("", "singapore")
    if not _is_global:
        print(f"\n  {'#':<3} {'Property':<42} {'Yrs':>4} {'B(comp)':>8} "
              f"{'FTM':>7} {'AdjCR':>7} {'Rel':>4}")
        print("  " + "─" * 80)
        for c in processed:
            bc   = bala_factor(int(c.get("remaining_yrs") or 0))
            name = str(c.get("raw_description", "")).split("\n")[0][:40]
            print(f"  {c.get('map_marker',''):<3} {name:<42} "
                  f"{int(c.get('remaining_yrs') or 0):>4} {bc:>8.4f} "
                  f"{float(c.get('ftm_cap_rate', 0))*100:>6.2f}% "
                  f"{float(c.get('adj_cap_rate', 0))*100:>6.2f}% "
                  f"{c.get('relevance_score', ''):>4}")
    else:
        print(f"\n  {'#':<3} {'Property':<52} {'Price':>10}  {'Rel':>4}")
        print("  " + "─" * 74)
        for c in processed:
            name  = str(c.get("property_name") or c.get("raw_description", "")).split("\n")[0][:50]
            price = float(c.get("price_raw") or c.get("price_sgd_m") or 0)
            print(f"  {c.get('map_marker',''):<3} {name:<52} {price:>10.1f}  "
                  f"{c.get('relevance_score', ''):>4}")

    # ── 5. Render Excel + Map ─────────────────────────────────────────────────
    print(f"\n[5/5] Rendering")
    if _is_global:
        subj_row  = _global_sales_tbl.subject_to_row(subject_cfg, subject_cfg)
        comp_rows = [_global_sales_tbl.comp_to_row(c, subject_cfg) for c in processed]
        _global_sales_tbl.build_workbook(subj_row, comp_rows, subject_cfg, output_file)
    else:
        _pu       = subject_cfg.get("price_unit", "M")
        subj_row  = subject_to_row(subject_cfg)
        comp_rows = [comp_to_row(c, _pu) for c in processed]
        build_workbook(subj_row, comp_rows, subject_cfg, output_file)
    print(f"  Excel → {output_file}")
    _write_mapping_tab(output_file, _col_map, _input_headers, input_file)

    if s_lon is not None:
        _geo_comps = [
            {"map_marker": r["map_marker"],
             "property":   str(r.get("property_name") or ""),
             "address":    str(r.get("address") or ""),
             "lon": r.get("lon"), "lat": r.get("lat")}
            for r in processed
        ]
        write_geo_sidecar(out_geo, s_lon, s_lat, _geo_comps, mb_cfg)
        print(f"  Geo   → {out_geo}")

    if generate_map:
        if not mapbox_tok:
            print("  Map skipped  (no Mapbox token in config)")
        elif s_lon is None:
            print("  Map skipped  (subject geocoding failed)")
        else:
            comps_geo = [
                (r["map_marker"], r["lon"], r["lat"])
                for r in processed if r.get("lon") is not None
            ]
            render_map(
                subject_lonlat = (s_lon, s_lat),
                comps          = comps_geo,
                token          = mapbox_tok,
                output_path    = out_map,
                style          = mb_cfg.get("style",    "streets-v12"),
                width          = mb_cfg.get("width",    1200),
                height         = mb_cfg.get("height",   900),
                padding        = mb_cfg.get("padding",  100),
                pin_size       = mb_cfg.get("pin_size", "l"),
            )
            print(f"  Map   → {out_map}")
    else:
        print("  Map skipped  (pass --map to generate)")

    print(f"\n  Done — {len(processed)} comp(s) written to {output_file}\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate asset sales comps table from input Excel (no web search)"
    )
    parser.add_argument("--config", default="configs/deal_config.json",
                        help="Path to deal config JSON")
    parser.add_argument("--map", action="store_true",
                        help="Generate Mapbox map PNG")
    parser.add_argument("--name-only", action="store_true",
                        help="Qualify a comp on its property name alone (skip the price gate)")
    parser.add_argument("--from-records", metavar="JSON", default=None,
                        help=(
                            "Skip extraction; load records from a previously saved "
                            "_records.json file.  Useful for re-running with edits "
                            "or a different model without re-parsing input files.\n"
                            "Example:  --from-records output/88_Cecil/"
                            "Transaction_Comparables_88_Cecil_records.json"
                        ))
    parser.add_argument("--refinement-file", metavar="TXT", default=None,
                        help=(
                            "Path to a text file containing analyst instructions for "
                            "refining the records list.  Applied after --from-records "
                            "loading.  The LLM filters or modifies the records per the "
                            "instructions before classification proceeds."
                        ))
    parser.add_argument("--refine-engine", choices=["tools", "code_interpreter"],
                        default=None,
                        help="Refinement engine: 'tools' (default six-primitive agent) "
                             "or 'code_interpreter' (OpenAI hosted sandbox).")
    args = parser.parse_args()
    run(args.config, generate_map=args.map,
        from_records=args.from_records,
        refinement_file=args.refinement_file,
        name_only=args.name_only,
        refine_engine=args.refine_engine)
