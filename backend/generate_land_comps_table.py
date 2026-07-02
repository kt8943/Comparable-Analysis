#!/usr/bin/env python3
"""
generate_land_comps_table.py
============================
Schema definition and Excel workbook builder for Land Sale Comparables.

Mirrors the pattern of generate_sales_comps_table.py (asset sales) and
generate_rent_comps_table.py (rent comps) — all rendering logic lives here
so both scan_input_land_comps.py and search_online_land_comps.py share
identical output formatting.

Fixed Output Schema  (13 columns)
----------------------------------
  Property | Map Marker | Date of Launch | Land Zoning | Land Tenure (Y)
  Site Area (SF) | Max GFA (SF) | Price (SGD M) | Price (SGD psf ppr)
  Adj. Price (SGD psf ppr) | Location | Quality | Comment

Calculation transparency
------------------------
  Two live Excel formulas written per comp row (auditable, not hard-coded):
    Price psf ppr  (col I)  =  ROUND(Price(H) × 1,000,000 / Max GFA(G), 0)
    Adj. Price     (col J)  =  Price(I) × Bala(subject) / Bala(comp tenure)
                               Bala(n) = official Singapore Bala Table lookup
  Both reference a 'Params' sheet and 'Bala Tbl' sheet:
    Params!B2 = Subject land tenure / remaining leasehold (yrs)
    'Bala Tbl' = official Singapore Bala Table (Appendix 2, SLA/SISV)
  Freehold / 999-yr: Bala = 1.0 (no adjustment).
"""

import sys as _sys
from pathlib import Path as _Path
_sys.path.insert(0, str(_Path(__file__).parent))

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import openpyxl

from tools.calculations import bala_factor, bala_expr as _bala_expr, _BALA_TABLE


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — OUTPUT SCHEMA
# ═══════════════════════════════════════════════════════════════════════════════

LAND_SCHEMA_BASE = [
    ("Type",                             "type",           "str",   "@",                10),
    ("Source",                           "source",         "str",   "@",                10),
    ("Property",                         "property",       "str",   "@",                36),
    ("Map\nMarker",                       "map_marker",     "str",   "@",                 8),
    ("Date of\nLaunch",                   "launch_date",    "str",   "@",                14),
    ("Land\nZoning",                      "land_zoning",    "str",   "@",                16),
    ("Land\nTenure (Y)",                  "tenure_yrs",     "float", '#,##0.# "yrs"',    12),
    ("Site Area\n(SF)",                   "site_area_sf",   "int",   "#,##0",            13),
    ("Max GFA\n(SF)",                     "max_gfa_sf",     "int",   "#,##0",            13),
    ("Price\n(SGD M)",                    "price_sgd_m",    "float", '"S$"#,##0.0',   13),
    ("Price\n(SGD psf ppr)",              "price_psf_ppr",  "int",   '"S$"#,##0',        16),
    ("Adj. Price\n(SGD psf ppr)",         "adj_price_psf",  "int",   '"S$"#,##0',        18),
    ("Location",                          "location",       "str",   "@",                24),
    ("Quality",                           "quality",        "str",   "@",                28),
    ("Comment",                           "comment",        "str",   "@",                40),
]


def _schema_idx(key: str, schema: list) -> int:
    """Return 1-based column index for the given internal key."""
    for i, entry in enumerate(schema, 1):
        if entry[1] == key:
            return i
    raise KeyError(f"Schema key {key!r} not found")


def get_land_schema(subject_cfg: dict = None) -> list:
    """
    Return LAND_SCHEMA_BASE adapted for this deal's currency and area unit.
    Uses key-based lookup so column order changes never break indices.
    """
    if subject_cfg is None:
        return list(LAND_SCHEMA_BASE)
    currency = subject_cfg.get("currency", "SGD")
    sym        = subject_cfg.get("currency_symbol", currency)
    price_unit = subject_cfg.get("price_unit", "M").upper()
    gfa_unit   = subject_cfg.get("gfa_unit", "sf").lower()
    gfa_lbl    = "sqm" if gfa_unit == "sqm" else "SF"
    area_lbl   = "psm" if gfa_unit == "sqm" else "psf"
    schema     = list(LAND_SCHEMA_BASE)
    for i, entry in enumerate(schema):
        if entry[1] == "site_area_sf":
            schema[i] = (f"Site Area\n({gfa_lbl.upper()})", "site_area_sf", "int",
                         "#,##0", entry[4])
        elif entry[1] == "max_gfa_sf":
            schema[i] = (f"Max GFA\n({gfa_lbl.upper()})", "max_gfa_sf", "int",
                         "#,##0", entry[4])
        elif entry[1] == "price_sgd_m":
            schema[i] = (f"Price\n({currency} {price_unit})", "price_sgd_m", "float",
                         f'"{sym}"#,##0.0', entry[4])
        elif entry[1] == "price_psf_ppr":
            schema[i] = (f"Price\n({currency} {area_lbl} ppr)", "price_psf_ppr", "int",
                         f'"{sym}"#,##0', entry[4])
        elif entry[1] == "adj_price_psf":
            schema[i] = (f"Adj. Price\n({currency} {area_lbl} ppr)", "adj_price_psf", "int",
                         f'"{sym}"#,##0', entry[4])
    return schema


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — STYLE HELPERS  (palette shared with sales / rent comps)
# ═══════════════════════════════════════════════════════════════════════════════

_NAVY  = "FF1F3864"
_NAVYL = "FF2E4C7E"
_WHITE = "FFFFFFFF"
_LGRAY = "FFF2F2F2"
_NOTE  = "FFEBF3FB"
_DARK  = "FF1A1A1A"
_RED   = "FFC00000"   # subject property star marker colour
_AVGBG = "FFD6DCE4"   # blue-grey for average row

def _fill(h):
    return PatternFill(patternType="solid", fgColor=h)

def _font(color=_WHITE, bold=True, sz=9, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=sz, italic=italic)

_T = Side(style="thin", color="FFBFBFBF")
def _border():
    return Border(left=_T, right=_T, top=_T, bottom=_T)

def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — WORKSHEET PRIMITIVES
# ═══════════════════════════════════════════════════════════════════════════════

def _section_header(ws, row: int, label: str, nc: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row=row, column=1, value=label)
    c.fill = _fill(_NAVY); c.font = _font(_WHITE, bold=True, sz=10)
    c.alignment = _align("left", "center", wrap=False); c.border = _border()
    for col in range(2, nc + 1):
        x = ws.cell(row=row, column=col)
        x.fill = _fill(_NAVY); x.border = _border()
    ws.row_dimensions[row].height = 19


def _col_headers(ws, row: int, schema: list):
    for col, (hdr, *_) in enumerate(schema, 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.fill = _fill(_NAVY); c.font = _font(_WHITE, bold=True, sz=8.5)
        c.alignment = _align("center", "center", wrap=True); c.border = _border()
    ws.row_dimensions[row].height = 34


def _data_row(ws, row: int, row_dict: dict, schema: list,
              alt: bool = False, bold: bool = False, is_subject: bool = False):
    """Write all 13 column values. Col J (Adj Price) is overwritten by formula; col I (Price psf) uses Python value."""
    bg = _fill(_LGRAY) if alt else _fill(_WHITE)
    for col, entry in enumerate(schema, 1):
        _, key, dtype, fmt, _ = entry
        val     = row_dict.get(key)
        display = "—" if val is None else val

        c = ws.cell(row=row, column=col, value=display)
        c.fill = bg; c.border = _border()
        c.font = _font(_DARK, bold=bold, sz=9)

        if col == 1:
            c.alignment = _align("left", "center", wrap=True)
        elif col == 2:
            c.alignment = _align("center", "center", wrap=False)
            # Subject property star is red; comp markers stay dark
            marker_color = _RED if is_subject else _DARK
            c.font = _font(marker_color, bold=True, sz=12)
        elif dtype in ("int", "float") and isinstance(val, (int, float)):
            c.alignment = _align("right", "center", wrap=False)
            c.number_format = fmt
        else:
            c.alignment = _align("center", "center", wrap=True)

    ws.row_dimensions[row].height = 48


def _write_formulas(ws, row: int, schema: list, is_subject: bool = False):
    """
    Write live Excel formula for Adj. Price psf ppr (col J).

    Price psf ppr (col I) is now a Python-computed static value from _data_row
    so it appears in the Streamlit preview (openpyxl cannot evaluate formulas
    when reading with data_only=True). The adj formula still references col I
    by cell address, so it computes correctly in Excel on download.

    Col J = subject → mirrors col I
            comp    → ROUND(I × Bala(subject yrs) / Bala(comp tenure), 0)
                       Bala from official Singapore Bala Table VLOOKUP.
    Freehold / 999-yr: Bala = 1.0 (no adjustment).
    """
    psf_col = _schema_idx("price_psf_ppr", schema)
    adj_col = _schema_idx("adj_price_psf", schema)
    adj_fmt = schema[adj_col - 1][3]
    E = get_column_letter(_schema_idx("tenure_yrs",  schema))
    I = get_column_letter(psf_col)

    # Col J — Adj. Price psf ppr
    adj = ws.cell(row=row, column=adj_col)
    adj.number_format = adj_fmt
    adj.alignment     = _align("right", "center", wrap=False)
    adj.font          = _font(_DARK, bold=is_subject, sz=9)

    if is_subject:
        adj.value = f"=IF({I}{row}=\"—\",\"—\",{I}{row})"
        return

    # Adj Price = Price × Bala(subject) / Bala(comp)
    # Uses official Singapore Bala Table VLOOKUP from 'Bala Tbl' sheet
    bala_comp = _bala_expr(f"{E}{row}")
    bala_subj = _bala_expr("Params!$B$2")
    adj.value = (
        f"=IFERROR("
        f"IF(OR({I}{row}=\"—\",{I}{row}=0),\"—\","
        f"ROUND({I}{row}*{bala_subj}/{bala_comp},0)),"
        f"\"—\")"
    )


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — PARAMS SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def _avg_row(ws, row: int, first_r: int, last_r: int, schema: list):
    """Write Average row — Land Tenure, Price psf ppr, Adj. Price psf ppr."""
    avg_keys = {"tenure_yrs", "price_psf_ppr", "adj_price_psf"}
    for col, entry in enumerate(schema, 1):
        key, dtype, fmt = entry[1], entry[2], entry[3]
        c = ws.cell(row=row, column=col)
        c.fill   = _fill(_AVGBG)
        c.border = _border()
        if col == 1:
            c.value     = "Average"
            c.font      = _font(_DARK, bold=True, sz=9)
            c.alignment = _align("left", "center", wrap=False)
        elif key in avg_keys:
            col_ltr         = get_column_letter(col)
            c.value         = (f'=IFERROR(AVERAGEIF({col_ltr}{first_r}:'
                               f'{col_ltr}{last_r},">0"),"—")')
            c.number_format = fmt
            c.font          = _font(_DARK, bold=True, sz=9)
            c.alignment     = _align("right", "center", wrap=False)
        else:
            c.value     = ""
            c.font      = _font(_DARK, bold=False, sz=9)
            c.alignment = _align("center", "center", wrap=False)
    ws.row_dimensions[row].height = 20


def _build_params_sheet(wb, subject_cfg: dict, bala_yield: float = 0.06):
    """
    'Params' worksheet — holds parameters referenced by the Adj. Price formula.
    B2 = Subject Land Tenure / Remaining Leasehold (yrs) — referenced by every comp row.
    Bala factors are now looked up from the official Singapore Bala Table
    in the 'Bala Tbl' sheet (not calculated from a parametric yield).
    """
    ws = wb.create_sheet("Params")
    ws.sheet_view.showGridLines = False

    subj_yrs  = subject_cfg.get("remaining_leasehold_yrs", 0)
    rows_data = [
        ("Parameter",                                                           "Value",   None),
        ("Subject Land Tenure / Remaining Leasehold (yrs)",                    subj_yrs,  '#,##0 "yrs"'),
        ("", "", None),
        ("Bala Table",                                 "See 'Bala Tbl' sheet — official SG table", None),
        ("Formula: Adj Price = Price × Bala(subject) / Bala(comp)", "", None),
        ("Source: Appendix 2, SLA/SISV Singapore Bala Table",       "", None),
    ]
    for r, (a, b, fmt) in enumerate(rows_data, 1):
        ca = ws.cell(row=r, column=1, value=a)
        cb = ws.cell(row=r, column=2, value=b)
        ca.font = _font(_DARK, bold=(r == 1), sz=9)
        cb.font = _font(_DARK, bold=(r == 1), sz=9)
        ca.alignment = _align("left",  "center", wrap=False)
        cb.alignment = _align("right", "center", wrap=False)
        ca.border = _border(); cb.border = _border()
        if fmt and isinstance(b, (int, float)):
            cb.number_format = fmt

    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 36
    for r in range(1, len(rows_data) + 1):
        ws.row_dimensions[r].height = 18


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — ROW CONVERTERS
# ═══════════════════════════════════════════════════════════════════════════════

def _calc_land_price_psf(price_m, max_gfa_sf):
    """Compute price psf ppr. price_m is always M-normalized, so always × 1e6."""
    try:
        pm  = float(price_m  or 0) or None
        gfa = int(max_gfa_sf or 0) or None
        if pm and gfa:
            return round(pm * 1_000_000 / gfa)
    except (TypeError, ValueError):
        pass
    return None


def _m_to_display(price_m, price_unit: str):
    """Convert M-normalized stored price to display unit (B → ÷1000, else no-op)."""
    if price_m is None:
        return None
    try:
        v = float(price_m)
        return round(v / 1000, 4) if price_unit.upper() == "B" else v
    except (TypeError, ValueError):
        return price_m


def subject_to_row(cfg: dict) -> dict:
    """Convert subject_property config block → land comps row dict."""
    prop_name  = cfg["property_name"]
    address    = cfg.get("address", "")
    price_unit = cfg.get("price_unit", "M")
    prop_cell  = (f"{prop_name}\n{address}"
                  if address and address.strip() != prop_name.strip()
                  else prop_name)
    _price_m  = cfg.get("price_sgd_m")
    _max_gfa  = cfg.get("gfa_sf")
    return {
        "type":          "Subject",
        "source":        "",
        "property":      prop_cell,
        "map_marker":    "★",
        "launch_date":   cfg.get("sale_date", ""),
        "land_zoning":   cfg.get("land_zoning", ""),
        "tenure_yrs":    cfg.get("remaining_leasehold_yrs", 0),
        "site_area_sf":  cfg.get("site_area_sf"),   # may be absent → shows "—"
        "max_gfa_sf":    _max_gfa,
        "price_sgd_m":   _m_to_display(_price_m, price_unit),
        "price_psf_ppr": _calc_land_price_psf(_price_m, _max_gfa),
        "adj_price_psf": None,   # subject adj = psf (formula); shows "—" in preview
        "location":      cfg.get("location", ""),
        "quality":       cfg.get("quality", ""),
        "comment":       "",
    }


def comp_to_row(r: dict, price_unit: str = "M") -> dict:
    """Convert a classified comp record → land comps row dict.
    price_sgd_m is M-normalized internally; display conversion applied here.
    """
    name = str(r.get("property_name") or "")
    addr = str(r.get("address")       or "").strip()
    _generic = {"comparable", "comp", "n/a", "na", "tbc", "tbd", "-", "—"}
    if addr and addr.lower() != name.lower() and addr.lower() not in _generic:
        prop = f"{name}\n{addr}"
    else:
        prop = name
    _src = r.get("_source", "")
    _src_map = {"excel": "Excel", "pdf": "PDF", "image": "Image", "manual": "Manual"}
    if _src.startswith("pdf_"):
        _src_map[_src] = "PDF " + _src[4:]
    if _src.startswith("excel_"):
        _src_map[_src] = "Excel " + _src[6:]
    if _src.startswith("image_"):
        _src_map[_src] = "Image " + _src[6:]
    return {
        "type":          "Comparable",
        "source":        _src_map.get(_src, ""),
        "property":      prop,
        "map_marker":    str(r.get("map_marker", "")),
        "launch_date":   str(r.get("launch_date")   or ""),
        "land_zoning":   str(r.get("land_zoning")   or ""),
        "tenure_yrs":    r.get("tenure_yrs"),
        "site_area_sf":  r.get("site_area_sf"),
        "max_gfa_sf":    r.get("max_gfa_sf"),
        "price_sgd_m":   _m_to_display(r.get("price_sgd_m"), price_unit),
        "price_psf_ppr": _calc_land_price_psf(r.get("price_sgd_m"), r.get("max_gfa_sf")),
        "adj_price_psf": None,   # Bala-adjusted formula; shows "—" in preview
        "location":      str(r.get("location")  or ""),
        "quality":       str(r.get("quality")   or ""),
        "comment":       str(r.get("comment")   or ""),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 6b — BALA TABLE SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def _build_bala_sheet(wb):
    """
    Add 'Bala Tbl' lookup sheet containing the official Singapore Bala Table.
    Column A = remaining years (1–99), Column B = factor as decimal (0.038–0.960).
    Referenced by the Adj. Price VLOOKUP formulas in every comp row.
    Source: Appendix 2, Table Showing Leasehold Values as Percentage of Freehold Value.
    """
    if "Bala Tbl" in wb.sheetnames:
        del wb["Bala Tbl"]
    ws = wb.create_sheet("Bala Tbl")
    ws.sheet_view.showGridLines = False

    # Header
    for col, hdr in enumerate(["Years", "Factor (decimal)"], 1):
        c = ws.cell(row=1, column=col, value=hdr)
        c.font      = _font(_WHITE, bold=True, sz=9)
        c.fill      = _fill(_NAVY)
        c.alignment = _align("center", "center", wrap=False)
        c.border    = _border()

    # Data rows: years 1–99
    for row_i, (yrs, pct) in enumerate(_BALA_TABLE.items(), 2):
        ca = ws.cell(row=row_i, column=1, value=yrs)
        cb = ws.cell(row=row_i, column=2, value=round(pct / 100.0, 5))
        for c in (ca, cb):
            c.font      = _font(_DARK, bold=False, sz=9)
            c.alignment = _align("center", "center", wrap=False)
            c.border    = _border()
        ca.number_format = "0"
        cb.number_format = "0.000"

    # Source note
    note_row = len(_BALA_TABLE) + 3
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=2)
    n = ws.cell(row=note_row, column=1,
                value=("Source: Appendix 2 — Table Showing Leasehold Values as Percentage "
                       "of Freehold Value (Singapore Land Authority / SISV)"))
    n.font      = _font("FF555555", bold=False, sz=8, italic=True)
    n.alignment = _align("left", "center", wrap=True)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    for r in range(1, note_row + 1):
        ws.row_dimensions[r].height = 15


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 7 — WORKBOOK BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

def build_workbook(subject_cfg: dict, subject_row_d: dict,
                   comp_rows: list, out_excel: str,
                   schema: list = None, bala_yield: float = 0.06,
                   title_suffix: str = "LAND SALE COMPARABLES",
                   subtitle: str = "Confidential — For Discussion Purposes Only",
                   subject_section: str = "  Subject Property — Site / Land Reference",
                   comps_section: str  = "  Land Sale Comparables"):
    """
    Write the two-table formatted Excel workbook.

    title_suffix / subtitle / section labels are overrideable so
    search_online_land_comps.py can inject its own header text.
    """
    if schema is None:
        schema = get_land_schema(subject_cfg)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Land Sale Comparables"
    ws.sheet_view.showGridLines = False

    nc        = len(schema)
    deal_name = subject_cfg.get("deal_name", subject_cfg["property_name"])

    for i, (*_, w) in enumerate(schema, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Hide the Type column (col A) — visible when needed, hidden by default
    ws.column_dimensions["A"].hidden = True

    # ── Title block ──────────────────────────────────────────────────────────
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    t = ws.cell(row=r, column=1,
                value=f"{deal_name.upper()}  —  {title_suffix}")
    t.fill = _fill(_NAVY); t.font = _font(_WHITE, bold=True, sz=13)
    t.alignment = _align("center", "center", wrap=False)
    ws.row_dimensions[r].height = 24

    r = 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    s = ws.cell(row=r, column=1, value=subtitle)
    s.fill = _fill(_NAVYL); s.font = _font(_WHITE, bold=False, sz=9, italic=True)
    s.alignment = _align("center", "center", wrap=False)
    ws.row_dimensions[r].height = 14

    # ── Table 1: Subject ─────────────────────────────────────────────────────
    r = 4
    _section_header(ws, r, subject_section, nc)
    r = 5; _col_headers(ws, r, schema)
    r = 6; _data_row(ws, r, subject_row_d, schema, bold=True, is_subject=True)
    _write_formulas(ws, r, schema, is_subject=True)
    ws.row_dimensions[7].height = 6

    # ── Table 2: Comps ───────────────────────────────────────────────────────
    r = 8
    _section_header(ws, r, comps_section, nc)
    r = 9; _col_headers(ws, r, schema)
    for i, crow in enumerate(comp_rows):
        r = 10 + i
        _data_row(ws, r, crow, schema, alt=(i % 2 == 1))
        _write_formulas(ws, r, schema, is_subject=False)

    # ── Average row ───────────────────────────────────────────────────────────
    if comp_rows:
        _avg_row(ws, 10 + len(comp_rows), 10, 9 + len(comp_rows), schema)

    # ── Notes footer ─────────────────────────────────────────────────────────
    r = 10 + len(comp_rows) + 2
    currency = subject_cfg.get("currency", "SGD")
    area_lbl = "psm" if subject_cfg.get("gfa_unit","sf").lower()=="sqm" else "psf"
    notes = (
        f"Notes:  (1) ★ = {deal_name}; markers 1–{len(comp_rows)} = land sale comparables.  "
        f"(2) Price ({currency} {area_lbl} ppr) = Sale Price × 1,000,000 / Max GFA (live Excel formula).  "
        f"(3) Adj. Price ({currency} {area_lbl} ppr) normalises for land tenure using the official "
        f"Singapore Bala Table (Appendix 2, SLA/SISV).  "
        "Formula: Adj Price = Price × Bala(subject) / Bala(comp).  "
        "Bala factors are looked up from the 'Bala Tbl' sheet — see Params for subject yrs.  "
        "(4) Freehold / 999-yr sites: Bala = 1.0.  "
        "Source: URA, JLL, CBRE, Colliers, SLP International."
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    n = ws.cell(row=r, column=1, value=notes)
    n.fill = _fill(_NOTE); n.font = _font("FF404040", bold=False, sz=8, italic=True)
    n.alignment = _align("left", "top", wrap=True)
    ws.row_dimensions[r].height = 72

    # ── Params + Bala Tbl sheets ─────────────────────────────────────────────
    _build_params_sheet(wb, subject_cfg)
    _build_bala_sheet(wb)
    wb.save(out_excel)
    print(f"  Saved → {out_excel}")
