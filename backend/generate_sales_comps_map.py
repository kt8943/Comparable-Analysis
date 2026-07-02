#!/usr/bin/env python3
"""
generate_sales_comps_map.py
===========================
Asset Sales Comparables map module.

Geocoding and rendering engine lives in generate_comps_map_base.py —
shared with rent and land map modules so all three are independent.

Usage
-----
    python3 generate_sales_comps_map.py --config configs/deal_config_88_Cecil.json

Output
------
    output/<DealName>/Transaction_Comparables_<DealName>_map.png
"""

import sys as _sys
from pathlib import Path as _Path
_sys.path.insert(0, str(_Path(__file__).parent))

import argparse
import json
from pathlib import Path

import openpyxl

# All geocoding + rendering comes from the shared base — no dependency on
# rent or land map modules.
from generate_comps_map_base import (
    geocode_with_fallbacks,
    render_map,
    _parse_property_text,
)


# ─────────────────────────────────────────────────────────────────────────────
# READ COMPS FROM SALES EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def read_comps_from_excel(xlsx_path: str) -> list:
    """
    Return [(marker_label, property_name_raw), ...] from the
    'Transaction Comparables' sheet — skips subject (★) and blank rows.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Transaction Comparables"]

    comps = []
    for row in ws.iter_rows(min_row=10, max_row=ws.max_row, values_only=True):
        prop   = str(row[0] or "").strip()
        marker = str(row[1] or "").strip()
        if prop and marker and marker not in ("", "★", "—"):
            comps.append((marker, prop))
    return comps


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def run(config_path: str = "configs/deal_config.json"):
    with open(config_path, encoding="utf-8") as f:
        cfg = json.load(f)

    subject_cfg  = cfg["subject_property"]
    output_excel = cfg["output_file"]
    Path(output_excel).parent.mkdir(parents=True, exist_ok=True)
    mb_cfg       = cfg.get("mapbox", {})
    token        = mb_cfg.get("token")
    style        = mb_cfg.get("style",    "streets-v12")
    width        = mb_cfg.get("width",    1200)
    height       = mb_cfg.get("height",   900)
    padding      = mb_cfg.get("padding",  100)
    pin_size     = mb_cfg.get("pin_size", "l")
    bounds_tuple = tuple(mb_cfg["geocode_bounds"]) if mb_cfg.get("geocode_bounds") else None

    if not token:
        raise ValueError("mapbox.token missing in deal config.")

    deal_name    = subject_cfg.get("deal_name", subject_cfg["property_name"])
    map_output   = str(Path(output_excel).with_suffix("")) + "_map.png"
    address      = subject_cfg.get("address", "")
    country_code = cfg.get("country_code", "")
    country_name = subject_cfg.get("country_name", "")
    suffix       = f", {country_name}" if country_name else ""

    print(f"\n{'='*60}\n  Sales Comps Map : {deal_name}\n{'='*60}")

    print(f"\n[1/4] Geocoding subject")
    s_lon, s_lat, _ = geocode_with_fallbacks(
        [f"{subject_cfg['property_name']}, {address}", address, subject_cfg['property_name']],
        token, country_code, bounds=bounds_tuple,
    )
    print(f"      {subject_cfg['property_name']:<45} ({s_lon}, {s_lat})")

    print(f"\n[2/4] Reading comps from {output_excel}")
    comps_raw = read_comps_from_excel(output_excel)
    print(f"      → {len(comps_raw)} comps found")

    print(f"\n[3/4] Geocoding comparables")
    comps_geo = []
    for marker, prop in comps_raw:
        name, addr_line = _parse_property_text(prop)
        queries = []
        if addr_line:
            queries.append(f"{addr_line}{suffix}" if suffix not in addr_line else addr_line)
        queries.append(f"{name}{suffix}")
        try:
            lon, lat, _ = geocode_with_fallbacks(queries, token, country_code, bounds=bounds_tuple)
            print(f"      {marker:>2}. {(addr_line or name)[:45]:<45} ({lon}, {lat})")
            comps_geo.append((marker, lon, lat))
        except Exception as exc:
            print(f"      {marker:>2}. FAILED — {exc}")

    print(f"\n[4/4] Rendering map → {map_output}")
    render_map(
        subject_lonlat = (s_lon, s_lat),
        comps          = comps_geo,
        token          = token,
        output_path    = map_output,
        style          = style,
        width          = width,
        height         = height,
        padding        = padding,
        pin_size       = pin_size,
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate asset sales comps map")
    parser.add_argument("--config", default="configs/deal_config.json")
    args = parser.parse_args()
    run(args.config)
