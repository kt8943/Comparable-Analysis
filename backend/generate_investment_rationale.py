"""
generate_investment_rationale.py
=================================
Two-stage LLM pipeline that reads market report PDFs and writes an
institutional-quality investment rationale for a subject property.

─── HOW IT WORKS ────────────────────────────────────────────────────────────

Stage 1 — extract_report_insights()
    Each market report PDF is read and passed to the LLM, which extracts
    structured facts (vacancy, rents, yields, outlook, key statistics) into
    a JSON object.  The result is saved to a cache file keyed by the PDF's
    filename + size + modified-time, so the same unchanged file is never
    re-extracted.  Only new or force-refreshed PDFs trigger an LLM call.

    PDF / txt report  →  LLM (temp=0.1)  →  structured JSON
    Cached at:  Input_files/market_reports/cache/<stem>_<hash>.json

Stage 2 — generate_rationale()
    Uses the cached insights from Stage 1 plus the deal config (subject
    property details) to write a 3–5 section investment committee memo.

    This stage runs TWO separate LLM calls:

    Call 1 — Prose writing (sources anonymised)
        The research data is passed with source labels replaced by generic
        names ("Research Report 1", "Research Report 2"…) so the LLM cannot
        echo real PDF filenames into the client-facing memo.  The LLM follows
        a STEP 1→2→3→4 chain-of-thought: first retrieving all evidence,
        then planning sections, then writing prose, then self-verifying.

    Call 2 — Source audit JSON (real filenames visible)
        The finished prose + research data with real filenames are sent to a
        second LLM call whose only job is to match every specific claim back
        to its source PDF and page number.  Output populates Source_Audit.xlsx.

    Cached insights + deal config  →  LLM (temp=0.3)  →  markdown rationale
    Output saved to: output/<deal_dir>/Investment_Rationale.md
                     output/<deal_dir>/Source_Audit.xlsx

─── CLI USAGE ───────────────────────────────────────────────────────────────
python generate_investment_rationale.py \\
    --config  configs/deal_config_xxx.json \\
   [--reports Input_files/market_reports/report1.pdf report2.pdf ...] \\
   [--refresh]          # ignore cache and re-extract every report \\
   [--notes-file /tmp/analyst_notes.txt]

─── KEY DESIGN DECISIONS ────────────────────────────────────────────────────
• File-hash caching      — same PDF is never re-extracted unless it changes
• Two separate LLM calls — keeps filenames out of prose; tracks sources cleanly
• Anonymised sources     — prevents model echoing internal PDF names into output
• STEP 1 data retrieval  — forces model to list evidence before writing prose,
                           reducing hallucination
• CJK stripping          — removes Korean/Japanese/Chinese text before Stage 2
                           so non-English characters never appear in the output
• Smart truncation       — keeps front 75% + tail 25% of PDF text to capture
                           both executive summary and conclusions/outlook
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path

# ── Windows UTF-8 fix ────────────────────────────────────────────────────────
# Windows consoles default to cp1252 which cannot encode many Unicode characters
# produced by LLMs (figure dashes, box-drawing chars, CJK text, etc.).
# Reconfigure stdout/stderr to UTF-8 so print() never raises UnicodeEncodeError.
for _stream in (sys.stdout, sys.stderr):
    try:
        if getattr(_stream, "encoding", "utf-8").lower().replace("-", "") != "utf8":
            _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, OSError):
        pass  # reconfigure not available (Python < 3.7) or stream is not a TTY

# Project root (two levels up from this file: backend/ → PGIM/)
ROOT      = Path(__file__).parent.parent
# Where Stage 1 cached JSON files are stored — one file per PDF
CACHE_DIR = ROOT / "Input_files" / "market_reports" / "cache"
# Where market report PDFs are uploaded by the user
REPORTS_DIR = ROOT / "Input_files" / "market_reports"


# ═══════════════════════════════════════════════════════════════════════════════
# PDF / FILE READING
# ═══════════════════════════════════════════════════════════════════════════════

def _read_pdf(pdf_path: Path, max_pages: int = 50) -> str:
    """
    Extract plain text from a PDF using pypdf (up to max_pages pages).

    Each page is prefixed with a [PAGE N] tag so the LLM knows which page
    a statistic came from — this feeds the page references in the source audit.

    Limitation: pypdf only reads text layer; charts, images, and scanned PDFs
    produce empty or garbled output.  If a report is image-only, Stage 1
    extraction will return little or no data.
    """
    try:
        import pypdf
    except ImportError:
        raise ImportError("pypdf is required for PDF reading.  pip install pypdf")

    reader = pypdf.PdfReader(str(pdf_path))
    pages  = reader.pages[:max_pages]
    parts  = []
    for i, pg in enumerate(pages, 1):
        t = pg.extract_text() or ""
        if t.strip():
            parts.append(f"[PAGE {i}]\n{t.strip()}")
    return "\n\n".join(parts)


def _read_report(path: Path) -> str:
    """Read a report file — PDF or plain text."""
    if path.suffix.lower() == ".pdf":
        return _read_pdf(path)
    return path.read_text(encoding="utf-8", errors="ignore")


def _smart_truncate(text: str, max_chars: int = 14000) -> str:
    """
    Trim text to fit within the LLM's context limit (14,000 chars by default).

    Strategy: keep the first 75% + the last 25% of the text.
    - First 75% captures the executive summary and main findings (front of report)
    - Last 25% captures the outlook, forecasts, and conclusions (tail of report)
    - Middle sections (detailed tables, appendices) are dropped with a marker

    This is better than a simple head-truncation which would cut off the outlook.
    """
    if len(text) <= max_chars:
        return text
    front = int(max_chars * 0.75)
    back  = max_chars - front
    return text[:front] + "\n\n[…middle omitted…]\n\n" + text[-back:]


# ═══════════════════════════════════════════════════════════════════════════════
# LLM HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _strip_thinking_tags(text: str) -> str:
    """
    Remove <think>...</think> blocks that reasoning models emit before their answer.

    Models like deepseek-r1 and qwen3 output their internal chain-of-thought
    wrapped in <think> tags before giving the final answer.  These tags must be
    stripped so the pipeline only processes the actual output, not the reasoning.
    Safe to call on any text — has no effect if no <think> tags are present.
    """
    return re.sub(r'<think>[\s\S]*?</think>', '', text, flags=re.IGNORECASE).strip()


# CJK Unicode ranges: Hangul, CJK Unified Ideographs, Hiragana, Katakana,
# CJK Compatibility Ideographs, CJK Extension A/B, Hangul Jamo, etc.
_CJK_RE = re.compile(
    r'[ᄀ-ᇿ'   # Hangul Jamo
    r'぀-ヿ'    # Hiragana + Katakana
    r'㄰-㆏'    # Hangul Compatibility Jamo
    r'가-힣'    # Hangul Syllables
    r'㐀-䶿'    # CJK Extension A
    r'一-鿿'    # CJK Unified Ideographs
    r'豈-﫿'    # CJK Compatibility Ideographs
    r'\U00020000-\U0002A6DF]+'  # CJK Extension B
)


def _strip_cjk(text: str) -> str:
    """
    Remove Korean, Japanese, and Chinese characters from a string.

    Why this is needed:
    - Market reports for Asian deals (Korea, Japan, China) often contain
      text in the local language mixed with English.
    - pypdf extracts this text as-is, so Stage 1 cache may contain CJK characters.
    - If Stage 2 sees CJK text in the research summary, small LLMs sometimes
      output non-English characters in the final rationale.
    - This function strips all CJK before the text reaches Stage 2, ensuring
      the rationale is always written in English only.

    Consecutive CJK sequences are replaced with a single space so surrounding
    Latin text (numbers, punctuation) stays readable.
    """
    return _CJK_RE.sub(" ", text).strip()


def _ollama_chat(messages: list[dict], base_url: str, model: str,
                 temperature: float = 0.2) -> str:
    """
    Send a chat request to a locally running Ollama instance.

    Uses urllib (no extra dependencies) to POST to the /api/chat endpoint.
    stream=False means Ollama waits for the full response before returning
    (simpler than streaming but holds the connection open for longer runs).
    timeout=600 allows up to 10 minutes for large models on slow hardware.
    """
    import urllib.request
    payload = json.dumps({
        "model":   model,
        "messages": messages,
        "stream":  False,
        "options": {"temperature": temperature},
    }).encode()
    req = urllib.request.Request(
        f"{base_url.rstrip('/')}/api/chat",
        data=payload,
        headers={"Content-Type": "application/json"},
    )
    with urllib.request.urlopen(req, timeout=600) as resp:
        data = json.loads(resp.read())
    return _strip_thinking_tags(data["message"]["content"].strip())


def _openai_chat(messages: list[dict], model: str, api_key: str,
                 temperature: float = 0.3, json_mode: bool = False) -> str:
    """
    Send a chat request to OpenAI's API (gpt-4o, gpt-4o-mini, etc.).

    Requires the `openai` Python package (pip install openai).
    Used when the user selects a cloud model from the sidebar instead of
    a local Ollama model.  Much faster and higher quality than local 3b/7b models
    for investment rationale writing, but costs money per call.
    """
    try:
        import openai as _openai
    except ImportError:
        raise ImportError("openai package required for GPT models.  pip install openai")
    client = _openai.OpenAI(api_key=api_key)
    kwargs: dict = dict(model=model, messages=messages, temperature=temperature)
    if json_mode:
        kwargs["response_format"] = {"type": "json_object"}
    resp = client.chat.completions.create(**kwargs)
    return _strip_thinking_tags(resp.choices[0].message.content.strip())


def _llm_chat(messages: list[dict], llm_cfg: dict, openai_key: str = "",
              temperature: float = 0.3, json_mode: bool = False) -> str:
    """
    Central router — sends a chat request to either Ollama or OpenAI.

    The frontend sidebar injects the user's model choice into llm_cfg before
    calling the pipeline, so this function always receives the correct provider.

    llm_cfg expected shape (from deal config + frontend overrides):
      Ollama:  {"provider": "ollama",  "ollama": {"base_url": ..., "model": ...}}
      OpenAI:  {"provider": "openai",  "openai_model": "gpt-4o"}

    Falls back to Ollama qwen2.5:3b if llm_cfg is empty or missing.
    Prints the provider and model name to the console so the user can see
    which model is running during each pipeline step.
    """
    provider = llm_cfg.get("provider", "ollama")
    if provider == "openai":
        key   = openai_key or llm_cfg.get("openai_api_key", "")
        model = llm_cfg.get("openai_model", "gpt-4o")
        if not key:
            raise ValueError(
                "OpenAI model selected but no API key found. "
                "Set it in the deal config under openai.api_key or via OPENAI_API_KEY env var."
            )
        print(f"  [LLM] OpenAI / {model}")
        return _openai_chat(messages, model, key, temperature, json_mode=json_mode)
    else:
        ollama = llm_cfg.get("ollama", {})
        model  = ollama.get("model", "qwen2.5:3b")
        url    = ollama.get("base_url", "http://localhost:11434")
        print(f"  [LLM] Ollama / {model}")
        return _ollama_chat(messages, url, model, temperature)


def _parse_json_from_llm(text: str) -> dict:
    """
    Extract and parse a JSON object from an LLM response.

    LLMs often wrap their JSON in markdown code fences (```json ... ```)
    or add preamble text before the actual JSON.  This function handles both:
    1. Strips ```json ... ``` fences if present
    2. Falls back to finding the first { ... } block if no fences found
    3. Raises json.JSONDecodeError if no valid JSON can be found
    """
    # Remove ```json ... ``` fences
    m = re.search(r"```(?:json)?\s*([\s\S]*?)```", text)
    if m:
        text = m.group(1)
    # Fallback: find first { ... }
    m2 = re.search(r"\{[\s\S]*\}", text)
    if m2:
        text = m2.group(0)
    return json.loads(text.strip())


def _fill(template: str, **kwargs) -> str:
    """
    Safe template fill.  Replaces {key} placeholders without choking on
    curly braces that may appear in the substituted values.
    """
    for key, val in kwargs.items():
        template = template.replace("{" + key + "}", str(val))
    return template


# ═══════════════════════════════════════════════════════════════════════════════
# RAG AUDIT UTILITIES
# ═══════════════════════════════════════════════════════════════════════════════

# Finite verbs that only appear in complete clauses (prose), not in chart titles
# or axis labels.  A line containing any of these is almost certainly a sentence.
_FINITE_VERBS = frozenset({
    # To be / to have
    "is", "are", "was", "were", "been", "has", "have", "had",
    # Modals
    "will", "would", "could", "should", "may", "might",
    # Market report verbs (all tenses / forms)
    "increase", "increased", "increases",
    "decrease", "decreased", "decreases",
    "rise", "rose", "rises",
    "fall", "fell", "falls",
    "grow", "grew", "grows",
    "decline", "declined", "declines",
    "remain", "remained", "remains",
    "continue", "continued", "continues",
    "expect", "expected", "expects",
    "project", "projected", "projects",
    "record", "recorded", "records",
    "reach", "reached", "reaches",
    "show", "showed", "shows",
    "indicate", "indicated", "indicates",
    "suggest", "suggested", "suggests",
    "reflect", "reflected", "reflects",
    "represent", "represented", "represents",
    "absorb", "absorbed", "absorbs",
    "lease", "leased", "leases",
    "complete", "completed", "completes",
    "deliver", "delivered", "delivers",
    "tighten", "tightened", "tightens",
    "ease", "eased", "eases",
    "expand", "expanded", "expands",
    "contract", "contracted", "contracts",
    "improve", "improved", "improves",
    "weaken", "weakened", "weakens",
    "exceed", "exceeded", "exceeds",
    "account", "accounted", "accounts",
    "underpin", "underpinned", "underpins",
    "constrain", "constrained", "constrains",
    "support", "supported", "supports",
    "drive", "driven", "drives",
    "total", "totalled", "totals",
})

# Matches data values: decimals (4.1%), percentages (3.2%), or large figures
# (1,250,000 sf).  Excludes bare 4-digit years (2025) so chart year-axis labels
# do not falsely qualify as table rows.
_DATA_NUM_RE = re.compile(r'\d+\.\d+|\d+%|[1-9][\d,]{4,}')

# Matches at least one alphabetic word ≥ 3 chars (rules out pure number lines)
_ALPHA_WORD_RE = re.compile(r'[A-Za-z]{3,}')


def _filter_prose_and_tables(text: str, min_page_chars: int = 80) -> str:
    """Keep prose paragraphs and table rows; discard chart/graph labels.

    The distinction is grammatical, not visual:

    KEEP — Prose lines
        A line containing a finite verb is a complete clause and therefore prose.
        "the" as a standalone token is a near-certain prose signal because
        definite articles appear in sentences but rarely in chart titles.
        Example kept:   "Vacancy tightened to 3.2% as net absorption remained positive."

    KEEP — Table rows
        A line with ≥ 4 words that contains both a meaningful data number
        (decimal, percentage, or large figure) and an alphabetic label.
        Example kept:   "Marina Bay  1,250,000  4.1%  SGD 12.50 psf/month"

    DROP — Chart / graph labels
        Lines that are pure noun phrases, axis labels, legend entries, or
        copyright notices — they lack a finite verb and don't match the
        table row pattern.
        Example dropped: "GRADE A CBD OFFICE GROSS EFFECTIVE RENT & VACANCY RATE"
        Example dropped: "© 2025 Cushman & Wakefield"
        Example dropped: "Q1 2025  Q2 2025  Q3 2025  Q4 2025"

    Pages where no lines survive (chart-only / infographic pages) are returned
    as empty strings so _chunk_pdf_by_page skips them and they never enter
    the RAG index.
    """
    # Pass 1: filter lines, preserving paragraph boundaries via None markers.
    # All-caps section headers (e.g. "RENTS ROSE AMID LOWER CBD VACANCIES") are
    # dropped as content but used as paragraph break signals, since pypdf does
    # not emit blank lines between paragraphs.
    kept: list[str | None] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            if kept and kept[-1] is not None:
                kept.append(None)
            continue

        words = line.split()
        if len(words) < 3:
            continue

        # Copyright / source attribution — never prose
        if line.startswith("©") or line.lower().startswith("source:") \
                or line.lower().startswith("note:"):
            continue

        tokens = {w.strip(".,;:()[]%$'\"").lower() for w in words}

        # All-caps section header (checked first — before finite-verb rule, since
        # headers like "RENTS ROSE AMID LOWER CBD VACANCIES" contain verbs):
        # drop content but insert a paragraph break signal.
        if line == line.upper() and len(words) >= 2:
            if kept and kept[-1] is not None:
                kept.append(None)
            continue

        # Prose: contains a finite verb → complete clause
        if tokens & _FINITE_VERBS:
            kept.append(line)
            continue

        # Prose: "the" as a standalone token (definite article in a sentence)
        if "the" in tokens:
            kept.append(line)
            continue

        # Table row: has a data number + an alphabetic label + enough words
        if (len(words) >= 4
                and _DATA_NUM_RE.search(line)
                and _ALPHA_WORD_RE.search(line)):
            kept.append(line)
            continue

        # Everything else: chart title, axis label, legend entry — drop

    # Pass 2: re-join consecutive prose lines with a space (same paragraph),
    # using the None markers as paragraph breaks → double newline in output.
    paragraphs: list[str] = []
    current: list[str] = []
    for token in kept:
        if token is None:
            if current:
                paragraphs.append(" ".join(current))
                current = []
        else:
            current.append(token)
    if current:
        paragraphs.append(" ".join(current))
    result = "\n\n".join(paragraphs)
    return result if len(result) >= min_page_chars else ""


def _chunk_pdf_by_page(pdf_path: Path) -> list[dict]:
    """Read a PDF and return one dict per page: {page, text, source_file}.

    Each page is filtered through _filter_prose_and_tables before indexing so
    the RAG embedding only sees readable prose and table rows — not chart titles,
    axis labels, or copyright lines.  Pages that are entirely chart/infographic
    (no readable text survives filtering) are skipped entirely, which also
    prevents the common p.1 clustering problem where the cover page's topic
    keywords match every claim.
    """
    try:
        import pypdf
    except ImportError:
        raise ImportError("pypdf is required.  pip install pypdf")
    reader = pypdf.PdfReader(str(pdf_path))
    chunks = []
    for i, pg in enumerate(reader.pages, 1):
        raw = pg.extract_text() or ""
        raw = _strip_cjk(raw).strip()
        filtered = _filter_prose_and_tables(raw)
        if filtered:
            chunks.append({"page": i, "text": filtered, "source_file": pdf_path.name})
    return chunks


def _embed_batch(texts: list[str], api_key: str,
                 model: str = "text-embedding-3-small") -> list[list[float]]:
    """Embed a list of texts using the OpenAI embeddings API.

    Sends in batches of 100 to stay well within API limits.
    Uses text-embedding-3-small: fast, cheap (~$0.002 / 1M tokens), and
    accurate enough for page-level semantic matching.
    """
    try:
        import openai as _openai
    except ImportError:
        raise ImportError("openai package required.  pip install openai")
    client  = _openai.OpenAI(api_key=api_key)
    results: list[list[float]] = []
    for i in range(0, len(texts), 100):
        batch = texts[i : i + 100]
        resp  = client.embeddings.create(input=batch, model=model)
        results.extend(item.embedding for item in resp.data)
    return results


def _cosine_sim(a: list[float], b: list[float]) -> float:
    """Cosine similarity between two embedding vectors (pure Python, no numpy)."""
    dot    = sum(x * y for x, y in zip(a, b))
    norm_a = sum(x * x for x in a) ** 0.5
    norm_b = sum(y * y for y in b) ** 0.5
    if norm_a == 0 or norm_b == 0:
        return 0.0
    return dot / (norm_a * norm_b)




def _build_audit_rag_index(
    extracted_reports: list[dict],
    openai_key: str,
    force_refresh: bool = False,
) -> list[dict]:
    """Build (or load cached) a flat list of {page, text, source_file, embedding} dicts.

    Embeddings are cached to disk alongside the Stage 1 JSON cache so unchanged
    PDFs are never re-embedded on subsequent runs.  Cache key uses the same
    filename + size + mtime strategy as Stage 1 extraction.
    """
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    all_chunks: list[dict] = []

    for rpt in extracted_reports:
        src_path = rpt.get("source_path", "")
        if not src_path or not Path(src_path).exists():
            continue
        p = Path(src_path)

        try:
            stat = p.stat()
            seed = f"{p.name}:{stat.st_size}:{stat.st_mtime}:rag"
        except OSError:
            seed = f"{p.name}:rag"
        cache_key  = hashlib.md5(seed.encode()).hexdigest()[:10]
        cache_file = CACHE_DIR / f"{p.stem}_{cache_key}_rag.json"

        if cache_file.exists() and not force_refresh:
            print(f"  [rag cache] {p.name}")
            cached = json.loads(cache_file.read_text(encoding="utf-8"))
            all_chunks.extend(cached.get("chunks", []))
            continue

        print(f"  [rag embed] {p.name} ({p.stat().st_size // 1024} KB) ...")
        chunks = _chunk_pdf_by_page(p)
        if not chunks:
            print(f"  [warning]   No extractable pages in {p.name}")
            continue
        print(f"  [rag embed] {len(chunks)} pages")

        try:
            embeddings = _embed_batch([c["text"] for c in chunks], openai_key)
        except Exception as e:
            print(f"  [warning]   Embedding failed for {p.name}: {e}")
            continue

        for chunk, emb in zip(chunks, embeddings):
            chunk["embedding"] = emb

        cache_data = {
            "source_file": p.name,
            "model":       "text-embedding-3-small",
            "built_at":    datetime.now().isoformat(),
            "chunks":      chunks,
        }
        cache_file.write_text(
            json.dumps(cache_data, ensure_ascii=False), encoding="utf-8"
        )
        print(f"  [rag cached] -> {cache_file.name}")
        all_chunks.extend(chunks)

    return all_chunks


def _rag_find_source(
    claim_embedding: list[float],
    rag_index: list[dict],
    claim_text: str,
    min_score: float = 0.20,
) -> dict | None:
    """Find the best-matching page chunk for a claim using a two-stage strategy.

    Stage 1 — Exact number matching (for claims with figures):
        Extract data values from the claim (decimals, percentages, large quantities;
        bare 4-digit years like 2025 are excluded to avoid matching every page).
        Only pages that contain at least one of those exact figures are considered.
        If no page contains the figure, the claim is unverifiable → return None.
        Cosine similarity then picks the best-ranked page among those candidates.

    Stage 2 — Pure cosine similarity (for claims with no numbers):
        If the claim contains no data figures (e.g. "demand driven by tech sector"),
        rank all pages by semantic similarity and return the best match above
        min_score.

    The returned chunk text is a full page (prose + tables, chart labels stripped)
    that is guaranteed to contain the claimed number.  No secondary sentence
    extraction is needed — the reviewer can see the number in context.
    """
    if not rag_index:
        return None

    # Data values only — exclude bare 4-digit years to avoid false positives
    claim_numbers = re.findall(r'\d+\.\d+%?|\d+%|[1-9][\d,]{4,}', claim_text)

    if claim_numbers:
        # Only consider pages that contain at least one claimed figure
        candidates = [
            chunk for chunk in rag_index
            if any(
                n.replace(",", "") in chunk["text"].replace(",", "")
                for n in claim_numbers
            )
        ]
        if not candidates:
            # Figure not found in any page — claim cannot be sourced
            return None
    else:
        candidates = rag_index

    scored = sorted(
        ((_cosine_sim(claim_embedding, chunk["embedding"]), chunk)
         for chunk in candidates),
        key=lambda x: x[0],
        reverse=True,
    )
    if not scored or scored[0][0] < min_score:
        return None

    score, best = scored[0]
    verbatim = bool(claim_numbers) and any(
        n.replace(",", "") in best["text"].replace(",", "")
        for n in claim_numbers
    )

    display = _extract_relevant_paragraphs(best["text"], claim_numbers)

    return {
        "source_file":   best["source_file"],
        "page":          best["page"],
        "text":          display,
        "citation_type": "Verbatim" if verbatim else "Paraphrased",
        "score":         score,
    }


def _extract_relevant_paragraphs(page_text: str, claim_numbers: list[str]) -> str:
    """Return only the paragraph(s) from page_text that contain a claimed number.

    Splits on double-newlines (paragraph boundaries produced by _filter_prose_and_tables),
    keeps every paragraph that contains at least one figure from claim_numbers, and
    joins them with a blank line.  Falls back to the full page text when no paragraph
    matches (e.g. qualitative claims with no numbers).
    """
    if not claim_numbers:
        return page_text
    paras = [p.strip() for p in page_text.split("\n\n") if p.strip()]
    hits = [
        p for p in paras
        if any(n.replace(",", "") in p.replace(",", "") for n in claim_numbers)
    ]
    return "\n\n".join(hits) if hits else page_text


# ═══════════════════════════════════════════════════════════════════════════════
# STAGE 1 — EXTRACT REPORT INSIGHTS
# ═══════════════════════════════════════════════════════════════════════════════

_EXTRACT_SYSTEM = (
    "You are a senior real estate research analyst at an institutional investment firm. "
    "Your job is to extract factual, quantitative market intelligence from property research reports. "
    "Always return valid JSON only — no markdown, no preamble, no explanation. "
    "LANGUAGE RULE: All JSON values must be written in English only. "
    "If the report contains text, names, or statistics in any non-English language "
    "(Korean, Japanese, Chinese, etc.), translate or transliterate them into English "
    "before writing them into the JSON. Never include non-English characters in any field."
)

_EXTRACT_PROMPT = """\
Extract the key real estate market facts from the report text below.

Focus on:
- Geography and property sectors covered
- Supply and demand dynamics (vacancy rates, net absorption, new supply)
- Rental levels and growth trends
- Capital values, investment yields, and transaction volumes
- Key demand drivers (economic, structural, occupier trends)
- Development pipeline and completions schedule
- Market outlook, forecasts, and risks
- Named government policies, regulations, or programmes
- Specific submarket breakdowns where data differs from city-wide averages
- Named occupier sectors or specific tenants driving demand
- Specific transactions cited as investment comparables

Write ALL JSON values in English only. Translate any Korean, Japanese, Chinese, or other
non-English text into English before writing it. Never include non-English characters.

Return ONLY a JSON object with these exact keys (use null if data not available):
{
  "country_region": "country or region covered",
  "sectors_covered": ["e.g. logistics, office, industrial"],
  "report_period": "time period e.g. 2025-2026",
  "market_overview": "2-3 sentences summarising the overall market context",
  "supply_demand": "key supply/demand facts — vacancy rates, absorption, completions",
  "rental_trends": "rental rate levels, recent changes, forecast growth",
  "capital_values_transactions": "yield levels, capital value trends, investment volumes",
  "demand_drivers": "main factors driving occupier and investor demand",
  "supply_pipeline": "under-construction and planned development pipeline",
  "market_outlook": "key forecast statements and risks",
  "named_policies": [
    "Every named government policy, regulation, GLS programme, or planning rule — exact names only. e.g. 'Industrial Government Land Sales Programme Q2 2026'. Empty list if none mentioned."
  ],
  "named_submarkets": [
    "Every specific submarket, district, precinct, or location cluster that has its own data point — include the figure. e.g. 'Jurong Industrial Estate: vacancy 2.1%, rental SGD 1.85 psf/month'. Empty list if none."
  ],
  "named_occupiers_sectors": [
    "Every specific named occupier sector or demand group cited with supporting context. e.g. '3PL operators: 42% of net absorption H1 2026', 'e-commerce: structural driver, growing 18% p.a.'. Empty list if none."
  ],
  "transaction_comparables": [
    "Every specific transaction cited as a comparable. Include price, yield, GFA, date where stated. Prefix with page number. e.g. '[p.12] ABC Logistics Hub sold SGD 280 psf, 5.5% NPI yield, 85,000 sqm, Q4 2025'. Empty list if none."
  ],
  "key_statistics": [
    "Prefix EVERY item with its page number using [p.N]. Extract up to 30 statistics — every quantitative figure in the report. e.g. '[p.5] Vacancy rate: 3.2% (Q1 2026, Grade A CBD)'. Write [p.?] if page cannot be determined."
  ]
}

Report text:
---
REPORT_TEXT_PLACEHOLDER
"""


def extract_report_insights(
    report_path: str | Path,
    llm_cfg: dict,
    openai_key: str = "",
    force_refresh: bool = False,
) -> dict:
    """
    STAGE 1 — Extract structured market facts from a single report PDF.

    Reads the PDF, sends the text to the LLM with _EXTRACT_PROMPT, and
    parses the response into a structured JSON dict.  Results are cached
    to disk so the same unchanged PDF is never re-processed.

    Cache invalidation:  cache key = MD5(filename + file_size + mtime).
    If the PDF changes (re-uploaded, updated), the hash changes automatically
    and the file is re-extracted on the next run.

    force_refresh=True skips the cache check and always re-runs the LLM.
    Use this if the extraction model was recently changed or the cache is bad.

    Returns a dict:
    {
        "source_file":  "filename.pdf",
        "source_path":  "/full/path/to/file.pdf",
        "extracted_at": "2026-06-04T12:00:00",
        "insights": {
            "market_overview": "...",
            "supply_demand": "...",
            "rental_trends": "...",
            "capital_values_transactions": "...",
            "demand_drivers": "...",
            "supply_pipeline": "...",
            "market_outlook": "...",
            "key_statistics": ["[p.5] Vacancy: 3.2%", ...]
        }
    }
    Returns {} if the file cannot be read or the LLM call fails.
    """
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    p = Path(report_path)

    # Cache key = filename + file size + mtime — re-extracts automatically if file changes
    try:
        stat = p.stat()
        cache_seed = f"{p.name}:{stat.st_size}:{stat.st_mtime}"
    except OSError:
        cache_seed = p.name
    cache_key  = hashlib.md5(cache_seed.encode()).hexdigest()[:10]
    cache_file = CACHE_DIR / f"{p.stem}_{cache_key}.json"

    # Serve from cache if available and not force-refreshing
    if cache_file.exists() and not force_refresh:
        print(f"  [cache]     {p.name}")
        return json.loads(cache_file.read_text(encoding="utf-8"))

    print(f"  [extracting] {p.name} ...")
    try:
        raw_text = _read_report(p)
    except Exception as e:
        print(f"  [warning]   Could not read {p.name}: {e}")
        return {}

    if not raw_text.strip():
        print(f"  [warning]   No extractable text in {p.name}")
        return {}

    # OpenAI GPT-4o has a 128k token context window — send up to 100k chars (full report).
    # Ollama local models typically have 8k–32k context; keep the conservative 14k limit.
    max_chars = 100_000 if llm_cfg.get("provider") == "openai" else 14_000
    truncated = _smart_truncate(raw_text, max_chars=max_chars)

    prompt = _EXTRACT_PROMPT.replace("REPORT_TEXT_PLACEHOLDER", truncated)

    try:
        response = _llm_chat(
            [
                {"role": "system", "content": _EXTRACT_SYSTEM},
                {"role": "user",   "content": prompt},
            ],
            llm_cfg=llm_cfg,
            openai_key=openai_key,
            temperature=0.1,
            json_mode=True,
        )
        insights = _parse_json_from_llm(response)
    except Exception as e:
        print(f"  [warning]   LLM extraction failed for {p.name}: {e}")
        # Don't cache failures — let the next run retry with a live LLM call
        return {
            "source_file":  p.name,
            "source_path":  str(p),
            "extracted_at": datetime.now().isoformat(),
            "insights":     {"extraction_error": str(e), "raw_excerpt": raw_text[:1500]},
        }

    result = {
        "source_file":  p.name,
        "source_path":  str(p),
        "extracted_at": datetime.now().isoformat(),
        "insights":     insights,
    }
    cache_file.write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  [cached]    -> {cache_file.name}")
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# STAGE 2 — GENERATE INVESTMENT RATIONALE
# ═══════════════════════════════════════════════════════════════════════════════


# ── Rationale generation prompts (Call 1) ────────────────────────────────────
# _RATIONALE_SYSTEM sets the persona, writing rules, data integrity constraints,
# style guidelines, and section title quality standards.  It is sent as the
# system message so its rules take highest precedence over the user prompt.
#
# _RATIONALE_PROMPT is the user message.  It contains placeholder tokens
# (ALL_CAPS) that are replaced at runtime with the real deal config values and
# the anonymised market research summary.  The anonymisation is critical — the
# LLM sees source data labelled "Research Report 1 / 2 / …", never real PDF
# filenames, so it cannot echo them into the prose.
_RATIONALE_SYSTEM = """\
You are a senior partner at a top-tier global institutional real estate investment fund.
You have spent 20 years writing investment committee memos for assets across Asia-Pacific,
Europe, and North America. Your writing is direct, authoritative, and evidence-dense.
Every sentence you write earns its place — no filler, no hedging, no throat-clearing.
IC readers are sophisticated; they expect conclusions first and evidence second, not
scene-setting. Write as if every word costs money.

━━ ABSOLUTE RULE — NEVER SIGNAL A DATA GAP ━━━━━━━━━━━━━━━━━━━━━
This rule overrides everything else. If data for a point is missing, omit that point
entirely and develop a different angle that IS supported by the research.
You must NEVER write sentences of the following type — these are career-ending in a
real investment committee memo:
  ✗ "While specific details on tenancy and WALE are not provided…"
  ✗ "Although vacancy data is not available for this submarket…"
  ✗ "In the absence of specific figures, the overall market suggests…"
  ✗ "While this information is limited, conditions appear favourable…"
  ✗ "Specific data on X has not been provided, however…"
  ✗ "Without precise figures, it can be inferred that…"
  ✗ Any sentence that opens with "While", "Although", or "Despite" followed
    by an acknowledgment that data is missing or incomplete.
Silence is professional. Flagging the gap is not. Omit and move on.

━━ DATA INTEGRITY ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Only write statistics and facts that appear explicitly in the Market Research Summary,
  the Comparable Evidence, or Deal Config. If a figure is not there, omit it — never
  estimate, round, or infer.
• Every number (percentage, rate, area, price, volume, yield, index value) must appear
  word-for-word or digit-for-digit in the source data. Do not derive or combine figures.
• NEVER write bracketed placeholders: [X%], [X.X%], [Y], [Z units], [p.?], [p.N], etc.
  If the exact figure is unavailable, omit the data point — no placeholder substitutes.
• Every policy name, regulation, or named programme must be explicitly stated in the
  Market Research Summary. Do not cite policies from memory or general knowledge.
• Never use general market knowledge. Every claim must trace to the research, the
  Comparable Evidence, or Deal Config.

━━ VOICE AND TONE ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Lead with the conclusion. State the investment thesis first, then the evidence.
  Do not build up to conclusions — senior readers want the answer in sentence one.
• Write with conviction. "Vacancy stands at 3.2% — the tightest level in a decade"
  not "Vacancy appears relatively low at approximately 3.2%."
• Active voice throughout. Present tense for market conditions; past tense for
  completed transactions.
• No attribution hedges: never write "according to", "the report states", "data shows",
  "it is noted that", "research indicates", or any similar phrase. State facts directly.
  Sources are tracked in the audit JSON — they never appear in the prose.
• Banned filler transitions: "additionally", "furthermore", "moreover", "in addition",
  "it is worth noting", "it should be noted", "lastly", "to summarise", "in conclusion".
• No bullet lists anywhere in the output. Continuous prose only.
• Each section: EXACTLY 2 paragraphs. Separate paragraphs with a blank line.
  Each paragraph: 120–180 words. Structure every paragraph as follows:
    – Sentence 1: bold investment thesis statement (the conclusion, stated upfront)
    – Sentences 2–4: qualitative reasoning — the structural or cyclical mechanism
      behind the thesis (why the dynamic exists, what is driving it, sector context)
    – Sentences 5–7: quantitative evidence — specific figures, rates, volumes, dates
      from the research that prove the thesis and its magnitude
    – Final sentence: investment implication — what this means for entry pricing,
      rental growth, capital value, or risk in this specific deal
  Never write a paragraph that is all numbers with no reasoning, or all reasoning
  with no numbers. Every paragraph must contain both.
• Always write exactly 4 sections. Add a 5th only when a genuinely distinct,
  data-supported angle cannot be absorbed into the first four.
• Complete every section in full. No trailing headings, JSON, or separators after the
  final section.

━━ LANGUAGE ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• English only. Translate or transliterate any non-English text before using it.
  Never include Korean, Japanese, Chinese, or other non-Latin characters.

━━ EVIDENCE DISCIPLINE ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Every investment conclusion must be anchored to a specific figure, named trend, or
  fact from the research. Plain assertions with no data anchor are not permitted.
• No single statistic or named fact may appear more than 3 times across all sections.
  Introduce each key data point once in its primary section; do not repeat mechanically.

━━ SECTION TITLES ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Write the title directly after the section number (## 1. / ## 2. / ## 3. / ## 4. / ## 5.)
— no brackets, no angle brackets, no quotes, no period.  Length: 6–9 words.

A title must do two things at once:
  (A) name a specific, evidenced market condition or asset characteristic
  (B) imply the investment consequence — why it matters to returns or capital value

RULES:
  • Do NOT include the property name, deal name, or location name in the title.
  • Do NOT use a colon to separate a category from a location (e.g. "Market Cycle: Property in City" is wrong).
  • The title must stand alone as an investment thesis statement.

STRONG titles — specific mechanism + investment implication (DO NOT copy these):
  "Structurally Under-Supplied Market Entering a Landlord-Favoured Cycle"
  "Prime Last-Mile Node Commanding Seoul's Fastest-Growing Demand Catchment"
  "Defensive Income Profile Anchored by Long-WALE Institutional Covenants"
  "Tight Development Pipeline Sustaining Above-Inflation Rental Reversion"

WEAK titles — never write anything that follows these patterns:
  ✗ Generic:     "Strong Fundamentals Supporting Logistics Investment Returns"
  ✗ Categorical: "Market Overview" / "Location Analysis" / "Deal Highlights"
  ✗ Vague:       "Positive Outlook for Industrial Property in the Region"
  ✗ Hedged:      "Potential Upside Amid Evolving Market Conditions"
  ✗ Property+Location: "Market Cycle Position: Property Name in City"

Vocabulary to draw from (use what fits the data):
  yield compression, structural undersupply, rental reversion, last-mile logistics,
  e-commerce penetration, constrained land release, institutional-grade covenant,
  WALE, net absorption, anchor tenant, leasehold profile, supply pipeline, cap rate,
  occupier demand, submarket tightness, development moratorium, value-add angle

Derive every title from the specific data in this prompt — country, submarket,
asset class, and the actual numbers from the research.  Original titles only."""

_RATIONALE_PROMPT = """\
Write an investment committee rationale for the subject property below.
Use only the Market Research Summary, the Comparable Evidence, and the Subject Property
details as your evidence base.

═══ SUBJECT PROPERTY ════════════════════════════════════════════
Deal Name     : DEAL_NAME
Property Name : PROPERTY_NAME
Address       : ADDRESS
Asset Class   : ASSET_CLASS
Asset Type    : ASSET_TYPE
Location      : LOCATION
Quality/Grade : QUALITY
GFA           : GFA_SF GFA_UNIT
Country       : COUNTRY_NAME
EXTRA_FIELDS
═════════════════════════════════════════════════════════════════

═══ MARKET RESEARCH SUMMARY ════════════════════════════════════
COMBINED_INSIGHTS
════════════════════════════════════════════════════════════════

COMPARABLE_EVIDENCE_BLOCK
═══ ANALYST NOTES ══════════════════════════════════════════════
ANALYST_NOTES
════════════════════════════════════════════════════════════════
REFINEMENT_BLOCK

STEP 1 — RETRIEVE DATA BEFORE WRITING ANYTHING
Do not write any prose yet.

Read through the entire Market Research Summary and Subject Property details.
Decide what distinct categories of investment-relevant information are actually present
in the research — do not use a fixed list of categories. Name the categories yourself
based on what the research actually contains.

GUIDANCE (examples of typical categories — not an exhaustive list, not all required):
  • Market cycle position — e.g. vacancy rates, net absorption, rental growth rate,
    cap rate or yield levels, investment volume, capital value trends, market outlook
  • Submarket or location dynamics — e.g. submarket vacancy vs city average,
    infrastructure projects, location cluster or precinct performance, rental premiums
    by district, migration of occupiers across submarkets
  • Demand drivers — e.g. occupier sector trends (e-commerce, 3PL, tech, financial
    services), structural demand shifts, ESG or green-building requirements, lease
    demand and pre-commitment activity, floor-plate or specification preferences
  • Supply pipeline — e.g. completions schedule, under-construction volume, land
    scarcity or development moratorium, GLS or planning approval pipeline,
    construction cost pressures
  • Deal-specific details — e.g. tenancy profile, WALE, lease expiry schedule,
    anchor tenant, occupancy rate, passing rent vs market rent, pricing, GFA,
    remaining leasehold, reversionary potential
  • Comparable pricing evidence — price psf and cap rates of comparable asset sales,
    asking or effective rents of leasing comparables, land price psf ppr of comparable
    land sales, and how the subject's pricing sits versus those comparable averages
    (all drawn from the Comparable Evidence block, when present)
  • Other categories the research supports — e.g. investment market activity,
    ESG premium evidence, structural sector shifts with their own data set

Aim for 3–6 categories. Merge closely related data into one rather than splitting
into many thin categories. You may use fewer or more depending on what the research
actually contains.

For each category you identify:
  — Name the category clearly so you can reference it in STEP 2.
  — Quote every specific figure, percentage, named fact, and direct data point found.
  — Only list data explicitly stated in the Market Research Summary or Deal Config.
    No inference, no general knowledge, no paraphrasing that changes the numbers.
  — Write "— none found —" if a category has no data in either source.

STEP 2 — PLAN SECTIONS
Using only the data retrieved in STEP 1:
  a) You will always write exactly 4 sections. Add a 5th only when a genuinely distinct,
     data-supported angle cannot be absorbed into the first four.
  b) Assign each category to the section where it will serve as primary evidence.
     Categories with "— none found —" must not drive any section — omit that angle entirely.
  c) State in one sentence the investment thesis each section will argue.

STEP 3 — WRITE THE SECTIONS
For each section, replace the [write a title here...] instruction on the ## line with your
actual title — 6–9 words, no property name, no location name, no colon separating a category
from a location. Just the title itself, nothing else on that line.
Do NOT mention any report name, file name, or source document anywhere in the sections.
Do NOT use phrases like "according to", "the report indicates", or any attribution.
Write every fact as a direct, unattributed statement. All source tracking goes in the audit JSON only.

Use only the data points you listed in STEP 1. If a figure or named fact was not listed
in STEP 1, it must not appear in the prose — do not add new data at the writing stage.

Each section: EXACTLY 2 paragraphs (3 only if data clearly supports a third distinct point).
Each paragraph: 80–130 words. Separate paragraphs with a blank line.
Pack in specific figures and data points — concise and evidence-dense beats long and vague.

PRESCRIBED SECTIONS (always write all four):

## 1. [write a 6–9 word investment thesis title here — derived from your STEP 1 market-cycle data]
Draw from your market-cycle and supply-pipeline categories (whatever you named them in STEP 1).
Paragraph 1 — supply/demand balance: vacancy trajectory, net absorption, new completions,
  landlord pricing power. Anchor every sentence to a specific figure.
Paragraph 2 — rental and capital value momentum: rental growth rate and direction, yield
  context, development pipeline constraints, market outlook and entry timing thesis.

## 2. [write a 6–9 word investment thesis title here — derived from your STEP 1 location/demand data]
Draw from your location/submarket and demand-driver categories.
Paragraph 1 — submarket positioning: why this specific submarket commands premium occupier
  demand; performance vs city-wide average; proximity to key infrastructure or logistics nodes.
Paragraph 2 — demand drivers: structural occupier trends driving take-up; catchment depth;
  land scarcity or regulatory factors that entrench the location's defensibility.

## 3. [write a 6–9 word investment thesis title here — derived from your STEP 1 deal-specific data]
Draw from your deal-specific category. Use whichever deal-specific angles are supported by
the research: asset quality, green certification, GFA, passing rent vs market rent,
reversionary potential, pricing vs comparable transactions, key risks and mitigants.
If tenancy / WALE data is not in the research, develop the asset quality and pricing
angles instead — never acknowledge the absence of tenancy data in the prose.
Paragraph 1 — asset case: lead with the property's quality, specification, certification,
  and market positioning; then connect to reversionary or income upside.
Paragraph 2 — deal case: pricing vs comparable transactions (cite the Comparable Evidence —
  the subject's price psf / cap rate versus the comparable average and range), capital
  appreciation thesis, key risks and specific mitigants, why compelling at this cycle point.

## 4. [write a 6–9 word investment thesis title here — derived from your STEP 1 capital-markets data]
Draw from your capital values, transaction volumes, and yield data categories.
Paragraph 1 — investment market: recent transaction volumes, investor appetite, yield
  compression or expansion trend, and the comparable deals in the Comparable Evidence that
  benchmark this asset's pricing (name the price psf / cap-rate spread versus the average).
Paragraph 2 — capital value outlook: yield trajectory, cap rate context vs historical range,
  why the pricing is supportable and what drives capital appreciation from here.

OPTIONAL FIFTH SECTION (add only if your STEP 1 categories contain a genuinely distinct angle
that cannot be absorbed into sections 1–4 — e.g. a quantified ESG premium, a named supply
moratorium, or a structural demand shift with its own data set):

## 5. [write a 6–9 word investment thesis title here — only if a distinct data-supported angle exists]

STEP 4 — VERIFY BEFORE RETURNING
Before returning your response, check every section against this list.
Fix any failure before returning — do not return output that fails a check.

  [ ] RETRIEVAL    — Every data point in the prose was listed in STEP 1. Any figure or
                     named fact that was NOT in your STEP 1 list must be deleted from the prose.
  [ ] LANGUAGE     — No non-English characters anywhere. All foreign terms translated.
  [ ] DATA ANCHOR  — Every opinion, forecast, or investment conclusion contains or directly
                     follows a specific figure, percentage, or named fact. No plain assertions.
  [ ] SOURCE CHECK — Every number (vacancy rate, cap rate, rental rate, yield, price, volume,
                     GFA, growth %, index value) must appear explicitly in the Market Research
                     Summary, the Comparable Evidence, or Deal Config. Delete any figure you
                     cannot locate there.
  [ ] POLICY CHECK — Every policy name, regulation, government initiative, or named programme
                     must be explicitly named in the Market Research Summary. Delete any policy
                     reference you cannot locate there.
  [ ] REUSE LIMIT  — No single statistic or named fact appears more than 3 times in total
                     across all sections. If it does, remove the duplicate and replace with
                     a different supporting data point or omit that sentence.
  [ ] PARAGRAPHS   — Each section has exactly 2 paragraphs (3 only if data demands it).
                     Paragraphs are separated by a blank line. Each paragraph is 80–130 words.
  [ ] SECTION COUNT — Exactly 4 sections are present (5 only if a distinct angle exists).
  [ ] GAP-FLAGGING — Zero sentences signal a missing data point. Search for "not provided",
                     "not available", "without specific", "in the absence", "while specific",
                     "although X data", "cannot be determined". Delete any such sentence and
                     replace with a different data-supported point, or omit entirely.
  [ ] ATTRIBUTION  — No sentence contains "according to", "the report", "data shows",
                     or any source reference. No report filename anywhere.
  [ ] TRANSITIONS  — No banned transition words anywhere in the output.

No fabrication. No data gaps flagged in prose. Use only research and deal config data.
Write ALL sections completely. Stop immediately after the last paragraph of your final section.
Do not add any headings, JSON, commentary, or separator after the final section.
"""


def _merge_insights(extracted_reports: list[dict], anonymize: bool = False) -> str:
    """
    Combine cached insights from multiple reports into one text block for the prompt.

    Called twice in Stage 2 with different anonymize settings:

    Call 1 (prose writing)  — anonymize=True
        Source labels become "Research Report 1", "Research Report 2", etc.
        The LLM physically cannot echo real PDF filenames into the prose
        because it never sees them.  This keeps client-facing memos clean.

    Call 2 (source audit)   — anonymize=False
        Source labels show the real PDF filename (e.g. "AEWResearch_Korea.pdf").
        The audit LLM needs the real names to populate Source_Audit.xlsx correctly.

    CJK characters are stripped from every field before the text is included,
    ensuring no Korean/Japanese/Chinese text reaches Stage 2.
    """
    parts = []
    for idx, rpt in enumerate(extracted_reports):
        src = rpt.get("source_file", "Unknown")
        ins = rpt.get("insights", {})
        if not ins:
            continue

        label = f"Research Report {idx + 1}" if anonymize else f"Source: {src}"
        lines = [label]

        field_labels = [
            ("market_overview",             "Market Overview"),
            ("supply_demand",               "Supply & Demand"),
            ("rental_trends",               "Rental Trends"),
            ("capital_values_transactions", "Capital Values & Transactions"),
            ("demand_drivers",              "Demand Drivers"),
            ("supply_pipeline",             "Supply Pipeline"),
            ("market_outlook",              "Market Outlook"),
        ]
        for key, label in field_labels:
            val = ins.get(key)
            if val:
                lines.append(f"{label}: {_strip_cjk(str(val))}")

        # Enriched list fields extracted by GPT (absent in older Ollama caches — skip gracefully)
        list_field_labels = [
            ("named_policies",          "Named Policies & Regulations"),
            ("named_submarkets",        "Submarket Breakdown"),
            ("named_occupiers_sectors", "Named Occupiers & Sectors"),
            ("transaction_comparables", "Transaction Comparables"),
        ]
        for key, label in list_field_labels:
            val = ins.get(key)
            if val and isinstance(val, list):
                lines.append(f"{label}: " + " | ".join(
                    _strip_cjk(str(s)) for s in val[:10]))

        stats = ins.get("key_statistics", [])
        if stats:
            lines.append("Key Data Points: " + " | ".join(
                _strip_cjk(str(s)) for s in stats[:30]))

        parts.append("\n".join(lines))

    if not parts:
        return "(No market research provided — rationale will be based on property specifics only.)"
    return "\n\n─────\n\n".join(parts)


def _clean_rationale_body(text: str) -> str:
    """
    Remove any trailing audit/JSON sections the LLM sometimes appends after the prose.

    Some models (especially smaller ones) ignore the instruction to stop after the
    last section and append headings like "## Source Audit", "## Citations", or
    a raw JSON block.  This function detects those headings and truncates the text
    at the first one, returning only the clean prose sections.
    """
    audit_hdr = re.compile(
        r'^#{1,4}\s*(source\s*audit|audit\s*json|audit|citations?|references?|json)\s*$',
        re.IGNORECASE,
    )
    lines = text.splitlines()
    clean = []
    for line in lines:
        if audit_hdr.match(line.strip()):
            break          # stop collecting at first audit heading
        clean.append(line)
    # Drop trailing blank lines
    while clean and not clean[-1].strip():
        clean.pop()
    return "\n".join(clean)


# ── Audit-call prompts (separate second LLM call) ─────────────────────────────

_AUDIT_SYSTEM = (
    "You are a citation auditor for an institutional real estate investment memo. "
    "You identify specific claims and match each one to its source. "
    "Return ONLY a valid JSON array — no markdown, no preamble, no explanation."
)

_AUDIT_PROMPT_TEMPLATE = """\
Audit the investment rationale below against the market research sources provided.
You must capture a row for EVERY auditable item. Do not skip anything that contains a
specific claim, figure, official statement, policy reference, or named data point.

Audit ALL of the following — every instance must appear as a separate row:

DATA QUOTES — every quantitative figure cited in the rationale:
  - Vacancy rate, occupancy rate, take-up, net absorption, pre-commitment rate
  - Cap rate, NPI yield, net yield, passing rent, market rent, rental reversion
  - Price, price psf, GFA, site area, transaction volume, deal count
  - Completions, pipeline supply, under-construction quantum
  - Growth rate, CAGR, index level, percentage change, basis points move
  - Any other number, ratio, or metric used to support an argument

NAMED QUALITATIVE CLAIMS — every specific non-numeric assertion tied to a named fact:
  - Named trend, structural shift, or outlook statement tied to a direction or timeframe
  - Named location, precinct, infrastructure node, or submarket used to justify positioning
  - Named occupier sector, tenant type, or demand driver (e.g. "e-commerce", "3PL", "data centres")
  - Named building, development, or asset used as a comparable or benchmark
  - Any ESG certification, green building standard, or sustainability requirement cited

Only skip sentences that are purely structural connective tissue with zero data content
(e.g. "The asset benefits from strong fundamentals" with no figures, names, or specific claims).
Every number and every named fact in the rationale MUST appear as a row.

INVESTMENT RATIONALE:
────────────────────────────────────────────────────────────────
RATIONALE_BODY_PLACEHOLDER
────────────────────────────────────────────────────────────────

RESEARCH SOURCES (read these to match claims):
────────────────────────────────────────────────────────────────
SOURCE_DATA_PLACEHOLDER
────────────────────────────────────────────────────────────────

VALID source_file VALUES — copy one of these EXACTLY (character-for-character):
VALID_FILENAMES_PLACEHOLDER
  - Deal Config

Do NOT invent filenames, abbreviate them, or write "Research Report N".
Do NOT use "General Knowledge" — every claim must trace to one of the research reports or Deal Config.
Every source_file entry must be one of the exact strings listed above.

Return ONLY a JSON array, one entry per specific claim:
[
  {
    "section_num": 1,
    "section_title": "exact title of that section from the rationale",
    "claim": "specific assertion, data point, or official statement from the rationale text",
    "source_file": "one of the exact valid values listed above",
    "page_ref": "page number where this data appears, e.g. 'p.5' — use 'p.?' if uncertain, null if Deal Config",
    "supporting_text": "verbatim or near-verbatim passage from the research — null if Deal Config",
    "citation_type": "Verbatim | Paraphrased | Deal Config"
  }
]

citation_type guide:
  Verbatim    — the rationale quotes the source almost word-for-word
  Paraphrased — the rationale restates a figure or finding in different words
  Deal Config — the claim originates from the deal configuration / subject property data
"""


# ── Claim extraction prompt (RAG audit step 1: enumerate claims before matching) ──

_CLAIM_EXTRACT_SYSTEM = (
    "You are an auditor extracting verifiable factual claims from an investment memo. "
    "Return only a JSON array — no markdown, no preamble, no explanation."
)

_CLAIM_EXTRACT_PROMPT = """\
Extract every specific, verifiable claim from the investment rationale below.

Include every:
  - Quantitative figure: vacancy rate, rental rate, cap rate, yield, price, GFA, volume, growth %
  - Named trend or structural shift with a direction or timeframe
  - Named location, submarket, or infrastructure node used to support positioning
  - Named occupier sector or demand group with supporting context
  - Named government policy, regulation, or programme
  - Specific transaction, building, or comparable asset cited

Exclude purely transitional sentences that contain zero data content.

For each claim record:
  - section_num:   integer section number (1, 2, 3 …)
  - section_title: exact title of that section from the rationale
  - claim:         the specific self-contained claim text — one fact or figure per entry

Return ONLY a JSON array:
[
  {"section_num": 1, "section_title": "...", "claim": "specific fact or figure"},
  ...
]

INVESTMENT RATIONALE:
────────────────────────────────────────────────────────────────
RATIONALE_BODY_PLACEHOLDER
────────────────────────────────────────────────────────────────
"""


# ── Source audit Excel writer ──────────────────────────────────────────────────

def _cross_check_claim(claim_entry: dict, extracted_reports: list[dict]) -> str:
    """
    Python-level verification: check whether a citation's supporting text
    actually appears in the cached Stage 1 extract for that source file.

    This is a fast automated check — it does NOT re-read the original PDF.
    It searches the cached extract text (which is the LLM-summarised version,
    not the full PDF text), so results are approximate:

    ✓  Found in cached extract         — 4-word sliding window matched
    ⚠  Not found in cached extract     — no match; reviewer must check original PDF
    ⚠  No supporting text provided     — audit LLM left supporting_text blank
    ⚠  Source file not in selected reports — filename not found in any loaded report

    Matching strategy: 4-word sliding window — any 4 consecutive words from the
    supporting text must appear somewhere in the cached extract text.
    This is loose enough to handle minor paraphrasing but strict enough to
    catch complete fabrications.

    Two passes:
    1. Exact filename match — looks for source_file == rpt["source_file"]
    2. Fuzzy match — handles truncation differences (e.g. "AEWResearch..." vs full name)
    """
    src_file = claim_entry.get("source_file", "")
    support  = claim_entry.get("supporting_text") or ""
    ctype    = claim_entry.get("citation_type", "")

    # Deal Config citations come from the config file, not a research report — skip
    if ctype in ("Deal Config", "General Knowledge") or src_file in ("Deal Config", "General Knowledge"):
        return ctype

    if not support.strip():
        return "⚠  No supporting text provided"

    def _flatten(rpt: dict) -> str:
        """Flatten all cached insight fields into a single lowercase string for matching."""
        ins = rpt.get("insights", {})
        txt = " ".join(str(v) for v in ins.values() if v and not isinstance(v, list))
        txt += " " + " ".join(str(s) for s in (ins.get("key_statistics") or []))
        return txt.lower()

    def _text_match(support: str, full_text: str) -> bool:
        """Return True if any 4-word window from support appears in full_text."""
        words  = support.lower().split()
        chunks = [" ".join(words[i:i+4]) for i in range(max(len(words) - 3, 1))]
        return any(c in full_text for c in chunks)

    # Pass 1 — exact filename match
    for rpt in extracted_reports:
        if rpt.get("source_file", "") == src_file:
            full_text = _flatten(rpt)
            if _text_match(support, full_text):
                return "✓  Found in cached extract"
            return "⚠  Not found in cached extract — verify against original PDF"

    # Pass 2 — fuzzy filename match (handles truncation / casing differences)
    src_lower = src_file.lower()
    for rpt in extracted_reports:
        rpt_src = rpt.get("source_file", "").lower()
        if rpt_src and (src_lower in rpt_src or rpt_src in src_lower):
            full_text = _flatten(rpt)
            if _text_match(support, full_text):
                return "✓  Found in cached extract (fuzzy filename match)"
            return "⚠  Not found in cached extract — verify against original PDF"

    return "⚠  Source file not in selected reports"


def _get_cross_check(entry: dict, extracted_reports: list[dict]) -> str:
    """Return the verification status string for one audit entry.

    If the entry was produced by the RAG audit (has a rag_score field), the
    status is derived from the cosine similarity score.  Otherwise falls back
    to the original 4-word sliding-window text match against the cached extract.
    """
    rag_score = entry.get("rag_score")
    ctype     = entry.get("citation_type", "")

    if rag_score is not None:
        if ctype in ("Deal Config", "General Knowledge"):
            return ctype
        page = (entry.get("page_ref") or "?")
        if rag_score >= 0.5:
            return f"✓  RAG match {rag_score:.0%} ({page})"
        elif rag_score >= 0.3:
            return f"⚠  Moderate RAG match {rag_score:.0%} ({page}) — verify"
        else:
            return f"⚠  Low RAG confidence {rag_score:.0%} — verify against PDF"

    return _cross_check_claim(entry, extracted_reports)


def _write_source_audit_excel(
    audit_entries: list[dict],
    extracted_reports: list[dict],
    out_path: Path,
    subject_cfg: dict,
) -> None:
    """
    Write Source_Audit.xlsx — a formatted Excel table for human citation review.

    Each row = one specific claim from the investment rationale, with:
    - Which section and section title it came from
    - The exact claim text
    - The source PDF filename and page reference
    - The LLM's verbatim supporting quote from the research
    - Citation type (Verbatim / Paraphrased / Deal Config)
    - Backend cross-check status (auto-verified against the cached extract)
    - Blank columns for the human reviewer to mark CONFIRMED / INCORRECT / CANNOT VERIFY

    Row colour coding:
    - Red    = claim not found in cached extract → must verify against original PDF
    - Orange = no supporting text provided by audit LLM → review needed
    - White/Grey = alternating rows, auto-verified or Deal Config

    Columns
    -------
    #  |  Section  |  Section Title  |  Claim / Data Point  |  Source Report  |
    Supporting Text (LLM quote)  |  Citation Type  |  Backend Cross-Check  |
    Human Validation  |  Reviewer Notes
    """
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Source Audit"

    # ── Palette ───────────────────────────────────────────────────────────────
    NAVY    = "FF1A3A5C"
    WHITE   = "FFFFFFFF"
    LGREY   = "FFF2F5F8"
    DGREY   = "FFE0E7EF"
    GREEN   = "FFD6F0DC"
    ORANGE  = "FFFFF3CD"
    RED_BG  = "FFFDE8E8"
    AMBER   = "FFFF8C00"
    DARK    = "FF1A1A1A"

    def _hdr_font(bold=True):
        return Font(name="Calibri", bold=bold, color=WHITE, size=10)
    def _body_font(bold=False, color=DARK):
        return Font(name="Calibri", bold=bold, color=color, size=9)
    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)
    def _align(h="left", v="top", wrap=True):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def _thin_border():
        s = Side(style="thin", color="FFD0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)

    # ── Title row ─────────────────────────────────────────────────────────────
    prop_name = subject_cfg.get("property_name", subject_cfg.get("deal_name", "Property"))
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = (f"Source Audit — {prop_name}  "
                        f"({subject_cfg.get('asset_class','').title()}, "
                        f"{subject_cfg.get('country_name','')})  |  "
                        f"Generated {datetime.now().strftime('%d %b %Y')}")
    title_cell.font      = Font(name="Calibri", bold=True, color=WHITE, size=11)
    title_cell.fill      = _fill(NAVY)
    title_cell.alignment = _align("left", "center", wrap=False)
    ws.row_dimensions[1].height = 22

    # Instruction row
    ws.merge_cells("A2:K2")
    note_cell = ws["A2"]
    note_cell.value = (
        "VALIDATION INSTRUCTIONS:  Review every row where Backend Cross-Check = "
        "⚠  Not found in cached extract.  Open the original PDF and verify the claim.  "
        "Mark Human Validation as CONFIRMED / INCORRECT / CANNOT VERIFY, and add notes."
    )
    note_cell.font      = Font(name="Calibri", italic=True, color="FF555555", size=8)
    note_cell.fill      = _fill("FFFFF8DC")
    note_cell.alignment = _align("left", "center", wrap=True)
    ws.row_dimensions[2].height = 28

    # ── Column headers ─────────────────────────────────────────────────────────
    HEADERS = [
        "#",
        "Section",
        "Section Title",
        "Claim / Data Point Used",
        "Source Report",
        "Page\nRef",
        "LLM Supporting Text\n(verbatim quote provided by LLM)",
        "Citation Type",
        "Backend Cross-Check\n(auto-verified against cached extract)",
        "Human Validation\n(CONFIRMED / INCORRECT / CANNOT VERIFY)",
        "Reviewer Notes",
    ]
    COL_WIDTHS = [4, 9, 38, 45, 32, 8, 45, 14, 34, 20, 28]

    for col_idx, (hdr, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.font      = _hdr_font()
        cell.fill      = _fill(NAVY)
        cell.alignment = _align("center", "center", wrap=True)
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 36

    # ── Data rows ──────────────────────────────────────────────────────────────
    for row_num, entry in enumerate(audit_entries, 1):
        row_idx    = row_num + 3   # rows 4+
        cross_chk  = _get_cross_check(entry, extracted_reports)
        ctype      = entry.get("citation_type", "")

        # Row background
        if "⚠  Not found" in cross_chk:
            row_fill = _fill(RED_BG)
        elif "⚠  No supporting" in cross_chk:
            row_fill = _fill(ORANGE)
        elif row_num % 2 == 0:
            row_fill = _fill(LGREY)
        else:
            row_fill = _fill(WHITE)

        row_data = [
            row_num,
            f"Section {entry.get('section_num', '?')}",
            entry.get("section_title", ""),
            entry.get("claim", ""),
            entry.get("source_file", ""),
            entry.get("page_ref") or "—",
            entry.get("supporting_text") or "",
            ctype,
            cross_chk,
            "",   # Human Validation — blank for reviewer
            "",   # Reviewer Notes — blank for reviewer
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell            = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill       = row_fill
            cell.border     = _thin_border()
            cell.alignment  = _align(wrap=True)

            # Special formatting  (cols shifted +1 after adding Page Ref at col 6)
            if col_idx == 1:  # row number
                cell.alignment = _align("center", "top", wrap=False)
                cell.font      = _body_font(color="FF888888")
            elif col_idx == 6:  # page reference
                cell.alignment = _align("center", "top", wrap=False)
                cell.font      = _body_font(bold=True, color="FF1A3A5C")
            elif col_idx == 8:  # citation type
                color = AMBER if ctype in ("Inferred", "General Knowledge") else DARK
                cell.font = _body_font(bold=(ctype == "Inferred"), color=color)
            elif col_idx == 9:  # cross-check status
                if "✓" in cross_chk:
                    cell.font = _body_font(color="FF1A7A2A")
                elif "⚠" in cross_chk:
                    cell.font = _body_font(bold=True, color="FFC00000")
                else:
                    cell.font = _body_font(color="FF555555")
            elif col_idx == 10:  # human validation (blank but highlighted)
                cell.fill = _fill(GREEN)
                cell.font = _body_font(color="FF1A7A2A")
            else:
                cell.font = _body_font()

        ws.row_dimensions[row_idx].height = 52

    # ── Summary stats at the bottom ───────────────────────────────────────────
    total      = len(audit_entries)
    n_verified = sum(1 for e in audit_entries
                     if "✓" in _get_cross_check(e, extracted_reports))
    n_warn     = sum(1 for e in audit_entries
                     if "⚠" in _get_cross_check(e, extracted_reports))
    n_inferred = sum(1 for e in audit_entries
                     if e.get("citation_type") in ("Inferred", "General Knowledge"))

    sum_row = total + 4 + 2
    ws.merge_cells(f"A{sum_row}:K{sum_row}")
    sum_cell = ws[f"A{sum_row}"]
    sum_cell.value = (
        f"SUMMARY:  {total} citations total  |  "
        f"{n_verified} verified in cached extract  |  "
        f"{n_warn} require manual PDF verification  |  "
        f"{n_inferred} inferred / general knowledge (review carefully)"
    )
    sum_cell.font      = Font(name="Calibri", bold=True, color=WHITE, size=9)
    sum_cell.fill      = _fill(NAVY)
    sum_cell.alignment = _align("left", "center", wrap=False)
    ws.row_dimensions[sum_row].height = 18

    # Freeze panes below header
    ws.freeze_panes = "A4"

    wb.save(str(out_path))
    print(f"  [source audit] -> {out_path.name}")


# ── Parsing helper ─────────────────────────────────────────────────────────────

def _parse_audit_json(raw_audit: str) -> list[dict]:
    """
    Extract and parse the source audit JSON array from the LLM's raw response.

    The audit LLM is instructed to return a bare JSON array, but may still wrap
    it in markdown fences or add preamble text.  This function handles both cases:
    1. Strips ```json ... ``` fences if present
    2. Finds the first [ ... ] array block
    3. Returns [] if no valid JSON array can be parsed (triggers fallback)
    """
    text = raw_audit.strip()
    # Strip markdown fences
    m = re.search(r"```(?:json)?\s*([\s\S]*?)```", text)
    if m:
        text = m.group(1).strip()
    # Find JSON array
    m2 = re.search(r"\[[\s\S]*\]", text)
    if m2:
        text = m2.group(0)
    try:
        result = json.loads(text)
        if isinstance(result, list):
            return result
    except Exception as e:
        print(f"  [warning] Could not parse source audit JSON: {e}")
    return []


def _fallback_citations_from_text(rationale_body: str) -> list[dict]:
    """
    Last-resort fallback if the audit LLM fails to return valid JSON.

    Scans the rationale prose for inline [SOURCE: filename] markers and
    builds a minimal citation list from them.  These markers are not normally
    present in the output (the main pipeline uses a separate audit call), but
    some model configurations may produce them as a side effect.
    Returns [] if no markers are found.
    """
    entries = []
    # Find [SOURCE: something] markers and the preceding sentence
    pattern = re.compile(
        r'([^.\n]{10,200}?)\s*\[SOURCE:\s*([^\]]+)\]',
        re.DOTALL,
    )
    for m in pattern.finditer(rationale_body):
        claim     = m.group(1).strip().rstrip(",;:")
        src_raw   = m.group(2).strip()
        entries.append({
            "section_num":    "?",
            "section_title":  "",
            "claim":          claim,
            "source_file":    src_raw,
            "supporting_text": None,
            "citation_type":  "Paraphrased",
        })
    return entries


def _build_input_summary(extracted_reports: list) -> str:
    """Build the 'Input Summary' section from the Stage-1 extracted report facts.

    Summarises, per input PDF, the key points and what they convey for this
    investment — the evidence base that shapes the rationale that follows. Built
    directly from the already-extracted structured facts (no extra LLM call).
    """
    reports = [r for r in (extracted_reports or []) if r.get("source_file")]
    if not reports:
        return ""

    out = [
        "## Input Summary\n",
        "_Key points drawn from the market report(s) provided, and what they convey "
        "for this investment — the evidence base for the rationale that follows._\n",
    ]
    for r in reports:
        src = r.get("source_file", "report")
        # The extracted market facts live under the "insights" sub-dict; fall back to
        # the top level for older/flat caches.
        ins = r.get("insights") if isinstance(r.get("insights"), dict) else r
        sectors = ins.get("sectors_covered")
        sectors = ", ".join(sectors) if isinstance(sectors, list) else (sectors or "")
        meta    = " · ".join(x for x in (ins.get("country_region") or "", sectors,
                                         ins.get("report_period") or "") if x)
        out.append(f"### {src}" + (f"  \n_{meta}_" if meta else ""))

        _before = len(out)
        overview = (ins.get("market_overview") or "").strip()
        if overview:
            out.append(overview)

        # "What it conveys" — the decision-relevant angles, when present.
        for label, key in (
            ("Demand drivers",            "demand_drivers"),
            ("Supply & demand",           "supply_demand"),
            ("Rents",                     "rental_trends"),
            ("Capital values / yields",   "capital_values_transactions"),
            ("Development pipeline",       "supply_pipeline"),
            ("Outlook & risks",           "market_outlook"),
        ):
            val = (ins.get(key) or "").strip()
            if val:
                out.append(f"- **{label}:** {val}")

        stats = [s.strip() for s in (ins.get("key_statistics") or [])
                 if isinstance(s, str) and s.strip()][:3]
        if stats:
            out.append("- **Notable figures:** " + "; ".join(stats))

        if len(out) == _before:   # nothing extracted → tell the user why
            _why = ("extraction error" if ins.get("extraction_error")
                    else "no text extracted — likely a scanned / image-only PDF")
            out.append(f"- _No summary available ({_why}). Verify the source PDF._")
        out.append("")   # blank line between reports

    out.append("---\n")
    return "\n".join(out) + "\n"


def generate_rationale(
    extracted_reports: list[dict],
    subject_cfg: dict,
    llm_cfg: dict,
    openai_key: str = "",
    analyst_notes: str = "",
    refinement_notes: str = "",
    force_refresh: bool = False,
    comp_summary: str = "",
) -> tuple[str, list]:
    """
    Two-call LLM pipeline that produces the investment rationale and its
    source audit.

    WHY TWO SEPARATE CALLS?
    ─────────────────────────────────────────────────────────────────────────
    A single call that sees the real PDF filenames will naturally echo them
    into the prose — producing sentences like "According to AEWResearch_
    AsiaPacific_South-Korea-Logistics_WEBSITE.pdf, vacancy is 3%…", which is
    unprofessional in a client-facing investment memo and exposes internal
    source filenames.

    Splitting into two calls solves this cleanly:

    Call 1 — Prose writing (anonymised sources)
        The research data is passed with source labels replaced by generic
        names ("Research Report 1", "Research Report 2", …) via
        _merge_insights(anonymize=True).  The LLM physically cannot echo a
        real filename because it never sees one.  Output: clean 3-section
        investment rationale markdown.

    Call 2 — Source audit JSON (real filenames visible)
        The finished prose and the research data WITH real filenames are sent
        to a separate LLM call whose only job is to match each specific claim
        in the prose back to its real source PDF and page number.  Output: a
        JSON array of citation entries that populates Source_Audit.xlsx.
        This JSON is never shown in the front-end; it stays in the Excel only.

    Returns (rationale_markdown: str, audit_entries: list[dict])
    """
    # Build optional extra deal fields
    sym   = subject_cfg.get("currency_symbol", "")
    extra = []
    if subject_cfg.get("price_sgd_m"):
        extra.append(f"Asking Price  : {sym}{subject_cfg['price_sgd_m']:.1f}M")
    if subject_cfg.get("ftm_noi_cap_rate"):
        extra.append(f"Cap Rate (FTM): {subject_cfg['ftm_noi_cap_rate']*100:.2f}%")
    if subject_cfg.get("remaining_leasehold_yrs"):
        extra.append(f"Leasehold     : {subject_cfg['remaining_leasehold_yrs']} yrs remaining")
    if subject_cfg.get("land_zoning"):
        extra.append(f"Zoning        : {subject_cfg['land_zoning']}")

    gfa_display = (f"{int(subject_cfg.get('gfa_sf', 0)):,}"
                   if subject_cfg.get("gfa_sf") else "—")

    # ── CALL 1: prose rationale (anonymised sources so filenames never appear) ──
    combined_anon = _merge_insights(extracted_reports, anonymize=True)

    prose_prompt = (
        _RATIONALE_PROMPT
        .replace("DEAL_NAME",         subject_cfg.get("deal_name",     "—"))
        .replace("PROPERTY_NAME",     subject_cfg.get("property_name", "—"))
        .replace("ADDRESS",           subject_cfg.get("address",       "—"))
        .replace("ASSET_CLASS",       subject_cfg.get("asset_class",   "—").title())
        .replace("ASSET_TYPE",        subject_cfg.get("asset_type",    "—"))
        .replace("LOCATION",          subject_cfg.get("location",      "—"))
        .replace("QUALITY",           subject_cfg.get("quality",       "—"))
        .replace("GFA_SF",            gfa_display)
        .replace("GFA_UNIT",          subject_cfg.get("gfa_unit",      "sf").upper())
        .replace("COUNTRY_NAME",      subject_cfg.get("country_name",  "—"))
        .replace("EXTRA_FIELDS",      "\n".join(extra) if extra else "")
        .replace("COMBINED_INSIGHTS", combined_anon)
        .replace("COMPARABLE_EVIDENCE_BLOCK", (
            "═══ COMPARABLE EVIDENCE (verified comparable transactions / rents) ═\n"
            f"{comp_summary.strip()}\n"
            "════════════════════════════════════════════════════════════════\n"
        ) if comp_summary.strip() else "")
        .replace("ANALYST_NOTES",     analyst_notes.strip() if analyst_notes else "(none)")
        .replace("REFINEMENT_BLOCK",  (
            "═══ REFINEMENT INSTRUCTIONS (highest priority) ══════════════════\n"
            "The analyst has reviewed the previous draft and requests these specific changes.\n"
            "Address every point below before writing:\n"
            f"{refinement_notes.strip()}\n"
            "════════════════════════════════════════════════════════════════"
        ) if refinement_notes.strip() else "")
    )

    print("  [generating] Investment rationale (3 sections) ...")
    t0_rationale = time.perf_counter()
    try:
        raw_rationale = _llm_chat(
            [
                {"role": "system", "content": _RATIONALE_SYSTEM},
                {"role": "user",   "content": prose_prompt},
            ],
            llm_cfg=llm_cfg,
            openai_key=openai_key,
            temperature=0.3,
        )
    except Exception as e:
        return f"**Error generating rationale:** {e}", []
    t1_rationale = time.perf_counter()
    print(f"  [timing]    Rationale prose  : {t1_rationale - t0_rationale:.1f}s")

    # Clean up any audit headings or stray separator the model appended
    rationale_body = _clean_rationale_body(raw_rationale)

    # ── CALL 2 (RAG): page-chunk index → claim extraction → vector source match ──
    # When an OpenAI key is available, the audit uses real vector search against
    # the original PDF pages — no LLM memory required.  Each claim is embedded,
    # the most similar page chunk is retrieved, and its source file + page number
    # become the citation.  Falls back to the original LLM audit if RAG is
    # unavailable (no key, no PDF paths, or embedding failure).
    t0_audit = time.perf_counter()
    audit_entries: list[dict] = []

    can_rag = bool(openai_key) and any(
        rpt.get("source_path") and Path(rpt["source_path"]).exists()
        for rpt in extracted_reports
    )

    if can_rag:
        print("  [rag audit] Building page-chunk embedding index ...")
        try:
            rag_index = _build_audit_rag_index(
                extracted_reports, openai_key, force_refresh=force_refresh
            )
        except Exception as e:
            print(f"  [warning] RAG index build failed: {e} — falling back to LLM audit")
            rag_index = []

        if rag_index:
            # Step 1: extract the list of specific claims from the prose
            print("  [rag audit] Extracting claims from rationale ...")
            claims: list[dict] = []
            try:
                raw_claims = _llm_chat(
                    [
                        {"role": "system", "content": _CLAIM_EXTRACT_SYSTEM},
                        {"role": "user",   "content":
                            _CLAIM_EXTRACT_PROMPT.replace(
                                "RATIONALE_BODY_PLACEHOLDER", rationale_body)},
                    ],
                    llm_cfg=llm_cfg,
                    openai_key=openai_key,
                    temperature=0.0,
                )
                claims = _parse_audit_json(raw_claims)
            except Exception as e:
                print(f"  [warning] Claim extraction failed: {e}")

            if claims:
                # Step 2: embed all claims in a single batch call
                print(f"  [rag audit] Embedding {len(claims)} claims ...")
                try:
                    claim_texts      = [c.get("claim", "") for c in claims]
                    claim_embeddings = _embed_batch(claim_texts, openai_key)

                    # Step 3: for each claim, retrieve the best matching page chunk
                    for claim_dict, claim_emb in zip(claims, claim_embeddings):
                        match = _rag_find_source(
                            claim_emb, rag_index, claim_dict.get("claim", "")
                        )
                        if match and match["score"] >= 0.3:
                            audit_entries.append({
                                "section_num":    claim_dict.get("section_num"),
                                "section_title":  claim_dict.get("section_title", ""),
                                "claim":          claim_dict.get("claim", ""),
                                "source_file":    match["source_file"],
                                "page_ref":       f"p.{match['page']}",
                                "supporting_text": match["text"],
                                "citation_type":  match["citation_type"],
                                "rag_score":      round(match["score"], 3),
                            })
                        else:
                            # Low similarity → claim likely originates from deal config
                            audit_entries.append({
                                "section_num":    claim_dict.get("section_num"),
                                "section_title":  claim_dict.get("section_title", ""),
                                "claim":          claim_dict.get("claim", ""),
                                "source_file":    "Deal Config",
                                "page_ref":       None,
                                "supporting_text": None,
                                "citation_type":  "Deal Config",
                                "rag_score":      round(match["score"], 3) if match else 0.0,
                            })
                except Exception as e:
                    print(f"  [warning] RAG claim matching failed: {e}")

        if not audit_entries:
            print("  [fallback] RAG audit empty — falling back to LLM audit ...")
            can_rag = False  # trigger LLM fallback below

    if not can_rag:
        # LLM-based audit (original approach — Ollama / no OpenAI key / RAG failure)
        combined_with_src = _merge_insights(extracted_reports, anonymize=False)
        valid_filenames   = "\n".join(
            f"  - {rpt['source_file']}"
            for rpt in extracted_reports if rpt.get("source_file")
        )
        audit_prompt = (
            _AUDIT_PROMPT_TEMPLATE
            .replace("RATIONALE_BODY_PLACEHOLDER",  rationale_body)
            .replace("SOURCE_DATA_PLACEHOLDER",     combined_with_src)
            .replace("VALID_FILENAMES_PLACEHOLDER", valid_filenames)
        )
        print("  [generating] Source audit JSON (LLM) ...")
        try:
            raw_audit = _llm_chat(
                [
                    {"role": "system", "content": _AUDIT_SYSTEM},
                    {"role": "user",   "content": audit_prompt},
                ],
                llm_cfg=llm_cfg,
                openai_key=openai_key,
                temperature=0.1,
            )
            audit_entries = _parse_audit_json(raw_audit)
            if not audit_entries:
                print("  [fallback] Audit JSON parse failed — trying inline markers")
                audit_entries = _fallback_citations_from_text(rationale_body)
        except Exception as e:
            print(f"  [warning] Audit call failed: {e} — continuing without audit")

    t1_audit = time.perf_counter()
    print(f"  [timing]    Source audit       : {t1_audit - t0_audit:.1f}s")
    print(f"  [timing]    Total generation   : {t1_audit - t0_rationale:.1f}s")

    # ── Build document header ──────────────────────────────────────────────────
    prop_name   = subject_cfg.get("property_name", subject_cfg.get("deal_name", "Property"))
    report_srcs = ", ".join(r["source_file"] for r in extracted_reports if r.get("source_file"))
    header = (
        f"# Investment Rationale — {prop_name}\n\n"
        f"**Address:** {subject_cfg.get('address', '—')}  \n"
        f"**Asset Class:** {subject_cfg.get('asset_class','—').title()}  \n"
        f"**Country:** {subject_cfg.get('country_name', '—')}  \n"
        f"**Generated:** {datetime.now().strftime('%d %b %Y')}  \n"
    )
    if report_srcs:
        header += f"**Market Reports Used:** {report_srcs}  \n"
    header += ("\n> _LLM-generated from the cited market reports — verify all figures "
               "against the source PDFs before use._\n")
    header += "\n---\n\n"

    rationale_md = header + _build_input_summary(extracted_reports) + rationale_body

    if audit_entries:
        rationale_md += (
            "\n\n---\n\n"
            f"> **Source Audit:** {len(audit_entries)} citations logged.  "
            "See `Source_Audit.xlsx` for verification details.\n"
        )

    return rationale_md, audit_entries


def _summarize_comp_excels(out_dir: Path) -> str:
    """Fallback comparable-evidence summary for standalone CLI runs (the Streamlit
    orchestrator passes a richer, analyst-visible summary via --comps-file).

    Reads the latest comp Excel per type from the deal's output dir and dumps its
    header + rows as compact text so the rationale can cite comparable pricing.
    Verify-only: it echoes the numbers already in the Excel, never derives new ones.
    Returns "" when no comp output exists."""
    specs = [
        ("Comparable Asset Sales", ["Transaction_Comparables", "Online_Comparables"]),
        ("Leasing Comparables",    ["Rent_Comps",              "Online_Rent_Comps"]),
        ("Comparable Land Sales",  ["Land_Sale_Comps",         "Online_Land_Comps"]),
    ]
    if not out_dir.exists():
        return ""
    try:
        import openpyxl
    except ImportError:
        return ""
    blocks = []
    for title, prefixes in specs:
        files = []
        for pfx in prefixes:
            files += [f for f in out_dir.glob(f"{pfx}*.xlsx")
                      if not f.name.startswith("~")]
        if not files:
            continue
        latest = max(files, key=lambda f: f.stat().st_mtime)
        try:
            ws = openpyxl.load_workbook(latest, data_only=True).active
        except Exception:
            continue
        rows = []
        for r in ws.iter_rows(values_only=True):
            vals = [("" if c is None else str(c).replace("\n", " ").strip()) for c in r]
            if any(vals):
                rows.append([v for v in vals if v.lower() != "source"] or vals)
        if len(rows) < 2:
            continue
        lines = [f"### {title} — {len(rows) - 1} data row(s)"]
        for r in rows[:20]:
            lines.append(" | ".join(v for v in r if v))
        blocks.append("\n".join(lines))
    return "\n\n".join(blocks)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT  (called from CLI or via run.py)
# ═══════════════════════════════════════════════════════════════════════════════

def run(
    config_path: str,
    report_paths: list[str] | None = None,
    force_refresh: bool = False,
    analyst_notes: str = "",
    refinement_notes: str = "",
    comp_summary: str = "",
) -> str:
    """
    Full pipeline orchestrator: extract insights → generate rationale → save outputs.

    Called by the Streamlit frontend (via subprocess) and directly from the CLI.
    The frontend always injects the sidebar model choice into the config before
    calling this function, so llm_cfg reflects the user's current model selection.

    Pipeline steps
    --------------
    1. Load config  — read the deal config JSON; build llm_cfg and resolve the
                      OpenAI API key (config field → env var fallback).

    2. Resolve reports  — if report_paths is empty, auto-discover all PDFs in
                          Input_files/market_reports/.

    3. Stage 1: Extract  — for each PDF, call extract_report_insights().
                           Already-extracted reports are served from the cache
                           in under a second.  Only new or force-refreshed reports
                           go through an LLM call.

    4. Stage 2: Generate  — call generate_rationale() which runs two LLM calls:
                            prose writing (anonymised sources) then audit JSON
                            (real filenames).  See generate_rationale() docstring
                            for the full explanation of why two calls are needed.

    5. Save outputs  — write Investment_Rationale.md and Source_Audit.xlsx into
                       the deal's output directory (output/<DealName>/).

    Returns the rationale markdown string.
    """
    import os
    cfg        = json.loads(Path(config_path).read_text(encoding="utf-8"))
    subj       = cfg["subject_property"]
    llm_cfg    = cfg.get("llm") or {}
    # Ensure a default Ollama block exists if the config has no llm section
    # (e.g. a freshly created config before the user has run any analysis).
    if not llm_cfg:
        llm_cfg = {"provider": "ollama",
                   "ollama": {"base_url": "http://localhost:11434", "model": "qwen2.5:3b"}}
    openai_key = (cfg.get("openai", {}).get("api_key")
                  or os.environ.get("OPENAI_API_KEY", ""))

    # Comparable evidence — the orchestrator passes a rich, analyst-visible summary
    # via --comps-file; when run standalone we fall back to reading the deal's comp
    # Excel outputs directly.  Verify-only: no numbers are derived here.
    _out_dir = ROOT / Path(cfg.get("output_file", "output/deal/deal.xlsx")).parent
    if not comp_summary:
        try:
            comp_summary = _summarize_comp_excels(_out_dir)
        except Exception as e:
            print(f"  [note] Comp-summary auto-read skipped: {e}")
            comp_summary = ""
    if comp_summary:
        print(f"  Comps       : injected into rationale evidence "
              f"({len(comp_summary)} chars)")

    # Resolve report paths — auto-discover if caller did not specify any
    if not report_paths:
        report_paths = sorted(
            str(p) for p in REPORTS_DIR.glob("*.pdf")
            if not p.name.startswith(".")
        )
        if not report_paths:
            print("[investment_rationale] No market reports found in "
                  f"{REPORTS_DIR}  — generating from property info only.")

    print(f"\n[Investment Rationale] {subj.get('deal_name', '')}")
    print(f"  Asset class : {subj.get('asset_class','-')}")
    print(f"  Country     : {subj.get('country_name','-')}")
    print(f"  Reports     : {len(report_paths)}")

    # ── Stage 1: Extract insights from each market report ────────────────────
    # extract_report_insights() caches results by (filename + size + mtime).
    # If the file hasn't changed since the last run the LLM is not called at all
    # and the cached JSON is returned in milliseconds.
    t_pipeline_start = time.perf_counter()
    extracted = []
    for rp in report_paths:
        t0 = time.perf_counter()
        result = extract_report_insights(rp, llm_cfg, openai_key, force_refresh=force_refresh)
        if result:
            extracted.append(result)
            cached = "(cached)" if (time.perf_counter() - t0) < 0.5 else f"{time.perf_counter() - t0:.1f}s"
            print(f"  [timing]    Extract {Path(rp).name}: {cached}")

    # ── Stage 2: Generate rationale prose + source audit JSON ────────────────
    # Two separate LLM calls — see generate_rationale() docstring for why.
    rationale, audit_entries = generate_rationale(
        extracted, subj, llm_cfg, openai_key,
        analyst_notes=analyst_notes,
        refinement_notes=refinement_notes,
        force_refresh=force_refresh,
        comp_summary=comp_summary,
    )

    # ── Save outputs ──────────────────────────────────────────────────────────
    out_dir = ROOT / Path(cfg.get("output_file", "output/deal/deal.xlsx")).parent
    out_dir.mkdir(parents=True, exist_ok=True)

    # Investment_Rationale.md — the prose shown in the dashboard and downloaded
    out_file = out_dir / "Investment_Rationale.md"
    out_file.write_text(rationale, encoding="utf-8")
    print(f"  [OK] Saved -> {out_file.relative_to(ROOT)}")

    # Source_Audit.xlsx — citation table for human verification (never shown inline)
    if audit_entries:
        audit_xlsx = out_dir / "Source_Audit.xlsx"
        try:
            _write_source_audit_excel(audit_entries, extracted, audit_xlsx, subj)
            print(f"  [OK] Saved -> {audit_xlsx.relative_to(ROOT)}")
        except Exception as e:
            print(f"  [warning] Could not write source audit Excel: {e}")
    else:
        print("  [note] No citations captured -- source audit Excel not generated.")

    print(f"\n  [timer] Total pipeline time: {time.perf_counter() - t_pipeline_start:.1f}s")
    return rationale


# ═══════════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    ap = argparse.ArgumentParser(
        description="Generate investment rationale from market reports + deal config"
    )
    ap.add_argument("--config",     required=True,
                    help="Path to deal config JSON")
    ap.add_argument("--reports",    nargs="*", default=[],
                    help="PDF/txt report paths.  Defaults to all PDFs in "
                         "Input_files/market_reports/")
    ap.add_argument("--refresh",    action="store_true",
                    help="Re-extract all reports even if cached")
    ap.add_argument("--notes-file", default="",
                    help="Path to a plain-text file with analyst notes")
    ap.add_argument("--refinement-file", default="",
                    help="Path to a plain-text file with refinement/feedback instructions")
    ap.add_argument("--comps-file", default="",
                    help="Path to a text/markdown file summarising the comparable "
                         "transactions/rents to inject as evidence (from the orchestrator). "
                         "If omitted, comps are auto-read from the deal's output dir.")
    args = ap.parse_args()

    notes = ""
    if args.notes_file:
        try:
            notes = Path(args.notes_file).read_text(encoding="utf-8")
        except Exception as e:
            print(f"[warning] Could not read notes file: {e}")

    refinement = ""
    if args.refinement_file:
        try:
            refinement = Path(args.refinement_file).read_text(encoding="utf-8")
        except Exception as e:
            print(f"[warning] Could not read refinement file: {e}")

    comps = ""
    if args.comps_file:
        try:
            comps = Path(args.comps_file).read_text(encoding="utf-8")
        except Exception as e:
            print(f"[warning] Could not read comps file: {e}")

    result = run(
        config_path=args.config,
        report_paths=args.reports or None,
        force_refresh=args.refresh,
        analyst_notes=notes,
        refinement_notes=refinement,
        comp_summary=comps,
    )

    print("\n" + "=" * 70)
    # Print rationale safely — replace any remaining unencodable chars
    safe_result = result.encode(sys.stdout.encoding or "utf-8", errors="replace").decode(
        sys.stdout.encoding or "utf-8"
    )
    print(safe_result)
