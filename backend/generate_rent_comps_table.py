#!/usr/bin/env python3
"""
generate_rent_comps_table.py
============================
Deal-agnostic Rental Comparables table generator for commercial real estate
underwriting. Reads a curated rent comps Excel and produces a formatted output.

Usage
-----
    python3 generate_rent_comps_table.py                          # uses configs/deal_config.json
    python3 generate_rent_comps_table.py --config configs/my_deal.json

Pipeline  (4 stages)
--------------------
  1  PARSE      Read input Excel; auto-detect columns via Ollama (fuzzy fallback).
  2  CLASSIFY   Ollama assigns Location, Quality, Asset Type, relevance score.
                Falls back to keyword rules if Ollama is down.
  3  CALCULATE  Effective rent = asking rent adjusted for rent-free amortisation.
  4  RENDER     Formatted Excel with navy headers, alternating fills, Sources tab.

Output Schema  (13 visible columns)
-------------------------------------
  Property | Map Marker | Lease Date | Land Zoning | NLA
  Asking Rent | Eff. Rent | Lease Term (Yrs) | Rent-Free (Mths)
  Tenant | Location | Quality | Asset Type
"""

import argparse
import json
import re
import sys as _sys
import urllib.request
from pathlib import Path

# ── Corporate proxy TLS fix (trust OS cert store; no-op without truststore) ────
_sys.path.insert(0, str(Path(__file__).parent))
from tools import corp_ssl  # noqa: F401,E402  — must import before any HTTPS call

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — OUTPUT SCHEMA
# ═══════════════════════════════════════════════════════════════════════════════

RENT_SCHEMA_BASE = [
    ("Type",                                    "type",           "str",   "@",              10),
    ("Source",                                  "source",         "str",   "@",              10),
    ("Property",                                "property",       "str",   "@",              36),
    ("Map\nMarker",                             "map_marker",     "str",   "@",               8),
    ("Date of\nLease Start",                    "lease_date",     "str",   "@",              14),
    ("Leased GLA\n(SF)",                        "nla_sf",         "int",   "#,##0",          14),
    ("Lease\nTenure\n(Yrs)",                    "lease_term_yrs", "float", "0.0",            10),
    ("Gross Face Rents\n(SGD psf pm)",          "asking_rent",    "float", '"S$"#,##0.00',   20),
    ("Effective Rents\n(SGD psf pm)",           "eff_rent",       "float", '"S$"#,##0.00',   20),
    ("Location",                                "location",       "str",   "@",              24),
    ("Quality",                                 "quality",        "str",   "@",              28),
    ("Tenant",                                  "tenant",         "str",   "@",              24),
    ("Type of Lease\nArea / Comments",          "lease_type",     "str",   "@",              32),
]


def get_rent_schema(subject_cfg: dict = None) -> list:
    """Return RENT_SCHEMA adapted for this deal's currency and area unit."""
    if subject_cfg is None:
        return list(RENT_SCHEMA_BASE)

    currency = subject_cfg.get("currency", "SGD")
    gfa_unit = subject_cfg.get("gfa_unit", "sf").lower()
    sym      = subject_cfg.get("currency_symbol", currency)
    area_lbl = "psm" if gfa_unit == "sqm" else "psf"
    nla_lbl  = "SQM" if gfa_unit == "sqm" else "SF"
    period   = subject_cfg.get("rent_period", "mth")   # "mth" or "yr"

    # Match by field key (NOT index) so adding/reordering columns can't clobber
    # the wrong one — the unit/currency labels always land on the right column.
    schema = list(RENT_SCHEMA_BASE)
    for i, entry in enumerate(schema):
        key = entry[1]
        if key == "nla_sf":
            schema[i] = (f"Leased GLA\n({nla_lbl})",
                         "nla_sf", "int", "#,##0", entry[4])
        elif key == "asking_rent":
            schema[i] = (f"Gross Face Rents\n({sym} {area_lbl} pm)",
                         "asking_rent", "float", f'"{sym}"#,##0.00', entry[4])
        elif key == "eff_rent":
            schema[i] = (f"Effective Rents\n({sym} {area_lbl} pm)",
                         "eff_rent",    "float", f'"{sym}"#,##0.00', entry[4])
    return schema


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — STYLE HELPERS  (shared with sales comps)
# ═══════════════════════════════════════════════════════════════════════════════

_NAVY   = "1A3A5C"
_TEAL   = "1A5C4A"
_WHITE  = "FFFFFF"
_ALT    = "EBF5FB"
_SUBJ   = "FEF9E7"
_RED    = "C00000"   # subject property star marker colour
_AVGBG  = "D6DCE4"   # blue-grey for average row

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=_WHITE, size=10, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)

def _border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — COLUMN DETECTION (Ollama + fuzzy)
# ═══════════════════════════════════════════════════════════════════════════════

RENT_FIELDS = {
    "raw_description": "Property name and full address",
    "nla_sf":          "Net lettable area or GFA in sqft or sqm",
    "asking_rent":     "Asking / gross rent per sqft or sqm per month",
    "eff_rent":        "Effective / net rent after incentives per sqft or sqm per month",
    "lease_term_yrs":  "Lease term or tenure in years",
    "rent_free_mths":  "Rent-free period in months",
    "tenant":          "Tenant name or occupier",
    "lease_date":      "Date the lease was signed or commenced",
    "land_zoning":     "URA or local zoning class",
    "quarter":         "Quarter of lease: Q1 Q2 Q3 Q4",
    "year":            "Calendar year of lease",
}

_RENT_SYNONYMS = {
    "raw_description": ["property", "building", "asset", "address", "description"],
    "nla_sf":          ["nla", "net lettable", "gfa", "floor area", "area"],
    "asking_rent":     ["asking rent", "gross rent", "headline rent", "passing rent"],
    "eff_rent":        ["effective rent", "net rent", "net effective", "eff rent"],
    "lease_term_yrs":  ["lease term", "tenure", "term (yrs)", "lease period"],
    "rent_free_mths":  ["rent free", "rent-free", "free period", "incentive"],
    "tenant":          ["tenant", "occupier", "lessee", "occupant"],
    "lease_date":      ["lease date", "commencement", "signed", "quarter", "year"],
    "land_zoning":     ["zoning", "land use", "ura"],
    "quarter":         ["quarter", "q1", "q2", "q3", "q4"],
    "year":            ["year", "yr"],
}


def _fuzzy_detect(headers: list) -> dict:
    mapping = {}
    for col_idx, raw_hdr in enumerate(headers):
        h = str(raw_hdr or "").lower()
        for field, syns in _RENT_SYNONYMS.items():
            if any(s in h for s in syns) and field not in mapping.values():
                mapping[col_idx] = field
                break
    return mapping


def _ollama_post(base_url: str, model: str, messages: list, timeout: int = 60) -> str:
    payload = json.dumps({
        "model": model, "messages": messages,
        "stream": False, "format": "json", "options": {"temperature": 0},
    }).encode()
    req = urllib.request.Request(
        f"{base_url}/api/chat", data=payload,
        headers={"Content-Type": "application/json"}, method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read())["message"]["content"]


def auto_detect_columns(headers: list, llm_cfg: dict) -> dict:
    provider = llm_cfg.get("provider", "ollama")
    if provider == "ollama":
        try:
            ocfg = llm_cfg.get("ollama", {})
            base_url = ocfg.get("base_url", "http://localhost:11434")
            model    = ocfg.get("model",    "qwen2.5:3b")
            system = (
                "You map spreadsheet column headers to these internal field names. "
                "Return ONLY a JSON object {col_index_str: field_name}.\n\n"
                "Fields:\n" +
                "\n".join(f"  {k}: {v}" for k, v in RENT_FIELDS.items())
            )
            user = "Headers (0-indexed): " + json.dumps(
                {str(i): h for i, h in enumerate(headers)}
            )
            raw = _ollama_post(base_url, model, [
                {"role": "system", "content": system},
                {"role": "user",   "content": user},
            ])
            raw = re.sub(r"^```[a-z]*\n?", "", raw.strip())
            raw = re.sub(r"\n?```$", "", raw)
            parsed = json.loads(raw)
            return {int(k): v for k, v in parsed.items()}
        except Exception as e:
            print(f"  [LLM] Ollama column detection failed ({e}), using fuzzy fallback")
    return _fuzzy_detect(headers)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — PARSE INPUT EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def parse_rent_comps(input_file: str, llm_cfg: dict, max_comps: int = 10) -> list:
    wb = openpyxl.load_workbook(input_file, data_only=True)

    # Auto-select the sheet with the most non-empty data rows
    def _data_rows(sheet_name):
        return sum(1 for r in wb[sheet_name].iter_rows(values_only=True)
                   if any(c not in (None, "") for c in r))
    best_sheet = max(wb.sheetnames, key=_data_rows)
    if best_sheet != wb.active.title:
        print(f"  [Sheet] Using '{best_sheet}' (most data rows)")
    ws   = wb[best_sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Find header row (first row with 3+ non-empty cells)
    hdr_row = 0
    for i, row in enumerate(rows):
        if sum(1 for c in row if c not in (None, "")) >= 3:
            hdr_row = i
            break

    headers = [str(c or "").strip() for c in rows[hdr_row]]
    col_map = auto_detect_columns(headers, llm_cfg)
    print(f"  Column map: { {v: headers[k] for k, v in col_map.items()} }")

    comps = []
    for row in rows[hdr_row + 1:]:
        if all(c in (None, "") for c in row):
            continue
        def _get(field):
            for idx, f in col_map.items():
                if f == field and idx < len(row):
                    return row[idx]
            return None

        desc        = str(_get("raw_description") or "").strip()
        if not desc:
            continue

        # Parse lease date from quarter + year or direct field
        q   = str(_get("quarter") or "").strip().upper()
        yr  = str(_get("year")    or "").strip()
        dt_raw = str(_get("lease_date") or "").strip()
        if q and yr:
            lease_date = f"{q} {yr}"
        elif dt_raw:
            lease_date = dt_raw
        else:
            lease_date = ""

        def _num(field):
            v = _get(field)
            try:
                return float(str(v).replace(",", "").replace("$", "").strip())
            except Exception:
                return None

        comps.append({
            "raw_description": desc,
            "lease_date":      lease_date,
            "nla_sf":          _num("nla_sf"),
            "asking_rent":     _num("asking_rent"),
            "eff_rent":        _num("eff_rent"),
            "lease_term_yrs":  _num("lease_term_yrs"),   # kept internally for eff rent calc
            "rent_free_mths":  _num("rent_free_mths"),   # kept internally for eff rent calc
            "lease_type":      str(_get("lease_type") or "").strip(),
        })

    return comps[:max_comps]


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — LLM CLASSIFICATION
# ═══════════════════════════════════════════════════════════════════════════════

def classify_rent_comps(comps: list, subject_cfg: dict, max_comps: int,
                        llm_cfg: dict) -> list:
    if not comps:
        return []
    asset_class  = subject_cfg.get("asset_class", "office")
    country_name = subject_cfg.get("country_name", "Singapore")
    provider     = llm_cfg.get("provider", "ollama")

    entries = "\n".join(
        f'{i+1}. {c["raw_description"]}' for i, c in enumerate(comps)
    )
    system = (
        f"You classify {country_name} {asset_class} rental comparable properties. "
        "Return ONLY a JSON array, one object per comp.\n\n"
        "Each object must have:\n"
        "  index        : 1-based integer (matches input)\n"
        "  property     : clean property name (no floor or unit info)\n"
        "  address      : geocodable address — street address if present, or specific "
        "building name if no street address; null if the text is a generic description "
        "with no identifiable address or building name\n"
        "  relevance    : integer 0–10 (10 = most comparable to subject)\n"
        "  map_marker   : leave blank (assigned later)\n\n"
        "Do NOT include location, quality, or asset_type — those come from the source data."
    )
    user = f"Classify these {len(comps)} rental comparables:\n{entries}"

    result = []
    try:
        if provider == "ollama":
            ocfg     = llm_cfg.get("ollama", {})
            base_url = ocfg.get("base_url", "http://localhost:11434")
            model    = ocfg.get("model",    "qwen2.5:3b")
            raw = _ollama_post(base_url, model, [
                {"role": "system", "content": system},
                {"role": "user",   "content": user},
            ], timeout=120)
        else:
            raise ValueError(f"Unsupported provider: {provider}")

        raw = re.sub(r"^```[a-z]*\n?", "", raw.strip())
        raw = re.sub(r"\n?```$", "", raw)
        parsed = json.loads(raw)
        if isinstance(parsed, dict):
            parsed = parsed.get("comparables") or next(
                (v for v in parsed.values() if isinstance(v, list)), []
            )

        for item in parsed:
            idx = int(item.get("index", 0)) - 1
            if 0 <= idx < len(comps):
                merged = {**comps[idx], **item, "type": "Comparable"}
                result.append(merged)

        result.sort(key=lambda x: -int(x.get("relevance", 0)))
        result = result[:max_comps]
        for i, r in enumerate(result):
            r["map_marker"] = str(i + 1)

    except Exception as e:
        print(f"  [LLM] Classification failed ({e}), using raw order")
        for i, c in enumerate(comps[:max_comps]):
            c["map_marker"] = str(i + 1)
            c["type"]       = "Comparable"
            c.setdefault("property", c["raw_description"].split("\n")[0][:50])
            c.setdefault("address",  "")
        result = comps[:max_comps]

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — CALCULATE EFFECTIVE RENT
# ═══════════════════════════════════════════════════════════════════════════════

def compute_eff_rent(comps: list) -> list:
    """
    If eff_rent is missing, estimate from asking_rent and rent_free_mths:
        eff_rent = asking_rent * (lease_term_mths - rent_free_mths) / lease_term_mths
    """
    for c in comps:
        if c.get("eff_rent") is not None:
            continue
        asking = c.get("asking_rent")
        rf     = float(c.get("rent_free_mths") or 0)
        lt_yrs = float(c.get("lease_term_yrs") or 0)
        if asking and lt_yrs > 0:
            lt_mths = lt_yrs * 12
            c["eff_rent"] = round(asking * (lt_mths - rf) / lt_mths, 2)
    return comps


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 7 — EXCEL RENDERING
# ═══════════════════════════════════════════════════════════════════════════════

def _write_header(ws, row: int, schema: list, subj_yrs: int = 0):
    for col_idx, (hdr, _, _, _, width) in enumerate(schema, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value       = hdr.format(subj_yrs=subj_yrs)
        cell.fill        = _fill(_NAVY)
        cell.font        = _font(bold=True)
        cell.border      = _border()
        cell.alignment   = _align()
        col_letter       = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width


def _write_row(ws, row: int, data: dict, schema: list,
               alt: bool = False, bold: bool = False, is_subject: bool = False):
    bg = _SUBJ if is_subject else (_ALT if alt else _WHITE)
    fc = "000000"
    for col_idx, (_, key, dtype, fmt, _) in enumerate(schema, 1):
        cell = ws.cell(row=row, column=col_idx)
        val  = data.get(key)
        if val is not None and val != "":
            if dtype in ("int", "float"):
                try:
                    val = int(float(val)) if dtype == "int" else float(val)
                except (TypeError, ValueError):
                    pass
        cell.value       = val
        cell.number_format = fmt
        cell.fill        = _fill(bg)
        # Map marker column: subject star is red and larger; comp numbers stay dark
        if key == "map_marker" and is_subject:
            cell.font = _font(bold=True, color=_RED, size=12)
        else:
            cell.font = _font(bold=bold, color=fc)
        cell.border      = _border()
        cell.alignment   = _align(
            h="right" if dtype in ("int", "float") else "center"
        )


def _avg_row(ws, row: int, first_r: int, last_r: int, schema: list):
    """Write Average row — Gross Face Rents and Effective Rents."""
    avg_keys = {"asking_rent", "eff_rent"}
    for col, (_, key, dtype, fmt, _) in enumerate(schema, 1):
        c = ws.cell(row=row, column=col)
        c.fill   = _fill(_AVGBG)
        c.border = _border()
        if col == 1:
            c.value     = "Average"
            c.font      = _font(bold=True, color="000000")
            c.alignment = _align(h="left")
        elif key in avg_keys:
            col_ltr         = get_column_letter(col)
            c.value         = (f'=IFERROR(AVERAGEIF({col_ltr}{first_r}:'
                               f'{col_ltr}{last_r},">0"),"—")')
            c.number_format = fmt
            c.font          = _font(bold=True, color="000000")
            c.alignment     = _align(h="right")
        else:
            c.value     = ""
            c.font      = _font(color="000000")
            c.alignment = _align()
    ws.row_dimensions[row].height = 32


def _subject_row(subject_cfg: dict) -> dict:
    prop_name = subject_cfg["property_name"]
    address   = subject_cfg.get("address", "")
    prop_cell = (f"{prop_name}\n{address}"
                 if address and address.strip() != prop_name.strip()
                 else prop_name)
    return {
        "type":           "Subject",
        "source":         "",
        "property":       prop_cell,
        "map_marker":     "★",
        "lease_date":     subject_cfg.get("sale_date", ""),
        "nla_sf":         subject_cfg.get("gfa_sf"),
        "lease_term_yrs": None,
        "asking_rent":    subject_cfg.get("asking_rent"),
        "eff_rent":       subject_cfg.get("eff_rent"),
        "location":       subject_cfg.get("location", ""),
        "quality":        subject_cfg.get("quality", ""),
        "lease_type":     subject_cfg.get("asset_type", ""),
    }


def build_workbook(subject_cfg: dict, comp_rows: list,
                   output_path: str, schema: list = None):
    if schema is None:
        schema = get_rent_schema(subject_cfg)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rent Comparables"
    ws.freeze_panes = "A3"
    # Hide the Type column (col A) — visible when needed, hidden by default
    ws.column_dimensions["A"].hidden = True

    current_row = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    deal_name = subject_cfg.get("deal_name", subject_cfg["property_name"])
    title_cell = ws.cell(row=current_row, column=1,
                         value=f"Rental Comparables — {deal_name}")
    title_cell.font      = _font(bold=True, color="000000", size=12)
    title_cell.fill      = _fill(_WHITE)
    title_cell.alignment = _align(h="left")
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=len(schema))
    current_row += 1

    # ── Table 1: Subject property ─────────────────────────────────────────────
    _write_header(ws, current_row, schema)
    current_row += 1
    _write_row(ws, current_row, _subject_row(subject_cfg), schema,
               is_subject=True, bold=True)
    current_row += 2

    # ── Table 2: Comparable leases ────────────────────────────────────────────
    _write_header(ws, current_row, schema)
    current_row += 1
    comp_start = current_row
    for i, comp in enumerate(comp_rows):
        _write_row(ws, current_row, comp, schema, alt=(i % 2 == 1))
        current_row += 1

    # ── Average row ───────────────────────────────────────────────────────────
    if comp_rows:
        _avg_row(ws, current_row, comp_start, current_row - 1, schema)
        current_row += 1

    # ── Notes ─────────────────────────────────────────────────────────────────
    current_row += 1
    currency = subject_cfg.get("currency", "SGD")
    gfa_unit = subject_cfg.get("gfa_unit", "sf")
    period   = subject_cfg.get("rent_period", "mth")
    sym  = subject_cfg.get("currency_symbol", currency)
    area = "psm" if gfa_unit == "sqm" else "psf"
    note = (f"Notes: Rents in {sym} per {area} per {period}. "
            "Effective Rent estimated from rent-free period where direct figure unavailable. "
            "Source: Online search / broker data.")
    cell = ws.cell(row=current_row, column=1, value=note)
    cell.font      = Font(italic=True, size=9, color="666666", name="Calibri")
    cell.alignment = _align(h="left", wrap=False)
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=len(schema))

    ws.row_dimensions[1].height = 18
    for r in range(2, current_row):
        ws.row_dimensions[r].height = 32

    wb.save(output_path)
    print(f"  Saved → {output_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 8 — ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def run(config_path: str = "configs/deal_config.json"):
    with open(config_path, encoding="utf-8") as f:
        cfg = json.load(f)

    subject_cfg = cfg["subject_property"]
    input_file  = cfg.get("rent_input_file") or cfg.get("input_file")
    output_file = cfg.get("rent_output_file") or cfg.get("output_file",
                  f"output/Rent_Comps_{subject_cfg.get('deal_name','Deal').replace(' ','_')}.xlsx")
    params      = cfg.get("parameters", {})
    max_comps   = params.get("max_comps", 10)
    llm_cfg     = cfg.get("llm", {"provider": "ollama",
                                   "ollama": {"base_url": "http://localhost:11434",
                                              "model": "qwen2.5:3b"}})

    if not input_file:
        print("No rent_input_file specified in config. "
              "Use search_online_rent_comps.py for AI-sourced rent comps.")
        return

    Path(output_file).parent.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*60}\n  Rent Comps : {subject_cfg['deal_name']}\n{'='*60}")

    print(f"\n[1/4] PARSE   {input_file}")
    comps = parse_rent_comps(input_file, llm_cfg, max_comps)
    print(f"      → {len(comps)} comps parsed")

    print(f"\n[2/4] CLASSIFY  ({llm_cfg.get('provider','ollama')})")
    classified = classify_rent_comps(comps, subject_cfg, max_comps, llm_cfg)
    print(f"      → {len(classified)} records classified")

    print(f"\n[3/4] CALCULATE  (eff. rent from rent-free)")
    classified = compute_eff_rent(classified)

    print(f"\n[4/4] RENDER   {output_file}")
    schema = get_rent_schema(subject_cfg)
    build_workbook(subject_cfg, classified, output_file, schema)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Rent Comparables table")
    parser.add_argument("--config", default="configs/deal_config.json",
                        help="Path to deal config JSON")
    args = parser.parse_args()
    run(args.config)
